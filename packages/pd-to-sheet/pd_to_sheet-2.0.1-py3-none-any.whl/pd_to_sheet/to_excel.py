from pathlib import Path

import pandas
from openpyxl.reader.excel import load_workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

widths = [
    (126, 1), (159, 0), (687, 1), (710, 0), (711, 1), (727, 0), (733, 1), (879, 0), (1154, 1), (1161, 0), (4347, 1),
    (4447, 2), (7467, 1), (7521, 0), (8369, 1), (8426, 0), (9000, 1), (9002, 2), (11021, 1), (12350, 2), (12351, 1),
    (12438, 2), (12442, 0), (19893, 2), (19967, 1), (55203, 2), (63743, 1), (64106, 2), (65039, 1), (65059, 0),
    (65131, 2), (65279, 1), (65376, 2), (65500, 1), (65510, 2), (120831, 1), (262141, 2), (1114109, 1),
]


def get_width(character):
    global widths
    if character == 0xe or character == 0xf:
        return 0
    for num, wid in widths:
        if character <= num:
            return wid
    return 1


# 计算一个字符串占多少个半角字符宽带，用于在表格里设置列宽
def get_string_width(string):
    result_width = 0
    for character in string:
        character_width = get_width(ord(character))
        result_width += character_width
    return result_width


# 将多个 DataFrame 保存到同一个 Excel 文件的不同 Sheet
def save_df_to_sheet(excel_writer, sheet_name, pandas_df, wrap_columns=None):
    print(f"正在保存 {sheet_name}")
    pandas_df.reset_index(drop=True, inplace=True)
    # 检查中是否有列“序号"如果有则删除后重新重新设置，否则插入一列到最前面并从1赋值
    if '序号' in pandas_df.columns:
        pandas_df.drop(columns=['序号'], inplace=True)
        pandas_df.insert(loc=0, column='序号', value=pandas_df.index + 1)
    else:
        pandas_df.insert(loc=0, column='序号', value=pandas_df.index + 1)
    # 将 DataFrame 写入 Excel 文件中的指定 sheet
    pandas_df.to_excel(excel_writer, sheet_name=sheet_name, index=False, engine='xlsxwriter')

    # 获取当前 sheet 的 workbook 和 worksheet 对象
    workbook = excel_writer.book
    worksheet = excel_writer.sheets[sheet_name]
    # 定义表头样式
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',  # 水平和垂直居中
        'valign': 'vcenter',
        'fg_color': '#D7E4BC'
    })
    # 定义边框样式
    border_format = workbook.add_format({
        'border': 1,  # 外边框为蓝色实线
        'border_color': 'blue',
        'text_wrap': True,
        'align': 'center',  # 水平和垂直居中
        'valign': 'vcenter'
    })

    cell_format = workbook.add_format({
        'text_wrap': True,
        'align': 'center',  # 水平和垂直居中
        'valign': 'vcenter'
    })

    # 获取 DataFrame 的行数和列数
    num_rows, num_cols = pandas_df.shape
    # 设置每一行的行高
    for row in range(num_rows + 1):  # +1 是为了包括表头
        worksheet.set_row(row, height=25, cell_format=cell_format)
    # 如果有需要换行的列，则设置每一行的高度
    if wrap_columns:
        for row_idx in range(1, num_rows + 1):
            line_count = 0
            for col in wrap_columns:
                cell_value = pandas_df[col].iloc[row_idx - 1]
                if cell_value:
                    current_line_count = str(cell_value).count('\n')
                    line_count = max(line_count, current_line_count)
            if line_count > 0:
                worksheet.set_row(row_idx, height=25 + line_count * 13, cell_format=cell_format)

    # 为表头行的每个单元格设置样式
    for col_num, value in enumerate(pandas_df.columns.values):
        worksheet.write(0, col_num, value, header_format)
    # 设置整个表格的边框样式
    worksheet.conditional_format(first_row=0, first_col=0, last_row=num_rows, last_col=num_cols - 1, options={'type': 'no_blanks', 'format': border_format})
    worksheet.conditional_format(first_row=0, first_col=0, last_row=num_rows, last_col=num_cols - 1, options={'type': 'blanks', 'format': border_format})

    for col_idx, col_name in enumerate(pandas_df.columns):

        # 计算每行的列宽，取最宽的一个值，包括表头
        max_width = get_string_width(col_name)
        pandas_df[col_name] = pandas_df[col_name].apply(lambda x: "、".join(x) if isinstance(x, list) else x)
        for value in pandas_df[col_name].unique().tolist():
            line_count = f"{value}".count('\n')
            if line_count > 0:
                # print(f"{col_name} 有换行")
                # print(f"正在设置列宽 {col_name}")
                # print(f"{col_name} 的表头列宽为 {max_width}")
                cell_value_list = f"{value}".split('\n')
                for cell_value in cell_value_list:
                    cell_value = f"{cell_value}发"  # 添加一个字符的宽度
                    cell_width = get_string_width(cell_value)
                    max_width = max(max_width, cell_width)
                # print(f"{col_name} 换行最终列宽为 {max_width}")
            else:
                cell_value = f"{value}发"  # 添加一个字符的宽度
                cell_width = get_string_width(cell_value)
                max_width = max(max_width, cell_width)

        # 设置列宽和对齐
        worksheet.set_column(col_idx, col_idx, max_width)
    # 冻结首行
    worksheet.freeze_panes(1, 0)
    # 设置纸张方向为横向
    worksheet.set_landscape()
    # 设置顶端标题行（例如第一行）
    worksheet.repeat_rows(0)
    # 设置页脚页码，使用 HTML 标签 <font> 设置字体和字号
    worksheet.set_footer('第 &P / &N 页')
    # 设置打印时水平居中
    worksheet.center_horizontally()
    # 缩放将所有列打印在一页
    worksheet.fit_to_pages(1, 0)  # 宽度缩放到1页，高度不限制
    # 设置页边距（单位：英寸,厘米转换为英寸）
    worksheet.set_margins(left=1 / 2.54, right=1 / 2.54, top=1.5 / 2.54, bottom=1 / 2.54)
    # 设置纸张为 A4
    worksheet.set_paper(9)  # A4 的代码为 9

    # 水平居中对齐打印页面
    worksheet.center_horizontally()
    # worksheet.set_column('A:DC', 15, formater)
    return excel_writer


def set_column_width_and_merge_header(file_path, output_path, sheet_params):
    # retain是统计汇总描述所在列号，默认为5
    # 加载已有的工作簿
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        sheet_name = ws.title
        title = None
        description = None
        if sheet_name in sheet_params:
            custom_param = sheet_params[sheet_name]
            title = custom_param['title']
            description = custom_param['description']

        # 设置页脚内容
        footer = f"第 &P 页，共 &N 页"
        ws.oddFooter.center.text = footer
        ws.oddFooter.center.size = 11  # 设置字体大小
        ws.oddFooter.center.font = "Arial"  # 设置字体

        # 设置单元格内容自动换行，水平居中，垂直居中
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # 设置第二行（索引为1，因为索引从0开始）的字体加粗
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # 获取当前工作表的总行数
        total_rows = ws.max_row

        for cell in ws[total_rows]:
            cell.font = Font(bold=True)
            # 找到包含“说明”的单元格后，合并从该行的第一列到此单元格
            if "说明" in str(cell.value):
                ws.merge_cells(start_row=total_rows + 1, start_column=1, end_row=total_rows + 1, end_column=ws.max_column)
                # ws.merge_cells(start_row=start_cell.row, start_column=start_cell.column, end_row=cell.row, end_column=cell.column)
                # 设置合并后的单元格的值为传入的参数
                ws.cell(row=total_rows, column=1).value = description
                break
        ws.row_dimensions[total_rows + 1].height = 30
        # 获取要保留的单元格的值

        # retain_value = ws.cell(row=total_rows, column=retain).value
        # 合并单元格
        # ws.merge_cells(start_row=total_rows + 1, start_column=1, end_row=total_rows + 1, end_column=retain)
        # 设置合并后的单元格的值为要保留的单元格的值
        # ws.cell(row=total_rows, column=1).value = retain_value
        # ws.cell(row=total_rows, column=1).value = "未被删除的文件记录合计"
        # 在顶部添加一行
        ws.insert_rows(1)
        ws.row_dimensions[1].height = 30
        # 设置要合并的单元格范围（新添加的行），这里假设合并整行的表头
        merge_start_row = 1
        merge_end_row = 1
        merge_start_col = 1
        merge_end_col = ws.max_column

        # 设置合并单元格的内容为sheet名
        if title:
            # 合并单元格
            ws.merge_cells(start_row=merge_start_row, start_column=merge_start_col, end_row=merge_end_row, end_column=merge_end_col)

            sheet_name = ws.title
            ws.cell(row=merge_start_row, column=merge_start_col).value = title

            # 设置字体为方正小标宋24号字体，注意openpyxl可能不支持所有中文字体名称
            # 如果方正小标宋字体不可用，可以尝试使用其他常见的中文字体，如'宋体'或'微软雅黑'
            font = Font(size=18, name='方正小标宋简体')
            ws.cell(row=merge_start_row, column=merge_start_col).font = font
            # 设置对齐方式
            alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=merge_start_row, column=merge_start_col).alignment = alignment

            # 设置打印表头，这里假设只打印合并后的那一行作为表头
            ws.print_title_rows = '1:2'
            # 冻结前两行
            # 设置 freeze_panes 的值为 'A3'，意味着从第三行开始显示，前两行固定不动
            ws.freeze_panes = 'A3'
        else:
            # 设置打印表头，这里假设只打印合并后的那一行作为表头
            ws.print_title_rows = '1:1'
            # 冻结前两行
            # 设置 freeze_panes 的值为 'A3'，意味着从第三行开始显示，前两行固定不动
            ws.freeze_panes = 'A2'

        # 设置打印页边距，毫米
        left = 15
        right = 15
        top = 15
        bottom = 10
        header = 8
        footer = 8

        ws.page_margins.left = left / 25.4  # 左边距
        ws.page_margins.right = right / 25.4  # 右边距
        ws.page_margins.top = top / 25.4  # 上边距
        ws.page_margins.bottom = bottom / 25.4
        ws.page_margins.header = header / 25.4
        ws.page_margins.footer = footer / 25.4

        # 设置所有列缩放到一页宽度
        # ws.page_setup.fitToWidth = True
        # ws.page_setup.fitToHeight = 0  # 不限制高度
        # ws.page_setuppaperSize = "9"  # 纸张类型A4
        # ws.page_setup.orientation = "portrait"
        # 获取所有列的宽度求和 字符像素宽度 = （字体宽度 * 字符个数 + 边距）*0.264583
        # 96 PPI：1 像素 ≈ 0.264583 毫米
        # 120 PPI：1 像素 ≈ 0.211667 毫米
        # 144 PPI：1 像素 ≈ 0.176389 毫米
        # 300 PPI：1 像素 ≈ 0.084667 毫米

        column_width_count = 0
        word_width = 7.4  # 11号宋体标准宽度为8像素
        for col in ws.column_dimensions:
            column_width = ws.column_dimensions[col].width
            column_width_count += (column_width * word_width) * 0.264583
        print(f"{sheet_name}的列宽度：{column_width_count}")
        # 其他打印设置（可选）
        if ws.max_column > 7:  # 如果超过6列则将纸张设置为横向
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 横向打印
            scale = (297 - left - right) / column_width_count
            if scale > 1:
                scale = 1
            ws.page_setup.scale = scale * 100  # 缩放比例
        else:  # 否则竖向打印并设置装订距离
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            left = 20
            ws.page_margins.left = left / 25.4  # 左边距
            scale = (210 - left - right) / column_width_count
            if scale > 1:
                scale = 1
            ws.page_setup.scale = scale * 100  # 缩放比例

        ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 设置纸张大小为 A4

    # 保存修改后的工作簿
    wb.save(output_path)


def format_excel_file(file_path, output_path, header=0):
    df_dict = pandas.read_excel(file_path, sheet_name=None, header=header)
    excel_writer = pandas.ExcelWriter(output_path)
    for sheet_name, df in df_dict.items():
        save_df_to_sheet(excel_writer, sheet_name, df, wrap_columns=None)
    excel_writer.close()


