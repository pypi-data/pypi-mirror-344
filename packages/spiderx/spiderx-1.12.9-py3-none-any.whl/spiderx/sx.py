# encoding: utf-8
import base64, configparser, datetime, json, os, pickle, time, warnings, random, re, string, requests,sys
from concurrent.futures import ThreadPoolExecutor  # 线程池模块
from urllib.parse import quote, unquote, urlparse, parse_qsl, urlencode
from colorama import init as cmd_color
from lxml.html import etree
from hashlib import md5, sha1
from functools import reduce
from io import StringIO, BytesIO
from smtplib import SMTP
from email.mime.text import MIMEText
from collections import namedtuple
from http.cookiejar import LWPCookieJar
from http import cookiejar
from html import unescape, escape
from contextlib import contextmanager
from pyee import EventEmitter
from .info import *
import threading
import subprocess
import psutil
import socket
import logging
import traceback
import binascii
import winreg
import jsonpath
import signal
import serial
import serial.tools.list_ports
from threading import Thread
from loguru import logger
from decimal import Decimal
import asyncio
import struct
import argparse
import zlib
import inspect
from typing import Union,Coroutine,Generator
threading.stack_size(4096*1000) #创建的线程的堆栈大小  默认40960
sys.setrecursionlimit(100000) #无限递归次数  默认1000
warnings.filterwarnings('ignore')
cmd_color(autoreset=True)
lock = threading.Lock()
emiter = EventEmitter()
#sx.emiter.on("data",lambda x: print(f'say:{x}'))
#sx.emiter.emit("data","hello")
#全局变量
HOME_PATH=os.path.expanduser('~')
FFMPEG_HREF= "http://ip.wgnms.top:9980/media/upload/user_1/2022-01-22/1642855445201.exe"
CLI_HREF= "http://ip.wgnms.top:9980/media/upload/user_1/2022-01-22/1642855494240.exe"
TEMP_DIR=os.path.join(HOME_PATH,'TEMP_DIR') #保存ts类似目录

CREATE_NEW_CONSOLE=16
CREATE_NO_WINDOW=134217728
DETACHED_PROCESS=8
STARTF_USESHOWWINDOW=1
CREATE_BREAKAWAY_FROM_JOB=16777216
CREATE_DEFAULT_ERROR_MODE=67108864

# 字体颜色
def pcolor(msg, _type='yes', end='\n'):
    print("\033[%s;%s%sm%s\033[0m" % (
    0, {'info': 32, 'warn': 33, 'msg': 33, 'error': 31, 'err': 31, 'yes': 36, 'ok': 35}[_type], '', msg),
          end=end)
def scolor(s, _type='warn'):
    return "\033[%s;%s%sm%s\033[0m" % (
    0, {'info': 32, 'warn': 33, 'msg': 33, 'error': 31, 'err': 31, 'yes': 36, 'ok': 35}[_type], '', s)
def 跟踪函数(n:int=None)->str:
    '''
    :param n: 倒数第几行
    :param _type: 输出颜色
    '''
    if n!=None:
        x=traceback.extract_stack()
        if abs(n)>len(x):
            return ''
        x=x[n]
        return f'{os.path.split(x.filename)[1]}/{x.name}/{x.lineno}'
    else:
        return '\n'.join([f'{os.path.split(x.filename)[1]}/{x.name}/{x.lineno}' for x in traceback.extract_stack()])
def 获取_编码(content):
    import chardet
    assert isinstance(content,bytes),'非bytes类型'
    coding=chardet.detect(content)
    return coding['encoding']
def 转编码(文本: str, 编码='utf-8') -> str:
    return 文本.encode(编码, 'ignore').decode(编码)
def decode(content:bytes,encodeing='utf-8'):
    '''
    ignore 忽略不可解码的字节，直接丢弃  Hello World
    replace 用 ? 替代不可解码的字节，保留大部分内容。 Hello ? World
    '''
    return content.decode(encoding=encodeing,errors='ignore')
def 绝对路径(fileName,stack=-1):
    if fileName:
        if os.path.isabs(fileName):
            #完整路径
            return fileName
        else:
            #相对路径
            if getattr(sys, 'frozen', False):
                #exe目录
                path = os.path.dirname(sys.executable)
            else:
                #第一次调用文件目录
                caller_file = inspect.stack()[stack].filename
                path = os.path.dirname(caller_file)
            return os.path.join(path,fileName)
    else:
        return fileName
# 装饰器
@contextmanager
def error(错误提示:str='',抛出异常=False,递归=1,ignore=False)->None:
    '''
    with sx.error(抛出异常=1):
        1/0 #try里面的代码
    '''
    try:
        yield
    except BaseException as e:
        if ignore:
            return
        elif 抛出异常:
            raise e
        else:
            if 递归:
                err = e.__traceback__  # 获取当前错误 赋值err
                while True:
                    if err.tb_next:
                        err = err.tb_next
                    else:
                        lno = err.tb_lineno
                        break

            else:
                lno = e.__traceback__.tb_next.tb_lineno
            if lno!=e.__traceback__.tb_lineno and 递归:
                pcolor('[{}>{}] 错误: {}'.format(e.__traceback__.tb_next.tb_lineno, lno, 错误提示 if 错误提示 else e), 'err')
            # pcolor('[错误 {}] : {}'.format(sys.exc_info()[2].tb_next.tb_lineno ,error if error else e.args),'err')
            else:
                pcolor('[{}] 错误: {}'.format(lno, 错误提示 if 错误提示 else e), 'err')
def exception_hook(exc_type, exc_value, ttraceback):
    '''
    sys.exception_hook=my_Exception
    1/0
    '''
    paths = []
    errors = []
    while ttraceback:
        errors.append(ttraceback.tb_frame.f_locals)
        tracebackCode = ttraceback.tb_frame.f_code
        module = tracebackCode.co_name
        lno = ttraceback.tb_lineno
        file = os.path.split(tracebackCode.co_filename)[1]
        if module == '<module>':
            paths.append(f'{file} -> {lno}')
        else:
            paths.append(f'{file} -> {module} -> {lno}')
        ttraceback = ttraceback.tb_next
    paths.reverse()
    track_path = '\n'.join(paths).strip()
    info = f'错误信息 : {exc_value}\n错误类型 : {exc_type}\n错误跟踪 :\n{track_path}'
    pcolor(info, 'err')
    os._exit(-1)
def exception_gui_hook(exc_type, exc_value, ttraceback):
    '''
    sys.exception_hook=my_Exception
    1/0
    '''
    paths = []
    errors = []
    while ttraceback:
        errors.append(ttraceback.tb_frame.f_locals)
        tracebackCode = ttraceback.tb_frame.f_code
        module = tracebackCode.co_name
        lno = ttraceback.tb_lineno
        file = os.path.split(tracebackCode.co_filename)[1]
        if module == '<module>':
            paths.append(f'{file} -> {lno}')
        else:
            paths.append(f'{file} -> {module} -> {lno}')
        ttraceback = ttraceback.tb_next
    paths.reverse()
    track_path = '\n'.join(paths).strip()
    info = f'错误信息 : {exc_value}\n错误类型 : {exc_type}\n错误跟踪 :\n{track_path}'
    pcolor(info, 'err')
    import tkinter
    root = tkinter.Tk()
    root.title('捕获异常错误')
    x = 600
    y = 600
    width = root.winfo_screenwidth()
    height = root.winfo_screenheight()
    # 'error', 'hourglass', 'info', 'questhead', 'question', 'warning', 'gray12', 'gray25','gray50', 'gray75', 'gray80'
    root.geometry('%dx%d+%d+%d' % (x, y, (width - x) // 2, (height - y) // 2))
    root.iconbitmap("hourglass")
    # root.resizable(width=0, height=0)
    text = tkinter.Text(root, font=('微软雅黑', 10))
    text.config(fg='black')
    text.insert(0.0, info)
    text.config(state=tkinter.DISABLED)
    text.pack(side=tkinter.TOP, fill=tkinter.BOTH, padx=5, pady=5, expand=True)
    btn = tkinter.Button(root, text="关闭", width=10, height=1, font=('微软雅黑', 10), command=root.quit)
    btn.pack(side=tkinter.TOP, pady=5)
    root.mainloop()
    os._exit(-1)
def 打印错误(e: BaseException, 递归=1):
    if 递归:
        err = e.__traceback__  # 获取当前错误 赋值err
        while True:
            if err.tb_next:
                err = err.tb_next
            else:
                lno = err.tb_lineno
                break
    else:
        lno = e.__traceback__.tb_lineno
    if lno!=e.__traceback__.tb_lineno and 递归:
        pcolor('[{}>{}] 错误: {}'.format(e.__traceback__.tb_lineno, lno, e), 'err')
    else:
        pcolor('[{}] 错误: {}'.format(lno, e), 'err')
def zsq_again_return(num=5, cback=None, sleep=0, 显示错误=True, 过滤错误=False, last_err=True):
    过滤错误列表 = ['ProxyError', 'SSLError', 'IncompleteRead']  # 过滤错误

    def rt(func):
        def wear(*args, **keyargs):
            for i in range(num):
                try:
                    return func(*args, **keyargs)
                except BaseException as e:
                    lno = e.__traceback__.tb_next.tb_lineno if e.__traceback__.tb_next else e.__traceback__.tb_lineno
                    放行 = False
                    for item in 过滤错误列表:
                        if item in str(e.args):
                            放行 = True
                            break
                    if 放行:
                        if 过滤错误:
                            pcolor('[{}][{}] 错误: {}'.format(lno,func.__name__, e), 'error')
                    else:
                        if 显示错误:
                            if last_err:
                                if i == num - 1:
                                    pcolor('[{}][{}] 错误: {}'.format(lno,func.__name__, e), 'error')
                            else:
                                pcolor('->{}  [{}][{}] 错误: {}'.format(i + 1, lno, func.__name__, e), 'error' if i == num - 1 else 'warn')
                    time.sleep(sleep)
            return cback
        return wear
    return rt
def zsq_try(func):
    '''    错误装饰器    '''
    def rt(*args, **keyargs):
        try:
            return func(*args, **keyargs)
        except BaseException as e:
            lno = e.__traceback__.tb_next.tb_lineno if e.__traceback__.tb_next else e.__traceback__.tb_lineno
            msg = '[{}][{}] 错误: {}'.format(lno,func.__name__,e)
            pcolor(msg, 'error')
    return rt
def zsq_try_Exception(error: str = None, 递归=1) -> None:
    '''自定义错误装饰器'''
    def rt(func):
        def wear(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except BaseException as e:
                if 递归:
                    err = e.__traceback__  # 获取当前错误 赋值err
                    while True:
                        if err.tb_next:
                            err = err.tb_next
                        else:
                            lno = err.tb_lineno
                            break
                else:
                    lno = e.__traceback__.tb_next.tb_lineno
                if lno !=e.__traceback__.tb_lineno:
                    pcolor('[{}>{}][{}] 错误: {}'.format(e.__traceback__.tb_next.tb_lineno, lno, func.__name__, error if error else e), 'err')
                else:
                    pcolor('[{}][{}] 错误: {}'.format(lno, func.__name__, error if error else e), 'err')
        return wear
    return rt
def zsq_thread(func):
    def wrapper(*args, **kwargs):
        thr = Thread(target=func, args=args, kwargs=kwargs)
        thr.setDaemon(True)  # 跟随程序关闭
        thr.start()
    return wrapper
def config_logging(文件路径: str='logging.log', 输出级别: int = logging.INFO, 写入级别: int = logging.DEBUG):
    '''
    #sx.config_logging('a.txt')
    import logging
    logging.info('hello123')

    logger = logging.getLogger(__name__)
    logger=logging.getLogger()
    logger.info("⼀般⽇志")
    logger.warning("警告⽇志")
    logger.error("错误⽇志")
    logger.debug("错误⽇志")

    %(levelno)s：打印⽇志级别的数值
    %(levelname)s：打印⽇志级别的名称
    %(pathname)s：打印当前执⾏程序的路径，其实就是sys.argv[0]
    %(filename)s：打印当前执⾏程序名
    %(funcName)s：打印⽇志的当前函数
    %(lineno)d：打印⽇志的当前⾏号
    %(asctime)s：打印⽇志的时间
    %(thread)d：打印线程ID
    %(threadName)s：打印线程名称
    %(process)d：打印进程ID
    %(message)s：打印⽇志信息
    '''
    文件路径 = 绝对路径(文件路径)
    file_handler = logging.FileHandler(文件路径, mode='a', encoding="utf8")
    file_handler.setFormatter(logging.Formatter('%(asctime)s %(filename)s LINE:%(lineno)d %(levelname)s >> %(message)s', datefmt="%Y-%m-%d %H:%M:%S"))
    file_handler.setLevel(写入级别)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(logging.Formatter('%(asctime)s %(filename)s LINE:%(lineno)d %(levelname)s >> %(message)s', datefmt="%Y-%m-%d %H:%M:%S"))
    console_handler.setLevel(输出级别)

    logging.basicConfig(level=min(输出级别, 写入级别), handlers=[file_handler, console_handler])
def get_fake_agent(浏览器:str='chrome')->str:
    '''chrome opera firefox internetexplorer safari'''
    return random.choice(fake_UserAgent['browsers'][浏览器])
def get_headers(浏览器:str=None)->dict:
    '''chrome opera firefox internetexplorer safari'''
    if 浏览器:
        agent = get_fake_agent(浏览器)
    else:
        agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    return {'User-Agent'.lower(): agent}
def 退出(信息:str='',类型=1)->None:
    if 类型 and 信息:  # 抛出异常退出
        sys.exit("异常退出 : {}".format(信息)) if 信息 else sys.exit(0)
    else:     # 强制退出
        os._exit(0)
# 功能函数
def 读取文件(文件路径: str) -> bytes:
    文件路径 = 绝对路径(文件路径)
    if os.path.exists(文件路径):
        with open(文件路径, 'rb') as f:
            return f.read()
    else:
        pcolor(f'{文件路径} 文件不存在。', 'error')
        return None
def 写入文件(文件路径, 字节流) -> None:
    文件路径 = 绝对路径(文件路径)
    with open(文件路径, 'wb') as f:
        f.write(字节流)
def 写入文件a(文件路径, 字节流) -> None:
    文件路径 = 绝对路径(文件路径)
    with open(文件路径, 'ab') as f:
        f.write(字节流)
def 加载文件(文件路径: str, 编码: str = 'utf-8-sig')->str:
    文件路径=绝对路径(文件路径)
    try:
        with open(文件路径, 'r', encoding=编码) as f:
            return f.read()
    except Exception as e:
        with open(文件路径, 'r', encoding='ANSI') as f:
            return f.read()
def 加载文件_创建(文件路径:str,编码:str='utf-8-sig',文本='')->str:
    '''自动创建文件 返回'''
    文件路径=绝对路径(文件路径)
    if os.path.exists(文件路径):
        return 加载文件(文件路径,编码)
    else:
        with open(文件路径,mode='w',encoding=编码) as f:
            f.write(文本)
            return 文本
def 保存文件(文件路径, 字符串, 编码='utf-8-sig')->None:
    文件路径 = 绝对路径(文件路径)
    with open(文件路径, 'w', encoding=编码) as f:
        f.write(字符串)
def 保存文件a(文件路径, 字符串, 编码='utf-8-sig')->None:
    文件路径 = 绝对路径(文件路径)
    with open(文件路径, 'a', encoding=编码) as f:
        f.write(字符串)
def 加载对象(文件路径)->object:
    文件路径 = 绝对路径(文件路径)
    with open(文件路径, 'rb') as f:
        return pickle.load(f)
def 保存对象(文件路径, 对象) -> None:
    文件路径 = 绝对路径(文件路径)
    with open(文件路径, 'wb') as f:
        pickle.dump(对象, f)
def 加载JSON(文件路径, 编码='utf-8-sig'):
    文件路径 = 绝对路径(文件路径)
    try:
        with open(文件路径, 'r', encoding=编码) as f:
            return json.load(f)
    except Exception as e:
        with open(文件路径, 'r', encoding='ANSI') as f:
            return json.load(f)
def 保存JSON(文件路径, JSON对象, 编码='utf-8-sig', indent=4):
    '''indent=None不格式化'''
    文件路径 = 绝对路径(文件路径)
    with open(文件路径, 'w', encoding=编码) as f:
        json.dump(JSON对象, f, ensure_ascii=False, indent=indent)
def 加载文件_xlsx(文件路径,min_row=None,max_row=None,min_col=None,max_col=None,sheet_name=None)->list:
    文件路径 = 绝对路径(文件路径)
    from openpyxl import load_workbook
    wb = load_workbook(文件路径)
    ws = wb[sheet_name] if sheet_name else wb.active
    data = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,min_col=min_col,max_col=max_col, values_only=True):
        data.append(list(row))
    wb.close()
    return data
def 保存文件_xlsx(文件路径,data=[[1],],sheet_name=None,width=30)->bool:
    '''data=[['a','2',1],['a','2',1]]'''
    文件路径 = 绝对路径(文件路径)
    try:
        from openpyxl import Workbook, utils
        wb = Workbook()
        ws = wb.create_sheet(title=sheet_name) if sheet_name else wb.active
        for row in data:
            ws.append(row)
        # 设置列宽
        for i in range(1, len(data[0]) + 1):
            ws.column_dimensions[utils.get_column_letter(i)].width = width
        wb.save(文件路径)
        wb.close()
        return True
    except Exception as e:
        打印错误(e)
        return False
def 字符串转日期格式(时间字符串, 中文=False):
    if 中文:
        return '{:%Y年%m月%d日 %H时%M分%S秒}'.format(datetime.datetime.strptime(时间字符串, "%Y-%m-%d %H:%M:%S"))
    else:
        return datetime.datetime.strptime(时间字符串, "%Y-%m-%d %H:%M:%S")
def 日期格式转时间戳(日期格式=None):
    if 日期格式:
        return 日期格式.timestamp()
    else:
        return datetime.datetime.today().timestamp()
def 时间戳转日期格式(时间戳, 中文=False):
    if 中文:
        dateArray = datetime.datetime.fromtimestamp(时间戳)
        return dateArray.strftime("%Y年%m月%d %H时%M分%S秒")
    else:
        return datetime.datetime.utcfromtimestamp(时间戳)
def 隧道代理(proxyMeta="http://user:pwd@host:port") -> dict:
    '''代理服务器 阿布云 或者 亿牛云代理'''
    return {"http": proxyMeta, "https": proxyMeta, }
def 列表分组(列表: list, step: int) -> list:
    return [列表[i:i + step] for i in range(0, len(列表), step)]
def 合并列表(*a, default=None) -> list:
    '''[1],[2],[3],...'''
    lst = []
    L = max([len(x) for x in a])
    for i in range(L):
        r = []
        for j in range(len(a)):
            try:
                r.append(a[j][i])
            except:
                r.append(default)
        lst.append(r)
    return lst
def base64加密(input_: str or bytes, 编码='utf-8')->str:
    if type(input_)==bytes:
        return base64.encodebytes(input_).decode(编码).strip()
    elif type(input_)==str:
        return base64.b64encode(input_.encode(编码)).decode(编码)
    else:
        raise Exception('base64加密 输入类型错误')
def base64解密(字符串: str,编码='utf-8')->str:
    '''
        base64 包括 ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/=

        1.
            标准的Base64并不适合直接放在URL里传输，因为URL编码器会把标准Base64中的“/”和“+”字符变为形如“%XX”的形式，
            而这些“%”号在存入数据库时还需要再进行转换，因为ANSI SQL中已将“%”号用作通配符
            可采用一种用于URL的改进Base64编码，它在末尾填充'='号，并将标准Base64中的“+”和“/”分别改成了“-”和“_”
        2.
            另有一种用于正则表达式的改进Base64变种，它将“+”和“/”改成了“!”和“-”，
            因为“+”,“*”以及前面在IRCu中用到的“[”和“]”在正则表达式中都可能具有特殊含义。
        '''
    字符串 = 字符串.replace('-', '+').replace('_', '/').strip('=')
    for i in range(3):
        try:
            return base64.b64decode(字符串+'=' * i).decode(编码)
        except:
            pass
def base64加密图片(字节流: bytes, 图片类型='jpg') -> str:
    '''图片转字符串  图片类型gif png jpeg x-icon'''
    return ','.join(['data:image/{};base64'.format(图片类型), base64.b64encode(字节流).decode()])
def base64解密图片(字符串: str) -> bytes:
    '''字符串转成图片  图片类型jpg 或者 png'''
    return base64.b64decode(re.sub('data:image/.*?;base64,', '', 字符串))
def base64图片转PIL(base64字符串:str):
    from PIL import Image
    return Image.open(BytesIO(base64解密图片(base64字符串)))
def url编码(s):
    return quote(s)
def url解码(s):
    return unquote(s)
def params_From_dict(query:dict)->str:
    result = []
    for k, v in query.items():
        result.append(f'{k}={v}')
    return '&'.join(result)
def str_From_dict(d: dict, 分隔1: str = ':', 分隔2: str = '\n') -> str:
    return f'{分隔2}'.join(['{}{}{}'.format(k.strip(), 分隔1, v.strip()) for k, v in d.items()])
def dict_From_Str(s: str, 分隔1: str = ':', 分隔2: str = '\n') -> dict:
    '''a:1\nb:2'''
    rt={}
    for x in s.strip().split(分隔2):
        x=x.strip()
        if x:
            d=x.split(分隔1, 1)
            if len(d) == 2:
                rt[d[0].strip()] = d[1].strip()
    return rt
def dict_From_HeadersStr(head: str = '', 浏览器=None) -> dict:
    ''' s="host:xxx.com"  '''
    headers = get_headers(浏览器)
    head = head.strip()
    if head:
        head = head.replace(':\n', ':')
        for row in head.split('\n'):
            row = row.strip()
            if row:
                y = row.split(':', 1)
                if len(y) == 2:
                    headers[y[0].lower().strip()] = y[1].strip()
    return headers
def dict_From_DataStr(s: str) -> dict:
    ''' s="a=1&b=2"   '''
    rt={}
    for x in s.strip().split('&'):
        x=x.strip()
        if x:
            d=x.split('=', 1)
            if len(d) == 2:
                rt[d[0].strip()] = d[1].strip()
    return rt
def dict_From_CookieStr(s: str) -> dict:
    '''a=b;b=c;'''
    rt={}
    for x in s.strip().split(';'):
        x=x.strip()
        if x:
            d=x.split('=', 1)
            if len(d) == 2:
                rt[d[0].strip()] = d[1].strip()
    return rt
def dict_From_CookieJar(cook):
    return {c.name:c.value for c in cook}
def dict_From_Cookiejar_Str(cookie_str: str = None, 列=[0, 1]) -> dict:
    ''' cookie_From_Cookiejar_Str('AIDUID	CCB5060E627C5BD5804CDD46A7C050FF:FG=1	.baidu.com	/	2022-08-17T02:14:53.615Z	44			') '''
    rt_dict = {}
    for x in cookie_str.strip().split('\n'):
        if x.strip():
            c = x.split('\t')
            rt_dict[c[列[0]].strip()] = c[列[1]].strip()
    return rt_dict
def dict_From_DataStr_QueryStr(s: str) -> dict:
    return dict_From_Str(s, ':', '\n')
def cookie_From_Cookies(cook: object) -> str:
    '''cook对象转字符串'''
    xx = []
    for k, v in cook.items():
        xx.append('{}={}'.format(k, v))
    return ';'.join(xx)
def cookie_From_Application_Cookies(cookie:str)->str:
    return str_From_dict(dict_From_Cookiejar_Str(cookie), '=',';')
def cookie_From_CookieJar(cook: object) -> str:
    '''CookieJar对象'''
    cookies=[]
    for c in cook:
        cookies.append('{}={}'.format(c.name, c.value))
    return ';'.join(cookies)
def cookiejar_From_CookieStr(cookie_str:str):
    import requests
    from http.cookiejar import CookieJar
    from requests.utils import cookiejar_from_dict
    cookies_dict = {}
    for cookie in cookie_str.split(';'):
        key, value = cookie.split('=',1)
        cookies_dict[key.strip()] = value.strip()
    cookie_jar = cookiejar_from_dict(cookies_dict, CookieJar())
    #session = requests.session()
    #session.cookies=cookie_jar
    return cookie_jar
def get_brower_cookies(domain_name:str=""):
    '''
    浏览器包括 [chrome, chromium, opera, opera_gx, brave, edge, vivaldi, firefox, safari]
    cookies=get_brower_cookies(".qq.com")
    requests.get(url,cookies=cookies)
    '''
    from .browserCookie3 import load
    cookjar=load(domain_name=domain_name)
    if not cookjar:
        print('没有登陆信息,请先登陆浏览器[chrome, chromium, opera, opera_gx, brave, edge, vivaldi, firefox, safari]')
    return cookjar
def 保存Cookiejar_From_CookiejarStr(文件路径: str = None, cookie_str: str = None) -> str or None:
    '''
    cookiejar_From_Str("cookie.txt",'AIDUID	CCB5060E627C5BD5804CDD46A7C050FF:FG=1	.baidu.com	/	2022-08-17T02:14:53.615Z	44			')
    :param 文件路径: cookie保存的路径
    :param cookie_str: 复制f12 application cookies的表格字符串
    :return:如果有文件路径直接保存返回空 如果没有则返回字符串
    '''
    # 字符串列
    文件路径 = 绝对路径(文件路径)
    keys = ['name', 'value', 'domain', 'path', 'expires-max-age', 'size', 'http', 'secure', 'samesite']
    rt_s = ''
    if cookie_str:
        for x in cookie_str.strip().split('\n'):
            if x.strip():
                values = x.split('\t')
                c = dict(合并列表(keys, values))
                # 保留一级域名如 .bilibili.com
                domain = c['domain'].split('.')
                domain[0] = ''
                domain = '.'.join(domain)
                row = [domain, 'TRUE', c['path'], 'FALSE', '4788470561', c['name'], c['value']]  # 过期时间100年
                rt_s += '\t'.join(row) + '\n'
    if 文件路径:
        with open(文件路径, 'w') as f:
            f.write(rt_s)
    else:
        return rt_s
def 保存LWPCookieJar_From_CookiejarStr(文件路径: str, cookie_str: str) -> None:
    '''cookiejar_From_Str("cookie.txt",'AIDUID	CCB5060E627C5BD5804CDD46A7C050FF:FG=1	.baidu.com	/	2022-08-17T02:14:53.615Z	44			')'''
    文件路径 = 绝对路径(文件路径)
    jar = LWPCookieJar()
    keys = ['name', 'value', 'domain', 'path', 'expires-max-age', 'size', 'http', 'secure', 'samesite']
    for x in cookie_str.split('\n'):
        if x.strip():
            values = x.split('\t')
            c = dict(合并列表(keys, values))
            c['expires-max-age'] = '4788470561'  # 过期时间100年
            jar.set_cookie(
                cookiejar.Cookie(version=0, name=c['name'], value=c['value'], domain=c['domain'], path=c['path'],
                                 secure=c['secure'],
                                 expires=c['expires-max-age'] if "expires-max-age" in c else None,
                                 domain_specified=True, domain_initial_dot=False,
                                 path_specified=True,
                                 rest={}, discard=False, comment=None, comment_url=None, rfc2109=False,
                                 port='80', port_specified=False,
                                 ))
    jar.save(文件路径)
def 加载配置文件(默认配置: dict, 文件路径='conf.ini', 编码='utf-8-sig') -> dict:
    文件路径 = 绝对路径(文件路径)
    config = configparser.ConfigParser()
    if not os.path.exists(文件路径):
        config['conf'] = 默认配置
        with open(文件路径, 'w', encoding=编码) as f:
            config.write(f)
    try:
        config.read(文件路径, encoding=编码)
    except Exception as e:
        config.read(文件路径, encoding='ANSI')
    return config['conf']
def 保存配置文件(配置对象: configparser.SectionProxy, 文件路径='conf.ini', 编码='utf-8-sig') -> None:
    文件路径 = 绝对路径(文件路径)
    config = configparser.ConfigParser()
    config['conf'] = 配置对象
    with open(文件路径, 'w', encoding=编码) as f:
        config.write(f)
def cmd_popen(*cmd) -> str:
    '''文本方式'''
    with os.popen(*cmd) as f:
        return f.read()  # 获取管道信息
def cmd_system(*cmd) -> int:
    '''
    返回命令执行结果的返回值  阻塞
    返回0 运行成功 1没有这个命令'''
    return os.system(*cmd)
def cmd_subprocess_run(cmd:Union[str,list], 类型=1,timeout=None)->bool:
    '''返回0 运行成功 1没有这个命令
    CREATE_NEW_CONSOLE：在新控制台窗口中启动子进程。
    DETACHED_PROCESS：将子进程从父进程分离，使其成为一个独立的进程组。
    CREATE_NO_WINDOW：在后台启动子进程，不创建窗口显示。
    CREATE_DEFAULT_ERROR_MODE：使用默认的错误模式处理子进程的错误。
    CREATE_BREAKAWAY_FROM_JOB：使子进程从作业对象脱离，成为一个独立的进程。
    CREATE_SUSPENDED：创建一个挂起的进程，需要调用ResumeThread()函数来启动执行。
    rt=subprocess.Popen(cmd)#非阻塞
    '''
    if 类型 == 1:
        # 不允许创建窗口 隐藏窗口  无输出
        si = subprocess.STARTUPINFO()
        si.dwFlags = STARTF_USESHOWWINDOW
        rt = subprocess.run(cmd,startupinfo=si,creationflags=CREATE_NO_WINDOW,timeout=timeout)
    elif 类型 == 2:
        # 隐藏窗口 有输出
        si = subprocess.STARTUPINFO()
        si.dwFlags = STARTF_USESHOWWINDOW
        rt = subprocess.run(cmd, startupinfo=si,timeout=timeout)
    elif 类型 == 3:
        #在新控制台窗口中启动子进程。
        rt = subprocess.run(cmd, creationflags=CREATE_NEW_CONSOLE,timeout=timeout)
    elif 类型 == 4:
        # 不允许创建窗口
        #在后台启动子进程，不创建窗口显示
        rt = subprocess.run(cmd, creationflags=CREATE_NO_WINDOW,timeout=timeout)
    else:
        # 不允许子窗口
        # 将子进程从父进程分离，使其成为一个独立的进程组。
        rt = subprocess.run(cmd, creationflags=DETACHED_PROCESS)
    if rt.returncode==0:
        return True
    else:
        return False
def cmd_subprocess_popen(cmd,类型=1)->str:
    if 类型==1:
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, stderr = proc.communicate()
        try:
            text = stdout.decode()
        except:
            text = stdout.decode("ANSI")
        return text
    elif 类型==2:
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, universal_newlines=True)
        text=''
        for line in iter(proc.stdout.readline, ''):
            print(line.strip())
            text+=line.strip()+'\n'
        proc.stdout.close()
        returncode = proc.wait()
        return text
def 设置_临时环境变量(k, v):
    os.environ[k] = v
def 设置_永久环境变量(k, v, tp=1):
    '''tp  1用户变量 2系统变量 '''
    if tp == 1:
        cmd_subprocess_run('setx {} {}'.format(k, v))  # 用户变量
    else:
        cmd_subprocess_run('setx {} {} /m'.format(k, v))  # 系统变量
def 随机字符串(长度=6, 类型=3):
    '''1 数字 2 字母 3字母数字 4字母数字特殊符号'''
    if 类型 == 1:
        s = string.digits
    elif 类型 == 2:
        s = string.ascii_letters
    elif 类型 == 3:
        s = string.ascii_letters + string.digits
    else:
        s = string.ascii_letters + string.digits + string.punctuation
    rt = []
    for i in range(长度):
        rt.append(random.choice(s))
    return ''.join(rt)
def 执行代码(python代码):
    '''
    print("hello")
    '''
    exec(python代码)
def 进制转换(进制对象: str, 当前进制: int, 转为进制: int, 去符号: bool = True) -> str:
    ''' print(进制转换(96,16,8)) print(进制转换('0x96',16,8)) '''
    try:
        十进制 = int(str(进制对象), 当前进制)
        if 转为进制 == 10:
            return str(十进制)
        elif 转为进制 == 2:
            rt = bin(十进制)
        elif 转为进制 == 8:
            rt = oct(十进制)
        elif 转为进制 == 16:
            rt = hex(十进制)
        return rt[2:] if 去符号 else rt
    except Exception as e:
        raise Exception('进制转换错误')
def 进制转换_16TO字符串(s16, 编码='utf-8'):
    from binascii import a2b_hex
    s16 = s16 if type(s16) == bytes else s16.encode(编码)
    return a2b_hex(s16).decode(编码)
def 进制转换_字符串TO16(s, 编码='utf-8'):
    from binascii import b2a_hex
    s = s if type(s) == bytes else s.encode(编码)
    return b2a_hex(s).decode(编码)
def get_uuid():
    import uuid
    return str(uuid.uuid1())
def byte_to_hex(b:bytes)->bytes:
    return binascii.b2a_hex(b)
    #return binascii.hexlify(b)
def hex_to_byte(b:bytes)->bytes:
    return binascii.a2b_hex(b)
    #return binascii.unhexlify(s)
def hex_to_str(b:bytes,编码='utf-8')->str:
    return binascii.a2b_hex(b).decode(编码)
def str_to_hex(s:str,编码='utf-8')->bytes:
    return binascii.b2a_hex(s.encode(编码))
def byte_to_base64(b:bytes)->bytes:
    return binascii.b2a_base64(b).strip()
def base64_to_byte(b:bytes)->bytes:
    'base64 包括 ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/='
    b = b.replace(b'-', b'+').replace(b'_', b'/').strip(b'=')
    for i in range(3):
        try:
            return binascii.a2b_base64(b+ b'=' * i)
        except:
            pass
def base64_to_hex(b: bytes or str) -> bytes:#base64->bytes->hex
    if isinstance(b, str):
        b=binascii.a2b_base64(b)
    return binascii.b2a_hex(b)
def byte_to_uint8array(b: bytes) -> list:
    '''[x for x in bytearray(b)]'''
    return list(b)
def uint8array_to_byte(lst: list) -> bytes:
    return bytes(lst)
def uint32array_to_byte(lst:list,fmt='>I') -> bytes:
    '''
    节在机器中存储的字节顺序  < 低位 >高位 ！network @ notive  = native
    uint32array_to_byte(fmt='<I',lst=[3854078970, 2917115795, 3887476043, 3350876132])
    '''
    rt=b''
    for num in lst:
        rt+=struct.pack(fmt, num)
    return rt
def byte_to_uint32array(b: bytes,fmt='>4I') -> list:
    '''
    UINT8、UINT16、UINT32、UINT64等数字类型，即分别对应将每1、2、4、8个字节放一起解释为一个数字
    strcut.unpack(fmt, byte) 返回值是一个list,从二进制流类型,变为int型
    struct.unpack(fmt , b) 低位在前 4*8 uint32array
    B uint8类型
    b int8类型
    H uint16类型
    h int16类型
    I uint32类型
    i int32类型
    L uint64类型
    l int64类型
    s ascii码，s前带数字表示个数
    byte_to_uint32array(bytes([250, 147, 184, 229, 147, 167, 223, 173, 75, 45, 182, 231, 228, 79, 186, 199]))
    '''
    return struct.unpack(fmt , b)
def hex_to_base64(b:bytes)->bytes: #h2x->bytes->base64
    return  binascii.b2a_base64(binascii.a2b_hex(b)).strip()
def str_to_ascii(字符串: str) -> list:
    return [ord(x) for x in 字符串]
def ascii_to_str(列表: list,字符串=True) -> list or str:
    if 字符串:
        return ''.join([chr(x) for x in 列表])
    else:
        return [chr(x) for x in 列表]
def int_to_byte(num:int,length=4,byteorder='big'):
    '''big little 高低位'''
    '''int_to_hex(97,4)  -> a '''
    return int(num).to_bytes(length=length,byteorder= byteorder)
def byte_to_int(b:bytes,byteorder='big'):
    '''big little 高低位'''
    '''byte_to_int(‘你’.encode())'''
    return int().from_bytes(b,byteorder=byteorder)
def html不转义(字符串):
    return unescape(字符串)
def html转义(字符串):
    return escape(字符串)
def 获取_搜索文件(文件夹目录: str, 关键字: str) -> Generator:
    '''
    sx.搜索文件(r'D:\python_project\aa', '*mp4 *.mp3')
    模糊查询  只能获取文件名 无法获取文件夹
    '''
    文件夹目录 = 绝对路径(文件夹目录)
    import glob
    keys = [x.strip() for x in 关键字.split(' ') if x.strip()]
    result=[]
    for root, lists, files in os.walk(文件夹目录):
        for key in keys:
            file_pattern = os.path.join(root, key)
            for fpath in glob.glob(file_pattern):
                if fpath not in result:
                    result.append(fpath)
                    yield fpath
def 获取_目录文件(文件夹目录: str, 打印=False):
    文件夹目录=绝对路径(文件夹目录)
    class DirFile():
        文件夹 = [];
        文件 = [];
        文件夹数 = 文件数 = 0

    result = DirFile()
    if not os.path.exists(文件夹目录):
        return result
    for root, lists, files in os.walk(文件夹目录):
        for file in files:
            file_path = os.path.join(root, file)
            if 打印:
                print('文件', file_path)
            if file_path not in result.文件:
                result.文件.append(file_path)
                result.文件数 += 1
        for dir in lists:
            dir_path = os.path.join(root, dir)
            if 打印:
                print('目录', dir_path)
            else:
                if dir_path not in result.文件夹:
                    result.文件夹.append(dir_path)
                    result.文件夹数 += 1
    return result
def dat_to_jpg(文件路径, 输出文件路径='out', 保留目录参数=1):
    def imageXor(f):
        """
        计算异或值
        各图片头部信息
        jpeg：ff d8 ff
        png：89 50 4e 47
        gif： 47 49 46 38
        """
        dat_read = open(f, "rb")
        try:
            a = [(0x89, 0x50, 0x4e), (0x47, 0x49, 0x46), (0xff, 0xd8, 0xff)]
            for now in dat_read:
                for xor in a:
                    i = 0
                    res = []
                    nowg = now[:3]
                    for nowByte in nowg:
                        res.append(nowByte ^ xor[i])
                        i += 1
                    if res[0] == res[1] == res[2]:
                        return res[0]
        except:
            pass
        finally:
            dat_read.close()
    def imageDecode(fileName):
        """ param f: 微信图片路径 """
        # 图片输出路径
        path, fn = os.path.split(fileName)
        out = '/'.join([输出文件路径] + path.split('\\')[-保留目录参数:]) + '/' + fn[:-4] + '.jpg'
        path, _ = os.path.split(out)
        if os.path.exists(out):
            return
        else:
            os.makedirs(path, exist_ok=1)
            # 先计算出偏移值
            change_Byte = imageXor(fileName)
            # 读取.bat
            dat_read = open(fileName, "rb")  # 图片写入
            png_write = open(out, "wb")  # 循环字节
            try:
                for now in dat_read:
                    for nowByte in now:
                        newByte = nowByte ^ change_Byte  # 转码计算
                        png_write.write(bytes([newByte]))  # 转码后重新写入
            except Exception as e:
                pass
            dat_read.close()
            png_write.close()
    pool = ThreadPoolExecutor(max_workers=30)
    for file in 获取_搜索文件(文件路径, 关键字='*.dat'):
        if os.path.isfile(file):
            if file[-4:] == '.dat':
                pool.submit(imageDecode, file)
    pool.shutdown(wait=True)
    print('done')
def 获取_正则搜索文件(文件夹目录: str,正则表达式='', 打印=False):
    文件夹目录 = 绝对路径(文件夹目录)
    class DirFile():
        文件夹 = [];
        文件 = [];
        文件夹数 = 文件数 = 0

    result = DirFile()
    if not os.path.exists(文件夹目录):
        return result
    for root, lists, files in os.walk(文件夹目录):
        for file in files:
            if re.match(正则表达式,file):
                file_path = os.path.join(root, file)
                if 打印:
                    print('文件', file_path)
                if file_path not in result.文件:
                    result.文件.append(file_path)
                    result.文件数 += 1
        for dir in lists:
            if re.match(正则表达式,dir):
                dir_path = os.path.join(root, dir)
                if 打印:
                    print('目录', dir_path)
                else:
                    if dir_path not in result.文件夹:
                        result.文件夹.append(dir_path)
                        result.文件夹数 += 1
    return result
def 检查文件名(s, 编码='utf-8',replace=' '):
    c = '\/:*?"<>|\x08'
    s = s.encode(编码, 'ignore').decode(编码).replace('\\',replace).replace('/',replace)
    s = ' '.join(s.split()) #去掉\r \n \t
    name = ''.join([x for x in s if x not in c]).strip() #去掉'\/:*?"<>|\x08'
    if not name:
        name = 随机字符串(11, 3)
    for i in range(len(name) - 1, -1, -1):  #去掉结尾字符串
        if name[i] in [' ', '.']:
            name=name[:-1]
        else:
            break
    return name
def 创建目录(文件路径: str, 提示类型: int = 1, 编码: str = 'utf-8'):
    '''
    作用：自动创建目录 并且检查是否可以创建 如果 已存在返回 False 不存在返回True
    if 创建目录(目录="abc/123",文件名='a.txt')['bl']:
        print(1)
    True 可以创建文件  False 不能创建
    提示类型 1 全路径 2 文件名
    '''
    文件路径 = 绝对路径(文件路径)
    class CreateFilepath:
        可创建 = 目录 = 文件名 = 文件路径 = 类型 = 描述 = ''
        def __str__(self):
            return str({x: self.__getattribute__(x) for x in dir(self) if '__' not in x})

    文件路径 = 文件路径.encode(编码, 'ignore').decode(编码).replace('/', "\\")
    文件路径 = ' '.join(文件路径.split())
    目录, 文件名 = os.path.split(文件路径)
    rt = CreateFilepath()

    新的文件名=检查文件名(文件名)
    新的路径 = ''

    if 目录:  # 新路径
        文件夹列表 = 目录.split('\\')
        if 文件夹列表:
            dir_new = []
            if re.match('^\w:$', 文件夹列表[0]):  # c: d:
                # 绝对路径
                dir_new.append(文件夹列表[0])
                for d in 文件夹列表[1:]:
                    dir_new.append(检查文件名(d))
            else:
                for d in 文件夹列表:
                    dir_new.append(检查文件名(d))
            新的路径 = '\\'.join(dir_new)
            if 新的路径:
                os.makedirs(os.path.abspath(新的路径), exist_ok=True)
    if (not 新的文件名) and (not 新的路径):
        rt.描述 = '文件路径错误'
        return rt
    elif 新的文件名:
        rt.类型 = '文件'
    else:
        rt.类型 = '目录'
    save_name = os.path.join(新的路径, 新的文件名) if 目录 else 新的文件名
    if os.path.exists(save_name):
        if 新的文件名:
            rt.可创建 = False
            rt.描述 += '已存在'
        else:
            rt.可创建 = False
            rt.描述 += '空'
    else:
        if 新的文件名:
            rt.可创建 = True
            rt.描述 += '不存在'
        else:
            rt.可创建 = False
            rt.描述 += '空'
    rt.目录 = 新的路径
    rt.文件名 = 新的文件名
    rt.文件路径 = save_name
    if not rt.可创建 and 提示类型:
        if 提示类型 == 1:
            pcolor('已存在 : {}'.format(save_name), 'ok')
        elif 提示类型 == 2:
            pcolor('已存在 : {}'.format(新的文件名), 'ok')
        else:
            pass
    return rt
def 定时运行(秒: int, 函数, 参数: list):
    '''#定时器 单位秒 只执行一次  sx.定时运行(1,task,(1,2,3))'''
    from threading import Timer
    t = Timer(interval=秒, function=函数, args=参数, kwargs=None)
    t.setDaemon(True)
    t.start()
def 排序_列表里字典(列表, 键, 倒序=False) -> list:
    return sorted(列表, key=lambda d: d[键], reverse=倒序)  # False 正序
def 排序_字典键值(字典, 位置: int, 倒序=False) -> dict:
    def sort_key(item):
        x = item[位置]
        if 位置 == 0:  # 按key排序
            s = str(x)
            if s.isdigit():
                return (0, int(s))
            else:
                return (1, s)
        else:  # 按value排序
            # 先把数字和字符串区分，数字优先按数字大小，字符串按字母排序
            if isinstance(x, (int, float)):
                return (0, x)
            elif isinstance(x, str):
                if x.isdigit():
                    return (0, int(x))  # 数字字符串也当数字
                else:
                    return (1, x)
            else:
                # 其他类型放后面，转成字符串比较
                return (2, str(x))
    return dict(sorted(字典.items(), key=sort_key, reverse=倒序))
def 排序_字典键(字典, 倒叙=False) -> dict:
    def sort_key(k):
        x = str(k[0])
        if x.isdigit():
            return (0, int(x))
        else:
            return (1, x)
    return dict(sorted(字典.items(),key=sort_key,reverse=倒叙))
def 排序_列表里元组(列表, 位置, 倒叙=False) -> list:
    return sorted(列表, key=lambda d: d[位置], reverse=倒叙)  # False 正序
def 排序_列表(列表, 倒叙=False) -> list:
    return sorted(列表, reverse=倒叙)
def 列表_字典分组(列表:list,健:str,排序健:str='',倒叙=False)->dict:
    # from operator import itemgetter
    # from itertools import groupby
    # from collections import defaultdict
    # lst=sorted(列表,key=lambda k:k[健],reverse=倒叙)
    # lst = groupby(lst, key=itemgetter(健))
    # rt=defaultdict(list)
    # for k, v in lst:
    #     for x in v:
    #         rt[k].append(x)
    # return rt
    from collections import defaultdict
    rt=defaultdict(list)
    列表.sort(key=lambda k:k[排序健 if 排序健 else 健],reverse=倒叙)
    for x in 列表:
        rt[x[健]].append(x)
    return rt
def 集合_交集(*args):
    '''[1,2,3],[2],[1,2,3]'''
    return list(reduce(lambda a, b: a & b, [set(x) for x in args]))
def 集合_并集(*args):
    return list(reduce(lambda a, b: a | b, [set(x) for x in args]))
def 集合_差集(*args):
    return list(reduce(lambda a, b: a - b, [set(x) for x in args]))
def 正则_提取中文(s):
    p = re.compile(r'[\u4e00-\u9fa5]')
    res = re.findall(p, s)
    result = ''.join(res)
    return result
def 加密_MD5(对象, 加密字符串=None) -> str:
    '''
    :param 对象: 加密字符串
    :param 加密字符串: 密码
    :return: 返回加密后16进制字符串
    '''
    hsobj = md5(str(对象).encode("utf-8"))
    if 加密字符串:
        hsobj.update(str(加密字符串).encode("utf-8"))
    return hsobj.hexdigest()
def 加密_SHA1(对象: str)->str:
    return sha1(str(对象).encode('utf-8')).hexdigest()
def 加密_HMAC_MD5(对象:bytes, 加密字符串:bytes)->str:
    '''
    :param 对象: 加密字符串
    :param 加密字符串: 密码
    :return: 返回加密后16进制字符串
    '''
    import hmac
    import hashlib
    m=hmac.new(加密字符串,对象,digestmod=hashlib.md5)
    return m.hexdigest()
def 加密_HMAC_SHA256(对象:bytes, 加密字符串:bytes)->str:
    '''
    HMAC-SHA256和SHA256是两种不同的加密算法。

    SHA256是一种单向散列函数，它将任意长度的输入数据转换为固定长度的输出数据，通常为256位。SHA256算法具有以下特点：
        不可逆：无法从SHA256的输出数据推导出原始输入数据。
        相同输入产生相同输出：对于相同的输入数据，SHA256算法总是会产生相同的输出数据。
        雪崩效应：即使输入数据发生微小的改变，SHA256的输出数据也会发生巨大的变化。
    HMAC-SHA256是在SHA256的基础上加入了密钥的散列算法，用于增加数据的安全性。HMAC-SHA256算法具有以下特点：
        需要一个密钥：HMAC-SHA256算法需要一个密钥作为输入，用于增加数据的安全性。
        可验证性：使用相同的密钥和输入数据进行HMAC-SHA256计算，可以验证计算结果是否一致。
        防止篡改：HMAC-SHA256算法可以防止数据在传输过程中被篡改。
    '''
    import hmac
    import hashlib
    #hashlib.sha256 hashlib.sha1 ...
    return hmac.new(加密字符串, 对象,digestmod=hashlib.sha256).hexdigest()
def 获取_TXT行列(文件路径: str, 分割行:str='\n' , 分割列:str='\t') -> list:
    '''
    s="第一个视频\thttps://cd15-ccd1-2.play.bokecc.c"
    sx.获取_TXT行列('urls.txt',分割列=r"\t",分割行='\n')
    -->  [['第一个视频', 'https://cd15-ccd1-2.play.bokecc.c']]
    '''''
    文件路径 = 绝对路径(文件路径)
    if os.path.exists(文件路径):
        s=加载文件(文件路径)
    else:
        pcolor(f'文件不存在:{文件路径}','err')
        return []
    return [[d for d in row.split(分割列)] if 分割列 else row.strip() for row in s.split(分割行) if row.strip()]
def 打印_进度条(字符串, 当前ID, 总数, 步长, 下载速度=None, 符号='█', 符号2='░', 进度条长度=30, 类型=4):
    '''打印_进度条('下载文件',0,100,类型=1)'''
    当前ID = 当前ID + 步长
    百分百 = 当前ID / 总数
    if 下载速度:
        speed = ' {}'.format(下载速度)
    else:
        speed = ''
    L = int(进度条长度 * 百分百)
    if 类型 == 1:
        print(('\r{:<%d} {:>4} {} {}/{}{}' % 进度条长度).format(L * 符号 + (进度条长度 - L) * 符号2, f'{int(100 * 百分百)}%', 字符串, 当前ID,
                                                           总数, speed), end='', flush=True)
    elif 类型 == 2:
        print(("\r{:>4} {:<%d} {} {}/{}{}" % 进度条长度).format(f'{int(100 * 百分百)}%', L * 符号 + (进度条长度 - L) * 符号2, 字符串, 当前ID,
                                                           总数, speed), end='', flush=True)
    elif 类型 == 3:
        print(('\r{:<%d} {:>4} {} {}/{}{}' % 进度条长度).format(L * 符号 + (进度条长度 - L) * 符号2, f'{int(100 * 百分百)}%', 字符串,
                                                           当前ID, 总数, speed), end='', flush=True)
    elif 类型 == 4:
        print(("\r{:>4} {:<%d} {} {}/{}{}" % 进度条长度).format(f'{int(100 * 百分百)}%', L * '#' + (进度条长度 - L) * '_', 字符串, 当前ID,
                                                           总数, speed), end='', flush=True)
def 打印_列表(列表:list)->None:
    [print(x) for x in 列表]
def 打印_字典(字典:dict,width:int=20)->None:
    [print('{1:>{0}} : {2}'.format(width,repr(k),repr(v))) for k,v in 字典.items()]
def 打印_JSON(json_):
    print(json.dumps(json_,indent=4,ensure_ascii=False))
#特殊功能函数
def 获取_进程名(打印=False):
    pid_dict = {}
    pids = psutil.pids()
    for pid in pids:
        p = psutil.Process(pid)
        pid_dict[pid] = p.name()
        if 打印:
            print("pid:%d\tpname:%s" %(pid,p.name()))
    return pid_dict
def 结束进程_by_id(进程id):
    try:
        kill_pid = os.kill(进程id, signal.SIGABRT)
    except Exception as e:
        pcolor('没有此进程','err')
def 结束进程_by_name(进程名=None):
    dic = 获取_进程名()
    for pid,pname in dic.items():
        if 进程名 and pname == 进程名:
            结束进程_by_id(pid)
def 删除目录树(path,显示错误=True):
    from shutil import rmtree
    path=绝对路径(path)
    try:
        rmtree(path)
    except Exception as e:
        if 显示错误:
            print(f'删除目录树错误 {path}\n{e}')
def 获取_页数(总数, 分页数):
    return 总数 // 分页数 if 总数 % 分页数 == 0 else (总数 // 分页数) + 1
def 字符串缩略(字符串, 位数, 结尾符号='...'):
    '''('xxxx',6,'...')'''
    if len(字符串) <= 位数:
        return 字符串
    else:
        return 字符串[:位数 - len(结尾符号)] + 结尾符号
def 获取_URL_参数(网址: str,解析:bool=True,lower=True) -> dict:
    '''
    lower 参数名全部小写
    解析  urlparse 解析参数
    获取_URL_参数（"a=1&b=2"）
    获取_URL_参数（"http://www.qq.com/a.php?a=1&b=2"）
    获取_URL_参数（"http://www.yixueks.com/ycweb/#/Forum?courseId=1074"） #带#号

    '''
    if 解析:
        网址=unquote(网址)
    query = {}
    #如果链接里面有#号
    param=urlparse(网址.rsplit('#', 1)[-1]).query
    if not param:
        param=网址
    for x in param.split('&'):
        if x:
            if '=' not in x:
                continue
            a = x.split('=',1)
            if lower:
                query[a[0].lower()] = a[1]
            else:
                query[a[0]] = a[1]
    return query
def 获取_URL_HOST(网址:str)->str:
    return urlparse(网址).hostname
def 获取_URL_HTTP(网址:str)->str:
    return urlparse(网址).scheme
def 获取_URL_HTTP_HOST(网址:str)->str:
    p=urlparse(网址)
    return f'{p.scheme}://{p.netloc}'
def 获取_URL_PATH(网址:str)->str:
    p=urlparse(网址)
    return f'{p.scheme}://{p.netloc}{p.path}'
def 获取_URL_QUERY(网址:str,解析:bool=True)->str:
    if 解析:
        网址=unquote(网址)
    param=urlparse(网址).query
    if not param:
        param=网址
    return param
def path_after(网址:str,path:str,解析:bool=True)->dict:
    if 解析:
        网址 = unquote(网址)
    lst = [x for x in urlparse(网址).path.split('/') if x]
    for i,p in enumerate(lst):
        if p.strip():
            if p.strip().lower()==path.lower():
                if i+1<len(lst):
                    return lst[i+1]
def path_before(网址:str,path:str,解析:bool=True)->dict:
    if 解析:
        网址 = unquote(网址)
    lst=[x for x in urlparse(网址).path.split('/') if x]
    for i,p in enumerate(lst):
        if p.strip():
            if p.strip().lower()==path.lower():
                if i-1>=0:
                    return lst[i-1]
def path_fileName(网址:str)->str:
    return urlparse(网址).path.rsplit('/',1)[1]
def 提取m3u8List(字符串:str,关键字='\.m3u8')->list():
    lst=[]
    for row in 字符串.split('\n'):
        a=re.search('^(.*?%s.*?)(\n|$|\s+)'%关键字, row.strip())
        if a:
            lst.append(a.group(1))
    return lst
def 单选(标题, items: list) -> dict:
    '''单选框('xx',[{'i':1,'name':'xxx'},{'i':2,'name':'yyy'}])'''
    while 1:
        try:
            pcolor('【{} 返回0】'.format(标题))
            value = int(input(':'))
            if value == 0:
                return -1
            if 1 <= value <= len(items):
                for x in items:
                    if x['i'] == value:
                        return x
        except:
            pass
def 多选(标题: str, items: list) -> list:
    ''' 多选('xx',[{'i':1,'name':'xxx'},{'i':2,'name':'yyy'}]) i大于等于1 '''
    while 1:
        try:
            ids = [x['i'] for x in items]
            pcolor('【{} 如1-3 1,2,3 返回0 全部all】'.format(标题))
            value = input(':')  # 1,2,3 1-3,all,0
            if value == '0':
                return -1
            elif value.lower() == 'all':
                return items
            else:
                if '-' in value:
                    value = value.split('-', 1)
                    start = int(value[0])
                    end = int(value[1])
                    selected = list(range(start, end + 1, 1))
                else:
                    selected = [int(x.strip()) for x in value.split(',') if x.strip()]
                selected = [items[x - 1] for x in selected if x in ids]
                return selected
        except:
            pass
def 文本对齐(文本:str,长度=20,对齐='L'):
    L=len(文本.encode('GBK'))
    if 对齐.upper()=='R':
        return '{:>{len}}'.format(文本, len=长度 - L + len(文本))
    elif 对齐.upper()=='M':
        return '{:^{len}}'.format(文本, len=长度-L+len(文本))
    else:
        return '{:<{len}}'.format(文本, len=长度 - L + len(文本))
def 获取更新时区(域名='time.windows.com',时区=8):
    '''
    pip install ntplib
    pool.ntp.org
    time.windows.com
    '''
    import ntplib
    c = ntplib.NTPClient()
    response = c.request(域名)
    ts_stamp = response.tx_time
    ts = time.localtime(ts_stamp)
    #ttime = time.localtime(time.mktime(ts) + 8 * 60 * 60)  # +东八区
    return ts
def 设置系统时间(time_str:str='2020-03-04 12:20:30'):
    try:
        import win32api
        if isinstance(time_str, time.struct_time):
            time_str = f'{time_str.tm_year}-{time_str.tm_mon}-{time_str.tm_mday} {time_str.tm_hour}:{time_str.tm_min}:{time_str.tm_sec}'
        elif isinstance(time_str, datetime.datetime):
            time_str = str(time_str)
        time_utc = time.mktime(time.strptime(time_str, '%Y-%m-%d %X'))
        tm_year, tm_mon, tm_mday, tm_hour, tm_min, tm_sec, tm_wday, tm_yday, tm_isdst = time.gmtime(time_utc)
        win32api.SetSystemTime(tm_year, tm_mon, tm_wday, tm_mday, tm_hour, tm_min, tm_sec, 0)
        print('设置时间:{}'.format(time_str))
        return True
    except Exception as e:
        打印错误(e)
        return False
def 设置_添加目录到环境(path):
    '临时搜索路径，程序退出后失效'
    if path:
        sys.path.append(path)
    else:
        sys.path.append(os.path.abspath(os.path.dirname(__file__)))
def 保存二维码图片(文件路径:str, qrcode_url:str,size=8,border=1):
    '''
    pip install qrcode
    :param 文件路径:  保存路径
    :param qrcode_url: 链接地址
    :return: bool
    '''
    文件路径 = 绝对路径(文件路径)
    try:
        import qrcode
        img = qrcode.make(qrcode_url, border=border, box_size=size, error_correction=qrcode.constants.ERROR_CORRECT_H, )
        img.save(文件路径)
        return True
    except Exception as e:
        打印错误(e)
        return False
def 解析二维码链接_pyzbar(文件路径)->str:
    '''pip install pyzbar'''
    文件路径 = 绝对路径(文件路径)
    if 文件路径 and os.path.exists(文件路径):
        from pyzbar import pyzbar
        from PIL import Image
        img=Image.open(文件路径)
        barcodes=pyzbar.decode(img)
        if barcodes:
            return barcodes[0].data.decode()
        else:
            return ''
def 解析二维码链接_wechat(文件路径)->str:
    # pip install opencv-python==4.5.2.54 opencv-contrib-python==4.5.2.54 numpy
    文件路径 = 绝对路径(文件路径)
    from cv2 import wechat_qrcode_WeChatQRCode,imread,imdecode
    import numpy as np
    detector = wechat_qrcode_WeChatQRCode()
    #image = cv2.imread(文件路径)
    image = imdecode(np.fromfile(文件路径, dtype=np.uint8),-1) #中文文件名
    barcodes, points = detector.detectAndDecode(image)  # 使用灰度图像将image换成gray
    if barcodes:
        return barcodes[0]  #返回多个二维码链接
    else:
        return ''
def 获取_屏幕分辨率():
    try:
        import win32print
        import win32gui
        import win32con
        hDC = win32gui.GetDC(0)
        width = win32print.GetDeviceCaps(hDC, win32con.DESKTOPHORZRES)  # 横向分辨率
        height = win32print.GetDeviceCaps(hDC, win32con.DESKTOPVERTRES)  # 纵向分辨率
        return width, height
    except Exception as e:
        打印错误(e)
        return None
def 设置_屏幕分辨率(width=1920,height=1080):
    try:
        import win32con
        import win32api
        import pywintypes
        devmode = pywintypes.DEVMODEType()
        devmode.PelsWidth = width
        devmode.PelsHeight = height
        devmode.Fields = win32con.DM_PELSWIDTH | win32con.DM_PELSHEIGHT
        win32api.ChangeDisplaySettings(devmode, 0)
    except Exception as e:
        打印错误(e)
        return True
def 获取_屏幕缩放比例():
    try:
        import win32api
        real_resolution=获取_屏幕分辨率()
        width = win32api.GetSystemMetrics(0)  # 获得屏幕分辨率X轴
        height = win32api.GetSystemMetrics(1)  # 获得屏幕分辨率Y轴
        screen_size = (width,height)
        screen_scale_rate = round(real_resolution[0] / screen_size[0], 2)
        screen_scale_rate = screen_scale_rate * 100
        return int(screen_scale_rate)
    except Exception as e:
        打印错误(e)
        return None
def 递归目录(列表, 子字段='child', 目录字段='name', names=[], 加编号=False):
    '''
    a=[
    {'name':'学校',
     'list':[
         {'name':'年级1',
          'list':[
              {'name':'班级1', 'list':{},'vid':'xxx1'},
              {'name':'班级2', 'list':[],'vid':'xxx2'},
              {'name':'班级3', 'vid':'xxx3'},
                  ]},
         {'name':'年级2',
          'list': [
              {'name': '班级1', 'list': [], 'vid': 'xxx4'},
              {'name': '班级2', 'list': None, 'vid': 'xxx5'},
              {'name': '班级3', 'list': '', 'vid': 'xxx6'},
          ]
      }]
    }
    ]
    for x in 递归目录(a, 子字段='list', 目录字段='name', 加编号=1):
        print(x)
    '''
    if isinstance(列表, list):
        for i,d in enumerate(列表):
            if d and 子字段 in d and d[子字段]:
                if 加编号:
                    name_ = f'{i + 1}--{d[目录字段]}'
                else:
                    name_ = d[目录字段]
                yield from 递归目录(列表=d[子字段], 子字段=子字段, 目录字段=目录字段, names=names + [检查文件名(name_)], 加编号=加编号)
            else:
                if 加编号:
                    name_ = f'{i+1}--{d[目录字段]}'
                else:
                    name_ = d[目录字段]

                path = '/'.join(names+[检查文件名(name_)])
                yield {'path':path,'data':d}
    else:
        raise Exception('递归目录类型错误')
def 资源文件路径(文件相对路径):
    """
    Get absolute path to resource, works for dev and for PyInstaller
    spec文件  datas=[('a.exe','.')],
    资源路径("a.exe")
    """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath("__file__")))
    return os.path.join(base_path, 文件相对路径)
def 小数(num):
    return Decimal(str(num))
def 加签名(url:str,sign_name='sign',key:bytes or str='',reverse=False):
    ''' url = "https://example.com/path?C=3&d=4&a=1&b=2" '''
    if isinstance(key,str):
        key=key.encode()
    elif not isinstance(key,bytes):
        raise Exception('加签名-->key类型错误')
    parsed_url = urlparse(url)
    query_params = {}
    for s in parsed_url.query.split('&'):
        (k,v)=s.split('=')
        query_params[k]=v
    dic = sorted(query_params.items(), key=lambda x: x[0],reverse=reverse)
    encoded_params = urlencode(dic)
    md5_obj = md5(encoded_params.encode('utf-8'))
    md5_obj.update(key)
    sign=md5_obj.hexdigest()
    query_params[sign_name] = sign
    dic.append((sign_name, sign))
    sorted_query = urlencode(dic)
    signed_url = f"{parsed_url.scheme}://{parsed_url.netloc}{parsed_url.path}?{sorted_query}"
    return signed_url
def params_排序(params:dict,reverse=False,query=True):
    sorted_params = sorted(params.items(),reverse=reverse)
    if query:
        param_str = '&'.join([f'{k}={v}' for k, v in sorted_params])
        return param_str
    else:
        return {k:v for k, v in sorted_params}
def bjcloudvod(url:str):
    '''
    x='bjcloudvod://SWl3eXR1PjI0Z3l3M3dsaWlxMnpqcXxlb3trbmZxMmZ0cDE1PjVlN2g0aGU4ZmVmaGJmNTo0aWg1ODtlOTo4OmYzNzI7NzZlO2Y2aDMyNDB9MHd0cnBkaTN4bWdqcjE2NjU4PDc7PDlkZDRlO2M3PGk4NDhpaWNqPzhoOWY7NTs3OTk1OjM2OGhhdnZ2akg4PXcxcnQ2'
    print(sx.bjcloudvod(x))
    #https://dws-video.wenzaizhibo.com/184b2d2db3ccbbac062ee059a3955b13/644a5e3c/00-x-upload/video/204573986_a2a5b47e605dfaf97e4b91826714233d_rsqgF47v.mp4
    '''
    # 如果前缀不是 bjcloudvod:// 则返回 None
    if not url.startswith("bjcloudvod://"):
        return None
    # 将 '-' 替换为 '+', 将 '_' 替换为 '/' 以保证 base64 字符串可解码
    url = url[13:].replace('-', '+').replace('_', '/')
    # 根据长度填充 '=' 字符
    padding = len(url) % 4
    if padding == 2:
        url += '=='
    elif padding == 3:
        url += '='
    # 将 base64 编码的字符串解码
    plaintext = base64.b64decode(url)
    # 根据解码后的数据进行解密，获取原始视频 URL
    c = plaintext[0] % 8
    ciphertext = plaintext[1:]
    result = []
    for i, char in enumerate(ciphertext):
        step = i % 4 * c + i % 3 + 1
        result.append(chr(char - step))
    return ''.join(result)
def zlib_解压_base64(base64str:str)->str:
    #base64str='eJyFU1my5CAMuxLIGzkO6/2PMDKZrrf8vKS6iTDYsi27unkV8+Li3R9PDBdxMRlf+Lf9g6V+xxJ/nScueLh+8HSNxvhpae6iMqRKjS4WRbYIsQu4TjLaKPLQnWHExo4HBQceFU0U4PnPfUPw3pZCNPn1/z5P1jyFjvATQ0K6LN4q/H9kYKO6R3DH5Ug37vHmsqIVmx7UNhoCnScO32bNTuYdQctinKAvk3wmEvEOHhv8Bf0cHHqu3E9e542nk/6CvuA9IgrZ07sFFqaHL1f66hBW0rC4072FZedQXj4+yKnwPOvDGJcPy3T80GP1puYrNmsPE94tdGTa9Mn4mhEqOwEs3tpQs1DGIDMfrAa7pySEaQ/9TO4h9I1XxIe0yShe2vCpy2128pyxvOvg3qC4TutkW2p/tgtjrBCwKthaM8LlKDA9VxWZ837tNkE6VIbyhEb3fitG7sx/0Y6LwyQrhuwAA2QPXuwvzv6C6+Q7yCQfZsXesseevac2aH/SJ2o2zY27HeuTMSOwEKz+ZBU7LDPGVMpWuxomcxnWKfdyBX9e/tF0q/N7QfmANR8UamOuLcVElPhaXpxVT+xXgrHuojCDUQ2yvVKdoVWzxpuV0yiWM3K0sCsnqk0JS+WSVRx28njhLDn5THrffvBIdrYC7B5n/arZQaWjvhNFxbFF9OBENRWm5c2fCpM7sVkRTgSKb6MtJ5a1E+KWWKZtqxfjJ+ak/LT/wsznG+YMDlux/gG0cBkU'
    _data = base64_to_byte(base64str.encode())
    decompressed_data = zlib.decompress(_data)
    return decompressed_data.decode('utf-8')
def zlib_压缩_base64(decompressed_str:str='test',level=-1)->str:
    decompressed_data=zlib.compress(decompressed_str.encode(),level=level) #Compression level, in 0-9 or -1.
    return base64.b64encode(decompressed_data).decode()
def jwt_base64_payload(jwted:str)->dict:
    #jwt_encode='eyJhbGciOiAiSFMyNTYiLCAidHlwIjogIkpXVCJ9.eyJleHAiOiAxNjIyNjEwMTM4LjU3MTk5MDUsICJpc3MiOiAiSXNzdWVyIiwgImlhdCI6IDE2MjI2MDY1MzguNTcxOTkwNSwgImRhdGEiOiB7InVzZXJuYW1lIjogInhqaiJ9fQ==.NmFjMzMxNmZlNzdhMDBmZTQxMWFjODQxOGVkNDViNzBlZWVmZGJjNDUyMmY3MjkyN2EwMTdlNTEwNTZjYTU4ZQ=='
    return json.loads(base64_to_byte(jwted.split('.')[1].encode()))
def crc32_hex(chunk:bytes):
    import zlib
    prev = zlib.crc32(chunk)
    crc32 = "%X" % (prev & 0xffffffff)
    return crc32
def 随机替换(文本, 替换字, 随机列表=[]):
    '''
    print(随机替换(文本="这是一个【话题】的【话题】 _ 【话题】示例。", 替换字="【话题】",随机列表=[1,2,3,4]))
    #这是一个1的4 _ 2示例。
    '''
    matches = list(re.finditer(re.escape(替换字), 文本))
    matches.reverse()
    rt = 文本
    for match in matches:
        if 随机列表:
            s = random.choice(随机列表)
            index=随机列表.index(s)
            随机列表.pop(index)
            rt = rt[:match.start()] + str(s) + rt[match.end():]
    return rt
def 禁用网卡(网卡名='以太网'):
    '''需要administrator账号执行代码'''
    print(f'禁用网卡 --> {网卡名}')
    os.system(f'netsh interface set interface "{网卡名}" admin=disable')
    time.sleep(0.5)
def 启用网卡(网卡名='以太网'):
    '''需要administrator账号执行代码'''
    print(f'启用网卡 --> {网卡名}')
    os.system(f'netsh interface set interface "{网卡名}" admin=enable')
    time.sleep(0.5)
def unicode_decode(encode_string):
    return re.sub(r'(\\u\w+)', lambda match: match.group(1).encode().decode('unicode-escape'), unquote(encode_string).replace('%u', '\\u'))
def 加载COOKIE(文件路径:str="cookie.txt",字典=False,编码:bool=False)->str or dict:
    cookie_str=加载文件_创建(文件路径)
    try:
        try:
            cookie_list=json.loads(cookie_str)
            # 如果是Cookie-Editor导出
            cookies = {}
            if isinstance(cookie_list,list):
                for cookie in cookie_list:
                    cookies[cookie['name']] = url编码(cookie['value']) if 编码 else cookie['value']
            # 如果是字典的cookie
            elif isinstance(cookie_list,dict):
                cookies=cookie_list
            else:
                raise Exception('类型错误')
        except:
            #如果是非json的字典格式  有单引号那种
            cookies=eval(cookie_str)
        if 字典:
            return cookies #dict
        else:
            return cookie_From_Cookies(cookies) #str
    except:
        pass
    if 字典:
        return dict_From_CookieStr(cookie_str) #dict
    else:
        return cookie_str  #str
# js函数转换
def 执行JS代码(js代码):
    '''执行JS代码('return 123') 无需js环境'''
    from js2py import EvalJs
    sss = '''
    function func(){
        %s
    }
    ''' % js代码
    js = EvalJs()
    js.execute(sss)
    return js.func()
def 执行JS代码_FUNC(js代码:str,函数名:str='',参数:tuple=()):
    '''
    # var func = function(a,b)
    # {return a+b;}
    # var a = function()
    # {return func(1,2);}
    # var b = function(x)
    # {return x+func(1,2);}
    # x=sx.执行JS代码_FUNC(js代码=js,函数名='a')
    # x=sx.执行JS代码_FUNC(js代码=js,函数名='b',参数=(1,))
    # print(x)
    '''
    from js2py import EvalJs
    js=EvalJs()
    js.execute(js代码)
    if isinstance(参数, (list, tuple)):
        return js.__getattr__(函数名)(*参数)
    else:
        return js.__getattr__(函数名)(参数)
def 执行EXECJS(js代码):
    '''执行JS代码('return 123')  需要nodejs环境'''
    from functools import partial
    subprocess.Popen = partial(subprocess.Popen, encoding='utf-8')
    from execjs import compile
    sss = ''' function func(){ %s } ''' % js代码
    js = compile(sss)
    return js.call('func')
def 执行EXECJS_FUNC(js代码:str,函数名:str,参数:tuple=()):
    '''执行EXECJS_FUNC(js代码=code,函数名='h',参数=(111,222))'''
    from functools import partial
    subprocess.Popen = partial(subprocess.Popen, encoding='utf-8')
    from execjs import compile
    if isinstance(参数,(list,tuple)):
        return compile(js代码).call(函数名,*参数)
    else:
        return compile(js代码).call(函数名, 参数)
def js对象转json(js对象字符串:str):
    '''js对象转json('{1:1}') 无需js环境'''
    from js2py import EvalJs
    sss = '''
    function func(){
        var res = %s
        return JSON.stringify(res)
    }
    ''' % js对象字符串
    js = EvalJs()
    js.execute(sss)
    return json.loads(js.func())
def js_Uint8Array(lst:list)->bytes:
    return bytes(lst)
def js_parseInt(a,b):
    return int(a,b)
def js_int8arry_to_uint8arry(lst:list)->list:
    return [x if x>=0 else x+256 for x in lst]
def join(列表:list,分割=''):
    return 分割.join(map(str,列表))
def json_path(josn对象, 表达式, first=True):
    '''
    # 查询store下的所有元素
    print(jsonpath.jsonpath(book_store, '$.store.*'))

    # 获取json中store下book下的所有author值
    print(jsonpath.jsonpath(book_store, '$.store.book[*].author'))

    # 获取所有json中所有author的值
    print(jsonpath.jsonpath(book_store, '$..author'))

    # 获取json中store下所有price的值
    print(jsonpath.jsonpath(book_store, '$.store..price'))

    # 获取json中book数组的第3个值
    print(jsonpath.jsonpath(book_store, '$.store.book[2]'))

    # 获取所有书
    print(jsonpath.jsonpath(book_store, '$..book[0:1]'))

    # 获取json中book数组中包含isbn的所有值
    print(jsonpath.jsonpath(book_store, '$..book[?(@.isbn)]'))

    # 获取json中book数组中price<10的所有值
    print(jsonpath.jsonpath(book_store, '$..book[?(@.price<10)]'))

    # 从根节点开始，匹配name节点
    jsonpath.jsonpath(json_obj, '$..name')

    # A 下面的节点
    jsonpath.jsonpath(json_obj, '$..A.*')

    # A 下面节点的name
    jsonpath.jsonpath(json_obj, '$..A.*.name')

    # C 下面节点的name
    jsonpath.jsonpath(json_obj, '$..C..name')

    # C 下面节点的第二个
    jsonpath.jsonpath(json_obj, '$..C[1]')

    # C 下面节点的第二个的name
    jsonpath.jsonpath(json_obj, '$..C[1].name')

    # C 下面节点的2到5的name
    jsonpath.jsonpath(json_obj, '$..C[1:5].name')

    # C 下面节点最后一个的name
    jsonpath.jsonpath(json_obj, '$..C[(@.length-1)].name')
    '''
    try:
        if first:
            return jsonpath.jsonpath(josn对象,表达式)[0]
        else:
            return jsonpath.jsonpath(josn对象,表达式)
    except:
        if first:
            return None
        else:
            return []
def hexXor(a:str,b:str)->str:
    '''
    a 和 b 都是十六进制字符串 异或操作
    a='547F1137EB0911475B97A0A8ED13DA58EEA2AFDF'
    b='3000176000856006061501533003690027800375'
    a["hoxXor"](b)
    '''
    return hex(int(a, 16) ^ int(b, 16))[2:]
# 请求
def get_proxies(ip_port:str='')->dict:
    '''
    自定义 代理IP  '127.0.0.1:8080'
    或者获取本地代理
    返回列表{} 或者 None
    '''
    try:
        if ip_port:
            return {'http': f'http://{ip_port}', 'https': f'http://{ip_port}'}
        else:
            hKey = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                  "Software\Microsoft\Windows\CurrentVersion\Internet Settings", 0, winreg.KEY_READ)
            retVal = winreg.QueryValueEx(hKey, "ProxyEnable")
            开启 = retVal[0]
            if 开启:
                res = winreg.QueryValueEx(hKey, "ProxyServer")
                lst = res[0].split(';')
                lst = list(set([x.split('=', 1)[1] if '=' in x else x for x in lst]))
                return {'http': f'http://{lst[0]}', 'https': f'http://{lst[0]}'} if lst else None
            else:
                return None
    except:
        return None
def set_proxies(ip_port='127.0.0.1:8080',开启代理=1,白名单="")->None:
    '''
    :param 开启代理: 1 或者 0
    :param ip_port: 127.0.0.1:8080;127.0.0.1:8888
    :param 白名单: 127.*;10.*;172.16.*;
    :return:
    '''
    hKey = winreg.OpenKey(winreg.HKEY_CURRENT_USER, "Software\Microsoft\Windows\CurrentVersion\Internet Settings", 0, winreg.KEY_WRITE)
    winreg.SetValueEx(hKey, "ProxyEnable", 0, winreg.REG_DWORD, 开启代理)
    winreg.SetValueEx(hKey, "ProxyServer", 0, winreg.REG_SZ, ip_port)
    winreg.SetValueEx(hKey, "ProxyOverride", 0, winreg.REG_SZ, 白名单)
    winreg.CloseKey(hKey)
default_proxies=get_proxies()
def add_headers(url:str=None,headers:str=None,浏览器:str=None,fake_useragent:bool=False)->dict:
    if headers:
        if isinstance(headers,str):
            headers=dict_From_HeadersStr(headers)
    else:
        if headers == None:
            pass
        else:
            p = urlparse(url)
            headers = {}
            headers['user-agent']='Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.71 Safari/537.36 Core/1.94.202.400 QQBrowser/11.9.5355.400'
            if url:
                headers['Referer']=f'{p.scheme}://{p.netloc}'
                headers['Origin']=f'{p.scheme}://{p.netloc}'
    if fake_useragent:
        headers['user-agent']=get_fake_agent(浏览器)
    return headers
def get_request(url, headers=None,params=None, verify=False, proxies=None, allow_redirects=True, cookies=None, stream=False,curl=False, timeout=30,fake_useragent=False, try_num=1):
    proxies = proxies if proxies else default_proxies
    headers = add_headers(url, headers, 浏览器='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str):cookies=dict_From_CookieStr(cookies)
    if params and isinstance(params,str):params=dict_From_HeadersStr(params)
    if curl and 'cookie' in headers and not cookies:
        cookies=dict_From_CookieStr(headers['cookie'])
    for i in range(try_num):
        try:
            if stream:
                return requests.get(url.strip(), timeout=timeout, headers=headers if headers else get_headers('chrome'),params=params, verify=verify, proxies=proxies, allow_redirects=allow_redirects, cookies=cookies, stream=stream)
            else:
                with requests.get(url.strip(), timeout=timeout, headers=headers if headers else get_headers('chrome'),params=params, verify=verify, proxies=proxies, allow_redirects=allow_redirects, cookies=cookies, stream=stream) as resp:
                    return resp
        except Exception as e:
            if i == try_num - 1:
                raise Exception(跟踪函数(-3)+' '+str(e))
def post_request(url, headers=None, data=None, verify=False, proxies=None, allow_redirects=True, cookies=None,stream=False, json_=None,curl=False, timeout=30,fake_useragent=False, try_num=1):
    proxies = proxies if proxies else default_proxies
    headers = add_headers(url, headers, 浏览器='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str): cookies=dict_From_CookieStr(cookies)
    if curl and 'cookie' in headers and not cookies:
        cookies=dict_From_CookieStr(headers['cookie'])
    for i in range(try_num):
        try:
            if stream:
                return requests.post(url.strip(), timeout=timeout, headers=headers if headers else get_headers('chrome'), verify=verify, proxies=proxies, allow_redirects=allow_redirects, data=data, json=json_, cookies=cookies,stream=stream)
            else:
                with requests.post(url.strip(), timeout=timeout, headers=headers if headers else get_headers('chrome'), verify=verify, proxies=proxies, allow_redirects=allow_redirects, data=data, json=json_, cookies=cookies,stream=stream) as resp:
                    return resp
        except Exception as e:
            if i == try_num - 1:
                raise Exception(跟踪函数(-3)+' '+str(e))
def curl_bash(curl_bash:str,code:bool=False)->requests.Response:
    '''copy->copy as cURL(bash)'''
    import uncurl
    rt_code = uncurl.parse(curl_bash)
    if default_proxies:
        lst=rt_code.split('\n')
        lst.insert(1,f'proxies={default_proxies},')
        rt_code='\n'.join(lst)
    if code:
        return rt_code
    else:
        return eval(rt_code)
def 获取_网络文件大小(网址,headers=None,params=None,proxies=None,cookies=None,verify=False,allow_redirects=True,fake_useragent=False)->int:
    proxies = proxies if proxies else default_proxies
    headers = add_headers(网址, headers, 浏览器='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str):cookies=dict_From_CookieStr(cookies)
    if params and isinstance(params,str):params=dict_From_HeadersStr(params)
    try:
        with requests.get(网址, stream=True, headers=headers if headers else get_headers('chrome'),params=params, proxies=proxies, cookies=cookies, verify=verify,allow_redirects=allow_redirects) as resp:
            return resp.headers['content-length']
    except Exception as e:
        打印错误(e)
    return 0
def 下载文件(文件路径: str = None, 网址: str = '', headers=None,params=None, proxies=None, verify=False, allow_redirects=True, cookies=None, fake_useragent=False,try_num=3) -> int:
    proxies = proxies if proxies else default_proxies
    headers = add_headers(网址, headers, 浏览器='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str):cookies=dict_From_CookieStr(cookies)
    if params and isinstance(params, str): params = dict_From_HeadersStr(params)

    if not 文件路径:
        文件路径 = 网址.rsplit('/', 1)[-1]
    文件路径 = 绝对路径(文件路径)
    if os.path.exists(文件路径):
        os.remove(文件路径)
    for i in range(try_num):
        try:
            with requests.get(网址.strip(), timeout=15, verify=verify, proxies=proxies, allow_redirects=allow_redirects, headers=headers if headers else get_headers('chrome'), params=params,cookies=cookies) as res:
                if res.status_code == 200:
                    res = res.content
                    with open(文件路径, 'wb') as f:
                        f.write(res)
                    return len(res)
                else:
                    if str(res.status_code) in http_err_code.keys():
                        raise Exception(res.status_code, ','.join(http_err_code[str(res.status_code)].values()))
                    else:
                        raise Exception('下载文件失败')
        except Exception as e:
            if i == (try_num - 1):
                pcolor('下载文件错误:{},{}'.format(e.args, e.__traceback__.tb_lineno), 'error')
    return 0
def 下载文件_进度条(文件路径: str = None, 网址: str = '', 分段长度: int = 5*1024, 多线程=False, 线程数=5, headers=None,params=None, proxies=None,allow_redirects=True, verify=False, cookies=None, 进度条函数=None, 打印错误=True,fake_useragent=False, try_num=2) -> int:  # 分段长度 kb
    '''覆盖已存在的文件'''
    if not 文件路径:
        文件路径 = 网址.rsplit('/', 1)[-1]
    文件路径 = 绝对路径(文件路径)
    proxies = proxies if proxies else default_proxies
    headers = add_headers(网址, headers, 浏览器='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str):cookies=dict_From_CookieStr(cookies)
    if params and isinstance(params, str): params = dict_From_HeadersStr(params)
    headers = headers if headers else get_headers('chrome')
    print('[ {} ] : {}'.format(scolor('下载文件', 'warn'), scolor(文件路径, 'yes')))
    for num in range(try_num):
        if not 文件路径:
            文件路径 = 网址.rsplit('/', 1)[-1]
        if os.path.exists(文件路径):
            os.remove(文件路径)
        try:
            with requests.get(网址.strip(), stream=True, headers=headers,params=params, proxies=proxies, cookies=cookies, verify=verify,allow_redirects=allow_redirects) as resp:
                if 'content-length' in resp.headers:
                    size = int(resp.headers['content-length'])
                    chunk_size = 1024 * 分段长度  # 分段接收
                    c = size / chunk_size
                    total_size = c / 1024 * 分段长度
                    count = int(size / chunk_size) if size % chunk_size == 0 else int(size / chunk_size) + 1
                    start_time = time.time()
                    if 多线程:
                        n = 0
                        flag_error=False
                        def get_one(网址, str_range):
                            nonlocal n
                            nonlocal flag_error
                            if flag_error:
                                return None
                            head = headers.copy()
                            head['range'] = str_range
                            for i in range(try_num):
                                try:
                                    with requests.get(网址.strip(), stream=True, headers=head, proxies=proxies,params=params, cookies=cookies, verify=verify,allow_redirects=allow_redirects) as resp:
                                        content = resp.content
                                        if 进度条函数:
                                            进度条函数(int((n + 1) / count * 100), '{:.2f}mb'.format(total_size))
                                        t=time.time()-start_time
                                        seconds = t * count / (i + 1) - t
                                        if seconds<1:
                                            剩余时间 = ' '*12
                                        else:
                                            m, s = divmod(seconds, 60)
                                            h, m = divmod(m, 60)
                                            剩余时间 = ' {:0=2.0f}:{:0=2.0f}:{:0=2.0f}'.format(h,m,s)
                                        speed = '{:.2f}MB/S{}'.format(
                                            (n + 1) * chunk_size / 1024 / 1024 / t,剩余时间)
                                        lock.acquire()
                                        打印_进度条('{:.2f}MB'.format(total_size), n, count, 1, 下载速度=speed)  # 步长1
                                        n += 1
                                        lock.release()
                                        return content
                                except:
                                    pass
                            flag_error=True
                        pool = ThreadPoolExecutor(max_workers=线程数)
                        tasks = []
                        for i in range(count):
                            if i == count - 1:
                                r = 'bytes={}-'.format(chunk_size * i)
                            else:
                                r = 'bytes={}-{}'.format(chunk_size * i, chunk_size * (i + 1) - 1)
                            tasks.append(pool.submit(get_one, 网址, r))
                        pool.shutdown(wait=True)
                        if not flag_error:
                            with open(文件路径, mode='wb') as f:
                                for task in tasks:
                                    f.write(task.result())
                        del tasks
                    else:
                        with open(文件路径, 'wb') as f:
                            for i, content in enumerate(resp.iter_content(chunk_size=chunk_size)):
                                f.write(content)
                                if 进度条函数:
                                    进度条函数(int((i + 1) / count * 100), '{:.2f}mb'.format(total_size))
                                t=time.time() - start_time
                                seconds=t * count / (i + 1) - t
                                if seconds<1:
                                    剩余时间 = ' '*12
                                else:
                                    m, s = divmod(seconds, 60)
                                    h, m = divmod(m, 60)
                                    剩余时间 = ' {:0=2.0f}:{:0=2.0f}:{:0=2.0f}'.format(h,m,s)
                                speed = '{:.2f}MB/S{}'.format((i + 1) * chunk_size / 1024 / 1024 / t,剩余时间)
                                打印_进度条('{:.2f}MB'.format(total_size), i, count, 1, 下载速度=speed)  # 步长1
                    print()
                    return size
                elif 'Content-Disposition' in resp.headers:
                    size=len(resp.content)
                    print('[ {} ] : {} MB'.format(scolor('文件大小', 'warn'), round(size / 1024 / 1024, 2)))
                    with open(文件路径, 'wb') as f:
                        f.write(resp.content)
                    return size
                else:
                    return 0
        except Exception as e:
            if num == (try_num - 1):
                if 打印错误:
                    pcolor('下载文件错误:{},{}'.format(e.args, e.__traceback__.tb_lineno), 'error')
    print()
    return 0
def m3u8DL_CLI(文件路径:str='test',网址:str="",headers:str or dict={},key:bytes=None,iv:bytes=None,cli_href:str="",ffmpeg_href:str="",cli_local:str="",ffmpeg_local:str="",istest:bool=False,options:list=[])->bool:
    '''
    文档地址 https://nilaoda.github.io/N_m3u8DL-CLI/Advanced.html
    --workDir    Directory      设定程序工作目录
    --saveName   Filename       设定存储文件名(不包括后缀)
    --baseUrl    BaseUrl        设定Baseurl
    --headers    headers        设定请求头，格式 key:value 使用|分割不同的key&value
    --maxThreads Thread         设定程序的最大线程数(默认为32)
    --minThreads Thread         设定程序的最小线程数(默认为16)
    --retryCount Count          设定程序的重试次数(默认为15)
    --timeOut    Sec            设定程序网络请求的超时时间(单位为秒，默认为10秒)
    --muxSetJson File           使用外部json文件定义混流选项
    --useKeyFile File           使用外部16字节文件定义AES-128解密KEY
    --useKeyBase64 Base64String 使用Base64字符串定义AES-128解密KEY
    --useKeyIV     HEXString    使用HEX字符串定义AES-128解密IV
    --downloadRange Range       仅下载视频的一部分分片或长度
    --liveRecDur HH:MM:SS       直播录制时，达到此长度自动退出软件
    --stopSpeed  Number         当速度低于此值时，重试(单位为KB/s)
    --maxSpeed   Number         设置下载速度上限(单位为KB/s)
    --proxyAddress http://xx    设置HTTP代理, 如 ["--proxyAddress", "http://127.0.0.1:8080"] ["--noProxy"]
    --enableDelAfterDone        开启下载后删除临时文件夹的功能
    --enableMuxFastStart        开启混流mp4的FastStart特性
    --enableBinaryMerge         开启二进制合并分片
    --enableParseOnly           开启仅解析模式(程序只进行到meta.json)
    --enableAudioOnly           合并时仅封装音频轨道
    --disableDateInfo           关闭混流中的日期写入
    --noMerge                   禁用自动合并
    --noProxy                   不自动使用系统代理
    --disableIntegrityCheck     不检测分片数量是否完整

    # commands=['C:\\Users\\Administrator\\FFMPEG\\m3u8dl.exe', 'https://hls.videocc.net/source/24560c93d4/d/24560c93d4c855d66ab155af0db215d1_1.m3u8', '--enableDelAfterDone', '--workDir', 'E:\\pytho
    # n_project\\下载合并视频\\我的下载器\\优酷视频_解码\\Downloads', '--saveName', 'aaa.mp4', '--downloadRange', '-9', '--headers', 'user-agent:Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWe
    # bKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.71 Safari/537.36 Core/1.94.201.400 QQBrowser/11.9.5325.400|accept:text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,ima
    # ge/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9|accept-encoding:gzip, deflate, br|accept-language:zh-CN,zh;q=0.9|cache-control:no-cache|pragma:no-cache|sec-ch-ua:"
    # ;Not A Brand";v="99", "Chromium";v="94"|sec-ch-ua-mobile:?0|sec-ch-ua-platform:"Windows"|sec-fetch-dest:document|sec-fetch-mode:navigate|sec-fetch-site:same-origin|sec-fetch-user:?1|u
    # pgrade-insecure-requests:1']

    '''
    文件路径 = 绝对路径(文件路径)
    save_path=文件路径.rstrip('.mp4')
    real_path=save_path+'.mp4'
    if os.path.exists(real_path):
        print('已存在 {}'.format(real_path))
        return True
    workDir,saveName=os.path.split(save_path)
    if not workDir:
        workDir=os.path.join(os.getcwd(),'Downloads')
    path = os.path.join(os.path.expanduser('~'), 'FFMPEG')
    os.makedirs(path, exist_ok=1)
    cli_path = cli_local if cli_local else os.path.join(path, "m3u8dl.exe")
    ffmpeg_path = ffmpeg_local if ffmpeg_local else os.path.join(path, "ffmpeg.exe")
    if not os.path.exists(cli_path):
        if not cli_href:
            cli_href=CLI_HREF
        下载文件_进度条(文件路径=cli_path, 网址=cli_href)
    if not os.path.exists(ffmpeg_path):
        if not ffmpeg_href:
            ffmpeg_href=FFMPEG_HREF
        下载文件_进度条(文件路径=ffmpeg_path, 网址=ffmpeg_href)
    workDir = os.path.join(os.getcwd(), workDir)
    commands=[cli_path,网址,"--enableDelAfterDone","--workDir",workDir,"--saveName",saveName]
    if istest and ('--downloadRange' not in options):
        commands+=['--downloadRange', '-9']
    commands+=options
    if headers and ('--headers' not in commands):
        if isinstance(headers,str):
            headers=dict_From_HeadersStr(headers)
        head=[]
        for k,v in headers.items():
            head.append(f'{k}:{v}')
        head='|'.join(head)
        commands+=['--headers',head]
    if key and ('--useKeyBase64' not in commands):
        commands += ['--useKeyBase64',byte_to_base64(key).decode()]
    if iv and ('--useKeyIV' not in commands):
        commands += ['--useKeyIV', byte_to_base64(iv).decode()]
    ''' CREATE_NEW_CONSOLE：在新控制台窗口中启动子进程。 '''
    returncode=subprocess.run(commands,creationflags=CREATE_NEW_CONSOLE).returncode
    if returncode==0:
        return True
    else:
        return False
def 获取_网络图片(图片网址: str, headers=None,params=None, proxies=None, cookies=None, verify=False,allow_redirects=True, pil=True, show=False,fake_useragent=False, try_num=3):
    '''获取网络图片("http://...")'''
    proxies = proxies if proxies else default_proxies
    headers = add_headers(图片网址, headers, 浏览器='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str):cookies=dict_From_CookieStr(cookies)
    if params and isinstance(params, str): params = dict_From_HeadersStr(params)
    for num in range(try_num):
        try:
            with requests.get(图片网址.strip(), headers=headers if headers else get_headers('chrome'),params=params, proxies=proxies, cookies=cookies, verify=verify,allow_redirects=allow_redirects) as res:
                if res.status_code == 200:
                    # 返回本地图片内存对象
                    # from PIL import Image
                    # img=Image.open(obj)  打开图片
                    # Image._show(img) 显示图片
                    img = BytesIO(res.content)
                    from PIL import Image
                    if pil:
                        img = Image.open(img)
                        if show:
                            Image._show(img)
                        return img
                    else:
                        return img
                else:
                    raise Exception('获取网络图片错误')
        except Exception as e:
            if num == (try_num - 1):
                raise Exception('{},{}'.format(e.args, e.__traceback__.tb_lineno))
def 获取_网络文件(文件网址: str, headers=None,params=None, proxies=None, cookies=None, verify=False,allow_redirects=True,fake_useragent=False, try_num=3) -> bytes:
    proxies = proxies if proxies else default_proxies
    headers = add_headers(文件网址, headers, 浏览器='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies, str): cookies = dict_From_CookieStr(cookies)
    if params and isinstance(params, str): params = dict_From_HeadersStr(params)
    for num in range(try_num):
        try:
            with requests.get(文件网址.strip(), headers=headers if headers else get_headers('chrome'),params=params, proxies=proxies, cookies=cookies, verify=verify,allow_redirects=allow_redirects) as res:
                if res.status_code == 200:
                    return res.content
                else:
                    raise Exception('获取网络文件错误')
        except Exception as e:
            if num == (try_num - 1):
                raise Exception('{},{}'.format(e.args, e.__traceback__.tb_lineno))
def 获取_IP信息(ip:str=None,proxies=None, timeout=30, try_num=3) -> json:
    proxies = proxies if proxies else default_proxies
    for num in range(try_num):
        try:
            # api_ips=[
            #     'http://ip.42.pl/raw',
            #     'http://icanhazip.com',
            #     'http://ifconfig.me/ip',
            #     'http://ipinfo.io/ip',
            # ]
            # http://ipinfo.io/json
            # http://ip-api.com/json
            if ip:
                return requests.get(f'http://ipinfo.io/{ip}/json', timeout=timeout, proxies=proxies, verify=False).json()
            else:
                return requests.get('http://ipinfo.io/json', timeout=timeout, proxies=proxies, verify=False).json()
        except Exception as e:
            if num == (try_num - 1):
                raise Exception('获取IP错误,{},{}'.format(e.args, e.__traceback__.tb_lineno))
def 获取_IP(proxies=None)->str:
    '''返回外网ip地址'''
    api_ips=[
        'http://ip.42.pl/raw',
        'http://icanhazip.com',
        'http://ifconfig.me/ip',
        'http://ipinfo.io/ip',
    ]
    for url in set(api_ips):
        req=get_request(url,proxies=proxies)
        if req.status_code==200:
            return req.text
def 多线程运行(运行函数,参数列表,回调函数=None, 线程数=10, 异步=True):
    pool = ThreadPoolExecutor(max_workers=线程数)
    if 异步:
        for 参数 in 参数列表:
            if 回调函数:
                pool.submit(运行函数, 参数).add_done_callback(lambda x: 回调函数(x.result()))  # 异步写入 无锁
            else:
                pool.submit(运行函数, 参数)  # 异步无锁
        pool.shutdown(wait=True)
    else:
        # 同步
        tasks = [pool.submit(运行函数, 参数) for 参数 in 参数列表]  # 异步爬取 无锁
        pool.shutdown(wait=True)
        if 回调函数:
            for task in tasks:
                回调函数(task.result())  # 同步写入
def 异步函数(func, *args, **kwargs):
    return asyncio.get_event_loop().run_until_complete(func(*args, **kwargs))
def pool_fetures(func,args,max_workers=10,callback=None):
    '''异步运行同步输出'''
    from concurrent import futures
    tasks=[]
    with futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        for i,arg in enumerate(args):
            if callback:
                task=executor.submit(func,arg).add_done_callback(callback)
            else:
                task = executor.submit(func, arg)
            tasks.append(task)
    result=[]
    for i,task in enumerate(tasks):
        if task.done():
            result.append(task.result())
    return result
def t(n:int=1000)->str:
    return str(int(time.time()*n))
def py2pyd(fname):
    r'''
    注意调用需要导入相应模块 否则找不到模块
    Microsoft Visual C++ 14.0 or greater is required. Get it with "Microsoft C++ Build Tools": https://visualstudio.microsoft.com/visual-cpp-build-tools/
    需要安装vs编译器C:\\Program Files (x86)\\Microsoft Visual Studio\\2022\\BuildTools\\VC\\Tools\\MSVC\\14.34.31933\\bin\\HostX86\\x64\\cl.exe
    安装win10 或者 win11 sdk
    C:\Program Files (x86)\Windows Kits\10\Include\10.0.17763.0\ucrt里的io.h文件
    复制到python的include目录中 也就是说 c:\softwares\anaconda3\include中。
    '''
    (path,name)=os.path.split(fname)
    cname=name[:-3]+'.c'
    cname=os.path.join(path,cname)
    with open('setup.py', 'w', encoding='utf-8-sig') as f:  # 自动生成单独的setup.py文件
        f.write('# encoding: utf-8\n')
        f.write('from setuptools import setup\n')
        f.write('from Cython.Build import cythonize\n')
        f.write('setup(\n')
        f.write("name='test',\n")
        f.write(f'ext_modules=cythonize(r"{fname}")\n')
        f.write(")\n")
    os.system('python setup.py build_ext --inplace')  # py编译开始
    if os.path.exists(cname):
        os.remove(cname)
def pypi_recent(pack_name='spiderx'):
    return get_request(f'https://pypistats.org/api/packages/{pack_name}/recent',proxies=default_proxies,verify=False).json()
# 实例类
class XPATH(etree.ElementBase):
    def __init__(self, html: str,显示不可见元素=True):
        '''类型str bytes etree._Element  display True 显示左右  False只显示可见'''
        if type(html) == str or type(html) == bytes:
            self.xp = etree.HTML(html)
        elif type(etree._Element):
            self.xp = etree.ElementTree(html)
        else:
            raise Exception('xpath输入类型错误:{}'.format(type(html)))
        if not 显示不可见元素: #只显示可见元素
            self.删除不可见元素()

    def title(self) -> str:
        return self.xp.xpath('normalize-space(string(//title//text()))')

    def 取首元素(self, 表达式: str, xpath=False, 去换行=False) -> str or etree._Element:
        '''
        第一个元素 //title
        第一个元素属性 //title/text()
        xpath 是否返回XPATH对象
        '''
        元素 = self.xp.xpath(表达式)
        if 元素:
            if 去换行:
                return self.取文本去掉换行(元素[0])  # 返回字符串
            elif xpath:
                return self.load(元素[0])
            else:
                return 元素[0]  # 返回元素
        else:
            return None

    def 取尾元素(self, 表达式: str, xpath=False, 去换行=False):
        元素 = self.xp.xpath(表达式)
        if 元素:
            if 去换行:
                return self.取文本去掉换行(元素[-1])  # 返回字符串
            elif xpath:
                return self.load(元素[-1])
            return 元素[-1]
        else:
            return None

    def 取多个元素(self, xpath_str: str, xpath=False , 去换行=False):
        '''
        多个元素 //img
        多个属性 //img/@src
        '''
        元素列表 = self.xp.xpath(xpath_str)
        if 去换行:
            return self.取文本去掉换行(元素列表)  # 返回字符串列表
        elif xpath:
            return [self.load(x) for x in 元素列表]
        else:
            return 元素列表

    def 替换换行(self, 元素: etree._Element = None) -> None:
        '''修改br 和给p加\n'''
        if not 元素:
            元素 = self.xp
        for p in 元素.xpath("//p"):
            if p.tail is not None:
                p.tail = p.tail + "\n"
            else:
                p.tail = "\n"
        for br in 元素.xpath("//br"):
            br.text = "\n"

    def 取文本(self, 元素: etree._Element = None, 拼接符='') -> str:
        if not 元素:
            body = self.xp.xpath('//body')
            if body:
                元素 = body[0]
            else:
                元素 = self.xp

        #处理表格
        tables = 元素.xpath('//table')
        if tables:
            tables.reverse()
            for table in tables:
                data = self.取表格(table, 去换行=1)
                new_tag = etree.Element("new_tag")
                new_tag.text = '\n'.join(['\t'.join([cell for cell in row]) for row in data])
                table.getparent().replace(table, new_tag)

        return 拼接符.join(元素.xpath('.//text()'))

    def 取文本_换行(self, 元素: etree._Element = None, 拼接符='') -> str:
        if not 元素:
            body = self.xp.xpath('//body')
            if body:
                元素 = body[0]
            else:
                元素 = self.xp

        #处理表格
        tables = 元素.xpath('//table')
        if tables:
            tables.reverse()
            for table in tables:
                data = self.取表格(table, 去换行=1)
                new_tag = etree.Element("new_tag")
                new_tag.text = '\n'.join(['\t'.join([cell for cell in row]) for row in data])
                table.getparent().replace(table, new_tag)

        self.替换换行(元素)
        return 拼接符.join(元素.xpath('.//text()'))

    def 取文本去掉换行(self, 表达式或元素: str or etree._Element) -> str or etree._Element:
        '''
        表达式 //a
        多个元素 出入多个元素
        单个元素
        '''
        a = 表达式或元素
        if type(a) == list:
            return [x.xpath('normalize-space(.)') for x in a]
        elif type(a) == etree._Element:
            return a.xpath('normalize-space(.)')
        elif type(a) == str:
            a = self.xp.xpath(a)
            return [x.xpath('normalize-space(.)') if type(x) == etree._Element else x for x in a]

    def 取正则查询元素(self, 表达式: str, 属性: str, 正则语句: str) -> list:
        s = '{}[re:match({},"{}")]'.format(表达式, 属性, 正则语句)
        return self.xp.xpath(s, namespaces={"re": "http://exslt.org/regular-expressions"})

    def 模糊查询元素(self, 表达式: str, dic: dict) -> list:
        '''
        xp.模糊查询('//a', {'text()':'美女','@src':True,'@data':False})
        '''
        s = []
        for k, v in dic.items():
            s.append('contains({},{})'.format(k, v if type(v) == bool else "\"{}\"".format(v)))
        s = ' and '.join(s)
        表达式 = '{}[{}]'.format(表达式, s)
        return self.xp.xpath(表达式)

    def 元素集取属性(self, 元素集: list, 表达式: str, 取首个: bool = True) -> list:
        if 取首个:
            return [x.xpath('string({})'.format(表达式)) for x in 元素集]
        else:
            return [x.xpath(表达式) for x in 元素集]

    def 取同级下个元素(self, 元素: etree._Element, 标签: str, N=1) -> etree._Element:
        表达式 = 'following-sibling::{}[{}]'.format(标签, N)
        res = 元素.xpath(表达式)
        return res[0] if res else None

    def 取同级上个元素(self, 元素: etree._Element, 标签: str, N=1) -> etree._Element:
        表达式 = 'preceding-sibling::{}[{}]'.format(标签, N)
        res = 元素.xpath(表达式)
        return res[0] if res else None

    def 取HTML(self, 元素: etree._Element, 编码='utf-8') -> str:
        return etree.tostring(元素, encoding=编码).decode(编码)

    def 删除不可见元素(self,元素: etree._Element=None)->None:
        '''删除一些 display 信息'''
        if 元素:
            lst= 元素.xpath('//*[re:match(@style,"{}")]'.format("display[\s]*:[\s]*none"), namespaces={"re": "http://exslt.org/regular-expressions"})
        else:
            lst= self.xp.xpath('//*[re:match(@style,"{}")]'.format("display[\s]*:[\s]*none"), namespaces={"re": "http://exslt.org/regular-expressions"})
        [elem.getparent().remove(elem) for elem in lst]

    def 删除标签(self,标签:str or list =['style','script'],元素:etree._Element=None)->None:
        '''默认删除style | script'''
        if not 元素:
            元素=self.xp
        tags=[]
        if isinstance(标签,list):
            tags=元素.xpath('|'.join([".//"+x for x in 标签]))
        if isinstance(标签,str):
            tags=元素.xpath('.//{}'.format(标签))
        for tag in tags:
            tag.getparent().remove(tag)

    @zsq_try
    def 取表格(self, 表格或表达式: etree._Element, 列: list = [],显示不可见元素=True, 去换行=False,列名=True) -> list:
        ''' （'//table',[1,2]）'''
        table = []
        if type(表格或表达式)==str:
            元素 = self.取首元素(表格或表达式)
        else:
            元素 = 表格或表达式
        if not 显示不可见元素:
            self.删除不可见元素(元素)
        if 列名:
            ths=元素.xpath('.//th')
            if ths:
                if 去换行:
                    ths = [self.取文本去掉换行(x) for x in ths]
                table.append(ths)
        trs = 元素.xpath('.//tr')
        if trs:
            for tr in trs:
                tds = tr.xpath('.//td')
                if tds:  # 排除空的
                    if 去换行:  # 取文本
                        tds = [self.取文本去掉换行(x) for x in tds]
                    table.append(tds)
        else:
            tds = 元素.xpath('.//td')
            if tds:  # 排除空的
                if 去换行:  # 取文本
                    tds = [self.取文本去掉换行(x) for x in tds]
                table.append(tds)
        if 列:
            return [[y for i, y in enumerate(x) if i in 列] for x in table]
        return table

    def xpath(self, *args, **keyargs):
        return self.xp.xpath(*args, **keyargs)

    def load(self, 元素: etree._Element):
        '''返回etree对象'''
        return XPATH(元素)
class RUNTIME():
    def __init__(self):
        pass

    def start(self):
        print('<<<' + '-' * 6)
        self.t1 = time.time()

    def end(self):
        self.t2 = time.time()
        print('-' * 6 + '>>>{:.5f}秒'.format(self.t2 - self.t1))
class MAIL():
    def __init__(self):
        self.服务器 = "smtp.qq.com"  # 设置服务器
        self.用户名 = "wgnms@qq.com"  # 用户名
        self.密码 = ""  # 第三方密码
        self.发件人 = 'wgnms@qq.com'  # 发件人
        self.收件人 = ['758000298@qq.com', 'wgnms@qq.com']  # 收件人
        self.附件 = None  # 文件绝对路径

    def send_mail(self, 标题="邮件测试标题", 邮件内容='邮件发送测试内容', 网页=True):
        try:
            if 网页:
                message = MIMEText(邮件内容, 'html', 'utf-8')
            else:
                message = MIMEText(邮件内容, 'plain', 'utf-8')
            message['Subject'] = 标题
            message['From'] = self.发件人
            message['To'] = ','.join(self.收件人)
            smtpObj = SMTP()
            smtpObj.connect(self.服务器, 25)  # 25 为 SMTP 端口号
            smtpObj.login(self.用户名, self.密码)
            smtpObj.sendmail(self.发件人, self.收件人, message.as_string())
            print("邮件发送成功 {}".format(','.join(self.收件人)))
            return True
        except BaseException as e:
            print("Error: 邮件发送失败")
            return False
class SESSION():
    def __init__(self, cookiePath='cookie.txt'):
        self.sess = requests.session()
        if os.path.exists(cookiePath):
            self.cookiePath = cookiePath
            self.sess.cookies = cookiejar.LWPCookieJar(cookiePath)
            try:
                # 加载cookie文件，ignore_discard = True,即使cookie被抛弃，也要保存下来
                self.sess.cookies.load(ignore_expires=True, ignore_discard=True)
            except:
                pass
        else:
            self.cookiePath = None

    def get(self, *args, **kwargs):
        resp = self.sess.get(*args, **kwargs)
        if self.cookiePath:
            self.sess.cookies.save()
        return resp

    def post(self, *args, **kwargs):
        resp = self.sess.post(*args, **kwargs)
        if self.cookiePath:
            self.sess.cookies.save()
        return resp
def userDataDir_创建(用户名=None,目录='user_data'):
    目录 = 绝对路径(目录)
    if not 用户名:
        i=0
        while True:
            path = os.path.join(目录, f'user_{i}')
            if not os.path.exists(path):
                用户名=f'user_{i}'
                break
            i+=1
    path = os.path.join(目录, 用户名)
    os.makedirs(path,exist_ok=1)
    return path
def userDataDir_读取(用户名=None,目录='user_data'):
    目录=绝对路径(目录)
    os.makedirs(目录,exist_ok=True)
    if 用户名:
        return os.path.join(目录,用户名)
    lst=os.listdir(目录)
    return os.path.join(目录,random.choice(lst)) if lst else None
class PYPP():
    host = ''
    port = ''
    user=''
    password=''
    def __init__(self, headless=False, executablePath=None, width=1200, height=800, userDataDir=None, 启用拦截器=False,
                 timeout=20):
        '''
        :param headless:
        :param executablePath:浏览器exe文件路径
        :param width:
        :param height:
        :param userDataDir: userDataDir='brower_temp'     None 不记录登录状态  填写数据目录则记录登录状态
        :param 启用拦截器:
        :param timeout:
        '''
        executablePath=绝对路径(executablePath)
        userDataDir=绝对路径(userDataDir)
        self.timeout = timeout * 1000
        self.headless = headless  # 无头模式 False
        self.executablePath = executablePath  # r'D:\pycharm_project\ChromePortable\App\Google Chrome\chrome.exe',
        self.width = width
        self.height = height
        self.userDataDir = userDataDir if userDataDir else userDataDir_创建(
            'user_0')  # r'D:\pycharm_project\ChromePortable\Data\User Data', #用户地址
        print(f'userDataDir-->{self.userDataDir}')
        self.option_networkidle0 = {'waitUntil': 'networkidle0', 'timeout': self.timeout}  # 在 500ms 内没有任何网络连接
        self.option_domcontentloaded = {'waitUntil': 'domcontentloaded', 'timeout': self.timeout}  # 状态树构建完成
        self.option_networkidle2 = {'waitUntil': 'networkidle2', 'timeout': self.timeout}  # 在 500ms 内网络连接个数不超过 2 个
        self.option_load = {'waitUntil': 'load', 'timeout': self.timeout}
        self.option_timeout = {'timeout': self.timeout}
        self.启用拦截器 = 启用拦截器
        self.help = '''
        page.waitForXPath：等待 xPath 对应的元素出现，返回对应的 ElementHandle 实例
        page.waitForSelector ：等待选择器对应的元素出现，返回对应的 ElementHandle 实例
        page.waitForResponse ：等待某个响应结束，返回 Response 实例
            await page.waitForResponse("https://www.qq.com")
            await page.waitForResponse(lambda res:res.url=="https://www.qq.com" and res.status==200)
        page.waitForRequest：等待某个请求出现，返回 Request 实例
            await page.waitForRequest("https://www.qq.com")
            await page.waitForeRequest(lambda req:req.url=="https://www.qq.com" and res.mothed=="GET")
        page.waitForFunction：等待在页面中自定义函数的执行结果，返回 JsHandle 实例
            await self.pypp.page.waitForFunction('showButtons')  填函数名
        page.waitFor：设置选择器 或者 方法 或者 等待时间

        self.page.on("request",lambda x:print(x.url()))  不需要开拦截器

        page.goto：打开新页面
        page.goBack ：回退到上一个页面
        page.goForward ：前进到下一个页面
        page.reload ：重新加载页面
        page.waitForNavigation：等待页面跳转
        '''

    async def 设置UserAgent(self, user_agent):
        await self.page.setUserAgent(user_agent)

    async def 加载浏览器(self, user_agent=None, opthon=None):
        from pyppeteer import launch, launcher
        # from pyppeteer.network_manager import Request, Response
        # from pyppeteer.dialog import Dialog

        # from pyppeteer import launcher
        # if '--enable-automation' in launcher.DEFAULT_ARGS:
        #     launcher.DEFAULT_ARGS.remove('--enable-automation')

        if not opthon:
            proxy_url= f'http://{self.host}:{self.port}' if (self.host and self.port) else ""
            if proxy_url:
                print(f'代理 --> {proxy_url}')
            opthon = {
                'headless': self.headless,  # 是否以”无头”的模式运行,，即是否显示窗口，默认为 True(不显示)
                'defaultViewport': {'width': self.width, 'height': self.height},
                'devtools': False,  # F12控制界面的显示，用来调试
                'ignoreHTTPSErrors': True,  # 是否忽略 Https 报错信息，默认为 False
                'executablePath': self.executablePath,
                # r'D:\pycharm_project\ChromePortable\App\Google Chrome\chrome.exe'
                'dumpio': True,  # 防止多开导致的假死
                'autoClose': False,  # 删除临时文件
                'args': [
                    '--mute-audio',  # 静音
                    f'--window-size={self.width + 20},{self.height}',  # 设置浏览器窗口大小，保持和页面大小一致
                    "--proxy-server=" + proxy_url,  #添加代理
                    # '--disable-infobars',                  #不显示信息栏，比如：chrome正在受到自动测试软件的控制
                    # "--start-maximized",                    # 最大化窗口
                    # '--no-sandbox',                        #取消沙盒模式，放开权限
                    # '--disable-extensions',  # 禁用拓展
                    # '--disable-gpu',
                    # '--disable-xss-auditor',
                ],
                'ignoreDefaultArgs': [
                    '--enable-automation',
                ]
            }
            if self.userDataDir:
                opthon['userDataDir'] = self.userDataDir  # 用户地址
        self.brower = await launch(opthon)
        self.page = await self.brower.newPage()
        if self.user and self.password:
            await self.page.authenticate({'username': self.user, 'password': self.password})
        if user_agent:
            await self.page.setUserAgent(user_agent)
        await self.page.setJavaScriptEnabled(enabled=True)
        self.page.setDefaultNavigationTimeout(1000 * self.timeout)  # 跳转超时
        await self.page.setViewport(viewport={'width': self.width, 'height': self.height})
        if self.启用拦截器:
            await self.page.setRequestInterception(True)
            # self.page.on("request",lambda x:print(x.url()))
            self.page.on("request", lambda x: asyncio.ensure_future(self.request拦截器(x)))
            self.page.on("response", lambda x: asyncio.ensure_future(self.response拦截器(x)))
            self.page.on('dialog', lambda x: asyncio.ensure_future(self.dialog拦截器(x)))
        # 以下为插入中间js，将淘宝会为了检测浏览器而调用的js修改其结果。
        await self.page.evaluate('''() =>{ Object.defineProperties(navigator,{ webdriver:{ get: () => false } }) }''')
        await self.page.evaluate('''() =>{ window.navigator.chrome = { runtime: {},  }; }''')
        await self.page.evaluate(
            '''() =>{ Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] }); }''');
        await self.page.evaluate(
            '''() =>{ Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5,6], }); }''')

    async def 获取页面(self, url, 跳转=False,次数=1):
        for i in range(次数):
            try:
                if 跳转:
                    await asyncio.gather(
                        self.page.goto(url, options=self.option_timeout),
                        self.page.waitForNavigation(options=self.option_networkidle2)
                    )
                    return await self.page.content()
                else:
                    await self.page.goto(url, options=self.option_timeout)
                    return await self.page.content()
            except Exception as e:
                print(f'获取页面错误 --> {e}')

    async def 关闭浏览器(self):
        if hasattr(self, 'brower'):
            try:
                pages = await self.brower.pages()
                for page in pages:
                    await page.close()
            except:
                pass
            finally:
                await self.brower.close()

    async def 元素属性(self, elem, attr: str = 'textContent'):
        '''元素 或者 元素列表'''
        if type(elem) == list:
            return [await self.page.evaluate('item=>item.{}'.format(attr), x) for x in elem]
        else:
            return await self.page.evaluate('item=>item.{}'.format(attr), elem)

    async def 设置COOKIE(self, cookie: str, domain='v.qq.com'):
        cookie = dict_From_CookieStr(cookie)
        [await self.page.setCookie({'name': k, 'value': v, 'domain': domain}) for k, v in cookie.items()]

    async def 获取COOKIE_IFRAME(self, iframe=0):
        return await self.page.evaluate(
            """ document.getElementsByTagName("iframe")[%(i)s].contentWindow.document.cookie """ % {'i': str(iframe)})

    async def 获取COOKIE(self):
        # return await self.page.evaluate('document.cookie', force_expr=True)
        cookie = await self.page.cookies()
        rt = []
        for c in cookie:
            rt.append(f"{c['name']}={c['value']}")
        return ';'.join(rt)

    async def 获取COOKIE_JSON(self, 文件路径='pypp_cookie.json'):
        cookies = await self.page.cookies()
        保存JSON(文件路径, cookies)

    async def 设置COOKIE_JSON(self, 文件路径='pypp_cookie.json'):
        cookies = 加载JSON(文件路径=文件路径)
        for cookie in cookies:
            await self.page.setCookie(cookie)

    async def 等待页面跳转(self, options):
        '''跳转超时'''
        await self.page.waitForNavigation(options=options if options else self.option_networkidle2)

    async def 清理多余窗口(self, 保留=2):
        while True:
            pages = await self.brower.pages()
            if len(pages) > 保留:
                await pages[-1].close()
                time.sleep(0.5)
            else:
                return

    async def 点击新窗口打开(self,btn,html=False):
        while True:
            await btn.click()
            pages = await self.brower.pages()
            if len(pages) >= 3:
                new_page=pages[-1]
                try:
                    if self.user and self.password:
                        await new_page.authenticate({'username': self.user, 'password': self.password})  # 设置代理
                    if html:
                        try:
                            await new_page.waitForNavigation(options=self.option_networkidle2)
                        except Exception as ee:
                            pcolor(f'点击新窗口打开错误 {ee.__traceback__.tb_lineno}:{ee}','error')
                        content=await new_page.content()
                        await new_page.close()
                        return content
                    else:
                        return new_page
                except Exception as e:
                    pcolor(f'点击新窗口打开错误 {e.__traceback__.tb_lineno}:{e}','warn')
                    await new_page.close()
            time.sleep(0.5)

    async def request拦截器(self, req):
        # from pyppeteer.network_manager import Request, Response
        resourceType = req.resourceType
        if resourceType in ['image']:  # 不加载资源文件
            await req.continue_()
            # ['document','stylesheet','image','media','font','script','texttrack','xhr','fetch','eventsource','websocket','manifest','other']
            # print('跳过图片',req.url)
            # await req.abort()
        elif 'searchCount' in req.url:
            '''
                * ``url`` (str): If set, the request url will be changed.
                * ``method`` (str): If set, change the request method (e.g. ``GET``).
                * ``postData`` (str): If set, change the post data or request.
                * ``headers`` (dict): If set, change the request HTTP header.
            '''
            data = {"url": "https://www.qq.com/", 'method': "GET", }
            await req.continue_(data)  # 修改url为xxx
        else:
            await req.continue_()

    async def response拦截器(self, resp):
        # from pyppeteer.network_manager import Request, Response
        if 'searchCount' in resp.url:
            response = await resp.text()
            print(response)  # 获得请求的text内容
            # js = await resp.json()
            # print(response)

    async def dialog拦截器(self, dialog):
        pass
        # from pyppeteer.dialog import Dialog
        # print(dialog.message)  # 打印出弹框的信息
        # print(dialog.type)  # 打印出弹框的类型，是alert、confirm、prompt哪种
        # print(dialog.defaultValue())#打印出默认的值只有prompt弹框才有
        # await page.waitFor(2000)  # 特意加两秒等可以看到弹框出现后取消
        # await dialog.dismiss()

        # await dialog.accept('000') #可以给弹窗设置默认值

    async def 执行js_return(self, js代码):
        '''
        login_token = await pypp.执行js_return('window.localStorage.token')
        :param js代码:
        :return:
        '''
        return await self.page.evaluate('''() =>{ return %s; }''' % js代码)

    async def 执行js(self, js代码):
        return await self.page.evaluate('''() =>{ %s }''' % js代码)

    async def 监听_标签文本(self, 标签='span', 文本='上传成功', 次数=30, 完全匹配=True):
        for i in range(次数):
            try:
                await self.page.waitForSelector(标签)
                # elems=await self.page.xpath(f'//{标签}')
                elems = await self.page.querySelectorAll(标签)
                if elems:
                    texts = await self.元素属性(elems)
                    print(f'监听 --> 标签:{标签} --> 文本:{texts}')
                    if 完全匹配:
                        if 文本 in texts:
                            return True
                    else:
                        for x in texts:
                            if 文本 in x:
                                return True
            except Exception as e:
                打印错误(e)
            time.sleep(1)
        return False

    def 异步函数(self, func, *args, **kwargs):
        return asyncio.get_event_loop().run_until_complete(func(*args, **kwargs))
class FFMPEG():
    def __init__(self, ffmpeg_href="", fpath=""):
        '''
        -loglevel quiet
        1."quiet"：最低日志级别，不输出任何信息。
        2."panic"：当发生严重错误时输出信息，并终止程序。
        3."fatal"：输出致命错误信息。
        4."error"：输出错误信息。
        5."warning"：输出警告信息。
        6."info"：输出一般信息，如编解码器信息、封装器信息等。
        7."verbose"：输出详细信息。
        8."debug"：输出调试信息。
        '''
        default_dir=os.path.join(os.path.expanduser('~'), 'FFMPEG')
        os.makedirs(default_dir,exist_ok=1)
        path =os.path.join(default_dir, "ffmpeg.exe")
        path2="ffmpeg.exe"
        if fpath:#本地优先级大于下载
            self.ffmpeg_path = fpath
        elif os.path.exists(path2):
            self.ffmpeg_path = path2
        else:
            self.ffmpeg_path = path
        self.ffmpeg_href = ffmpeg_href
    def 合并视频音频(self,文件路径,音频文件,视频文件):
        cmd=f'"{self.ffmpeg_path}" -i "{os.path.abspath(视频文件)}" -i "{os.path.abspath(音频文件)}" -vcodec copy -acodec copy -y {os.path.abspath(文件路径)}'
        self.执行(cmd)
    def ffmpeg_分离器合并(self,文件路径="out.mp4",文件名:str='file.txt'):
        #  ffmpeg -f concat -i filelist.txt -c copy output.mkv
        cmd=f'"{self.ffmpeg_path}" -f concat -safe 0 -i "{文件名}" -y -c copy "{文件路径}"'
        self.执行(cmd,show=1)
    def ffmpeg_拼接合并(self,文件路径="out.mp4",视频列表=[]):
        # ffmpeg -i "concat:input1.mpg|input2.mpg|input3.mpg" -c copy output.mpg
        s='|'.join(视频列表)
        cmd = f'"{self.ffmpeg_path}" -i concat"{s}" -y -c copy "{文件路径}"'
        self.执行(cmd,show=1)
    def ffmpeg_转格式合并(self,文件路径="out.mp4",ts目录=''):
        #00001.temp 00002.temp ...
        cmd=f'"{self.ffmpeg_path}" -i "1.temp" -c copy -f mpegts -bsf:v h264_mp4toannexb "1.ts"'
        cmd = f'copy /b "{os.path.abspath(ts目录)}\*.ts" "{文件路径}.temp"'
        cmd = f'"{self.ffmpeg_path}" -i "{os.path.abspath(文件路径)}.temp" -c copy -bsf:a aac_adtstoasc "{os.path.abspath(文件路径)}"'
    def 合并音频mp3(self,文件路径="out.mp3",音频列表=[]):
        #ffmpeg64.exe -i "concat:123.mp3|124.mp3" -acodec copy output.mp3
        音频列表=[os.path.abspath(x) for x in 音频列表]
        cmd='"{}" -i "concat:{}"  -c:a libfdk_aac -c:a copy -y "{}"'.format(self.ffmpeg_path, '|'.join(音频列表),os.path.abspath(文件路径))
        self.执行(cmd)
    def 合并音频m4a(self, 文件路径='out.m4a', 音频列表=[], show=False):
        '''
        合并 m4a mp3
        ffmpeg -i file1.m4a -acodec copy file1.aac
        ffmpeg -i file2.m4a -acodec copy file2.aac
        ffmpeg -i "concat:file1.aac|file2.aac" -c copy result.aac
        ffmpeg -i result.aac -acodec copy -bsf:a aac_adtstoasc filenew.m4a
        '''
        import shutil  # 删除aac文件夹
        try:
            os.makedirs('aac', exist_ok=1)
            all_files = {}
            for x in 音频列表:
                fpath, fname = os.path.split(x)
                ft = fname.rsplit('.', 1)
                name = ft[0]
                if len(ft) == 2:
                    hz = ft[1]
                else:
                    hz = ''
                all_files[int(name)] = {'name': name, 'path': x, 'dir': os.path.abspath(fpath), 'type': hz}
            all_files = sorted(all_files.items(), key=lambda k: k[0], reverse=False)
            path_aac = os.path.abspath('aac')
            if os.path.exists(path_aac):
                shutil.rmtree(path_aac)
            os.makedirs('aac', exist_ok=1)
            for i, file in enumerate(all_files):
                tp = file[1]['type']
                if tp == 'mp3':
                    # 设置音频编码和比特率  192k 指的是比特率设置为每秒 192 千比特（kbps）。比特率越高，音频的音质通常越好，但文件大小也相应增大
                    cmd = f'"{self.ffmpeg_path}" -y -i "{file[1]["path"]}" -c:a aac -b:a 192k "{os.path.abspath("aac/{}.aac".format(file[0]))}"'
                elif tp in ['m4a', '', 'aac']:
                    cmd = f'"{self.ffmpeg_path}" -y -i "{file[1]["path"]}" -acodec copy "{os.path.abspath("aac/{}.aac".format(file[0]))}"'
                else:
                    raise Exception('合并m4a错误 : 文件类型->{}'.format(tp))
                self.执行(cmd, show=show)
                # print('\r {}/{}'.format(i+1,len(all_files)),end='',flush=1)
            files = [os.path.abspath(f'aac/{x[0]}.aac') for x in all_files]
            resutl_aac = os.path.abspath("result.aac")
            resutl_m4a = os.path.abspath(文件路径)
            cmd2 = f'"{self.ffmpeg_path}" -y -i "concat:{"|".join(files)}" -acodec copy "{resutl_aac}"'
            self.执行(cmd2, show=show)
            cmd3 = f'"{self.ffmpeg_path}" -y -i "{resutl_aac}" -acodec copy "{resutl_m4a}"'
            self.执行(cmd3, show=show)
            if os.path.exists(resutl_aac):
                os.remove(resutl_aac)
            if os.path.exists(path_aac):
                shutil.rmtree(path_aac)
            return True
        except Exception as e:
            打印错误(e)
            return False
    def 合成图片视频(self, 图片路径, 音频路径, 输出路径):
        ffmpeg_command = [
            self.ffmpeg_path,
            '-loop', '1',
            '-i', 图片路径,
            '-i', 音频路径,
            '-c:v', 'libx264',
            '-c:a', 'aac',
            '-vf', 'scale=720:1280, format=yuv420p',
            '-r','30',
            '-shortest',
            输出路径,'-y'
        ]
        subprocess.call(ffmpeg_command,stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    def 合并视频(self,视频路径1, 视频路径2, 输出路径):
        ffmpeg_command = [
            self.ffmpeg_path,
            '-i', 视频路径1,
            '-i', 视频路径2,
            '-filter_complex', "[0:v:0][0:a:0][1:v:0][1:a:0]concat=n=2:v=1:a=1[v][a]",
            '-map','[v]',
            '-map','[a]',
            '-r', '30',
            '-shortest',
            输出路径,'-y'
        ]
        subprocess.call(ffmpeg_command,stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    def mpeg(self,输入路径,输出路径,show=True):
        cmd = f'"{self.ffmpeg_path}" -i "{输入路径}" -y -c copy -bsf:a aac_adtstoasc "{输出路径}" -loglevel error'
        self.执行(cmd, show=show)
    def 执行(self,cmd,show=False):
        if not os.path.exists(self.ffmpeg_path):
            if not self.ffmpeg_href:
                self.ffmpeg_href=FFMPEG_HREF
            下载文件_进度条(文件路径=self.ffmpeg_path, 网址=self.ffmpeg_href)
        if show:
            return cmd_subprocess_run(cmd,类型=2)
        else:
            return cmd_subprocess_run(cmd)
class PROXY():
    def __init__(self):
        '''
        设置代理
        enable: 0关闭，1开启
        proxyIp: 代理服务器ip及端口，如 "192.168.70.127:808"
        IgnoreIp:忽略代理的ip或网址，如 "172.*;192.*;"
        '''
        self.KEY_ProxyEnable = "ProxyEnable"
        self.KEY_ProxyServer = "ProxyServer"
        self.KEY_ProxyOverride = "ProxyOverride"
        self.KEY_XPATH = "Software\Microsoft\Windows\CurrentVersion\Internet Settings"
    def 设置代理(self,开启=1, 代理IP='127.0.0.1:8080', 白名单=""):
        '''
        :param 开启: 1 或者 0
        :param 代理IP: 127.0.0.1:8080;127.0.0.1:8888
        :param 白名单: 127.*;10.*;172.16.*;
        :return:
        '''
        hKey = winreg.OpenKey(winreg.HKEY_CURRENT_USER, self.KEY_XPATH, 0, winreg.KEY_WRITE)
        winreg.SetValueEx(hKey, self.KEY_ProxyEnable, 0, winreg.REG_DWORD, 开启)
        winreg.SetValueEx(hKey, self.KEY_ProxyServer, 0, winreg.REG_SZ, 代理IP)
        winreg.SetValueEx(hKey, self.KEY_ProxyOverride, 0, winreg.REG_SZ, 白名单)
        winreg.CloseKey(hKey)
    def 获取代理(self)->list:
        '''返回列表[{},{}]'''
        hKey = winreg.OpenKey(winreg.HKEY_CURRENT_USER, self.KEY_XPATH, 0, winreg.KEY_READ)
        retVal = winreg.QueryValueEx(hKey, self.KEY_ProxyEnable)
        开启=retVal[0]
        if 开启:
            res = winreg.QueryValueEx(hKey, self.KEY_ProxyServer)
            # http=127.0.0.1:8888;https=127.0.0.1:8888
            lst=res[0].split(';')
            lst= list(set([x.split('=',1)[1] if '=' in x else x for x in lst ]))
            return [{'http':f'http://{x}','https':f'http://{x}'} for x in lst]
        else:
            return []
    def get_proxies(self)->dict:
        '''返回列表{} 或者 None'''
        hKey = winreg.OpenKey(winreg.HKEY_CURRENT_USER, self.KEY_XPATH, 0, winreg.KEY_READ)
        retVal = winreg.QueryValueEx(hKey, self.KEY_ProxyEnable)
        开启 = retVal[0]
        if 开启:
            res = winreg.QueryValueEx(hKey, self.KEY_ProxyServer)
            # http=127.0.0.1:8888;https=127.0.0.1:8888
            lst = res[0].split(';')
            lst = list(set([x.split('=', 1)[1] if '=' in x else x for x in lst]))
            return {'http': f'http://{lst[0]}', 'https': f'http://{lst[0]}'} if lst else None
        else:
            return None
class SOCKET():
    def __init__(self, host: str = '127.0.0.1', port: int = 8888,连接数=10,byteSize=1024):
        self.host = host
        self.port = port
        self.addr = (host, port)
        self.conn_number=连接数
        self.__close__=False
        self.byteSize=byteSize
    def __监听__(self,callback=None):
        self.__close__ = False
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.bind(self.addr)
        sock.listen(self.conn_number)#连接数
        print('运行socket监听:{}'.format(self.addr))
        while not self.__close__:
            try:
                client, addr0 = sock.accept()
                data = client.recv(self.byteSize)
                if data:
                    if callback:
                        callback(data)
                    else:
                        print('socket接收到:{}'.format(data))
            except Exception as e:
                print(e,e.__traceback__.tb_lineno)
            time.sleep(0.2)
        print('已退出socket监听')
    def 运行监听(self,回调函数=None):
        self.__close__ = False
        th=threading.Thread(target=self.__监听__,args=(回调函数,))
        th.setDaemon(True)
        th.start()
    def 发送消息(self, data: bytes):
        try:
            sock = socket.socket()
            sock.connect(self.addr)
            sock.send(data)
            sock.close()
        except Exception as e:
            打印错误(e)
    def 关闭服务(self):
        self.__close__=True
class COM_SERIAL():
    def __init__(self, port=None,description=[], baudrate=9600, bytesize=1024, timeout=None, writeTimeout=None):
        '''
        port 端口
        baudrate 波特率
        bytesize 字节大小
        '''
        self.port=port
        self.description=description
        assert self.port or self.description,'缺少参数'
        self.ser=serial.Serial()
        self.ser.baudrate=baudrate
        self.bytesize = bytesize
        self.timeout=timeout
        self.writeTimeout = writeTimeout
        self.__close__=False
    def connect(self):
        self.ser.close()
        while True:
            try:
                if self.ser.isOpen():
                    return
                else:
                    if self.port:
                        self.ser.port=self.port
                        self.ser.open()
                        print('已连接:', self.port)
                    elif self.description:
                        for port in self.串口列表():  # 自动获取非COM1的端口
                            for desc in self.description:
                                if desc in port.description:
                                    self.ser.port = port.name
                                    self.ser.open()
                                    print('已连接:', port.description)
                                    break
            except Exception as e:
                打印错误(e)
            time.sleep(1)
    def __监听__(self):
        while not self.__close__:
            if self.ser.isOpen():
                try:
                    if self.ser.in_waiting:
                        d=self.ser.read(self.ser.in_waiting)
                        self.收到(d)
                except Exception as e:
                    self.ser.close()
                    print(e,e.__traceback__.tb_lineno)
            else:  # 重连
                self.connect()
            time.sleep(0.2)
        print('串口已关闭')
    def 串口列表(self,打印=False):
        lst=list(serial.tools.list_ports.comports())
        if 打印:
            for x in lst:
                print(x.name,x.description)
        return lst
    def 发送HEX(self,data:str or bytesarray='AA0102AB'):
        if isinstance(data,str): #16进制字符串转bytes
            data=bytes.fromhex(data)
        #data=bytearray([0xAA,0x01,0x02,0xAB])
        print('发送hex:',data.hex())
        self.ser.write(data)
    def 收到(self,data:bytes):
        hex_data=data.hex().encode()
        print('收到hex:',hex_data)
    def 关闭串口(self):
        self.__close__=True
    def 运行监听(self):
        th=threading.Thread(target=self.__监听__)
        th.setDaemon(True)
        th.start()
class ARGV():
    '''
    a=ARGV()
    a.设置参数('-i')
    a.设置参数('-uninstall')
    a.设置参数('--key',类型=str)
    args=a.解析参数()
    print(args)
    '''
    def __init__(self):
        self.parse=argparse.ArgumentParser(description='参数说明')
    def 设置参数(self,字段名:str,类型=str,默认值=None,必选参数=True,参数说明='没有参数说明',帮助提示="没有帮助提示")->None:
        '''
        字段名='-i'  或者  字段名='--install'
        '''
        self.parse.add_argument(字段名,type=类型,default=默认值,required=必选参数,metavar=参数说明,help=帮助提示)
    def 解析参数(self)->argparse.Namespace:
        return self.parse.parse_args()
class WINRAR():
    def __init__(self,winRAR_path=r'C:\Program Files\WinRAR\Rar.exe',password=None):
        #C:\Program Files\WinRAR\Rar.exe
        #C:\Program Files\WinRAR\WinRAR.exe
        self.exe_path=winRAR_path
        self.password=password
    def 解压(self, 文件路径:str, 解压目录:str= '', timeout=None)->bool:
        '''WINRAR().解压("a.rar","aaa")'''
        try:
            文件路径 = os.path.abspath(文件路径)
            if not 解压目录:
                解压目录 = 文件路径.rsplit('.')[0]
            else:
                解压目录 = os.path.abspath(解压目录)
            if not os.path.exists(文件路径):
                raise Exception('winRAR 解压错误:文件不存在')
            os.makedirs(解压目录, exist_ok=1)
            # x 用绝对路径解压文件
            # -o+ 导出覆盖文件
            command = [self.exe_path, 'x', 文件路径, '-o+', 解压目录]
            if self.password:
                command.append(f'-p{self.password}')
            si = subprocess.STARTUPINFO()
            si.dwFlags = STARTF_USESHOWWINDOW
            if subprocess.call(command, startupinfo=si,timeout=timeout,creationflags=CREATE_NO_WINDOW) == 0:
                return True
            else:
                return False
        except Exception as e:
            打印错误(e)
            return False
    def 压缩(self, 文件路径:str, 压缩目录:str= '', timeout=None):
        '''
        sx.WINRAR().压缩(文件路径="a.jpg",目录="b.rar")
        sx.WINRAR().压缩(文件路径="xxx",目录="b.rar")
        '''
        try:
            压缩目录 = os.path.abspath(压缩目录)
            文件路径 = os.path.abspath(文件路径)
            if os.path.isfile(文件路径):
                # -ep 只压缩文件本身
                command = [self.exe_path, 'a', '-ep', 压缩目录, 文件路径]
            else:
                # a 添加文件到压缩文件
                # r 递归
                # -ep1 表示去掉文件路径中的第一级目录
                command = [self.exe_path, 'a', '-ep1', '-r', 压缩目录, 文件路径]
            if self.password:
                command.append(f'-p{self.password}')
            si = subprocess.STARTUPINFO()
            si.dwFlags = STARTF_USESHOWWINDOW
            if subprocess.call(command, startupinfo=si,timeout=timeout,creationflags=CREATE_NO_WINDOW) == 0:
                return True
            else:
                return False
        except Exception as e:
            打印错误(e)
            return False
class JSRPC():
    def __init__(self,port=9000):
        '''
        运行服务器端
        rpc=JSRPC(port=9000)
        rpc.run_server()
        time.sleep(10000)

        运行客户端
        rpc=JSRPC(port=9000)
        rpc.run_client(name="cookie",data="hello")
        即向浏览器发送  {'name': 'cookie', 'data': 'hello'}
        浏览器如果没注册name则返回 {'data': '未注册action:cookie', 'code': False, 'name': None}
        如果已注册则返回 {'data': 'helloooo....', 'code': True, 'name': 'cookie'}

        浏览器注入js代码  建立通信

        const socket = new WebSocket('ws://127.0.0.1:9000')//修改服务器端口号
        socket.onerror = (e) => {console.error('rpc错误：', e)}
        socket.onclose = (e) => {console.log('rpc已关闭', e)}
        socket.onopen = function (e) {
            message = {reg: 'cookie'}
            console.log('发送:',message)
            socket.send(JSON.stringify(message));
        };
        socket.onmessage = (e) => {
            data = JSON.parse(e.data);
            console.log('收到:', data)
            websocket_id=data['websocket_id']
            if (websocket_id){
                message={'websocket_id':websocket_id}
                /*  下面修改自己的代码赋值给 message['data'] */



                message['data']="helloooo...."
                /*----------------------------------------*/
                console.log('发送:',message)
                socket.send(JSON.stringify(message));
            }
        }

        '''
        self.port=port
        self.conn = {}
        self.loop = asyncio.new_event_loop()
    def run_server(self):
        import websockets
        async def handle_client(websocket, url_path):
            current_id = id(websocket)
            #清理失效的连接
            for _ in [k for k, v in self.conn.items() if not v[1].open]:
                del self.conn[_]
            async for message in websocket:
                sent_data = {'data': '', 'code': True}
                try:
                    recv_data = json.loads(message)  # 消息数据
                    print(f"[ 收到 <-- {current_id} ] {recv_data}")
                    reg = recv_data.get('reg')  # 注册名
                    name = recv_data.get('name')  # 请求注册名
                    data = recv_data.get('data')  # 获取数据
                    desc_id = recv_data.get('websocket_id')  # 目标websocket id
                    recv_websocket = websocket
                    #如果是注册
                    if reg:
                        for websocket_id_, v in self.conn.items():
                            if reg == v[0]:
                                self.conn[current_id] = [reg, websocket]
                                del self.conn[websocket_id_]
                                print(f'[ 删除 {websocket_id_} ]')
                                break
                        else:
                            self.conn[current_id] = [reg, websocket]
                        sent_data['data'] = f'注册成功:{reg}'
                    #通过注册名找到websocket连接发送  py转发给浏览器消息
                    elif name:
                        for websocket_id_, v in self.conn.items():
                            if v[0] and name == v[0]:
                                recv_websocket = v[1]
                                sent_data['data'] = data
                                sent_data['websocket_id'] = current_id
                                self.conn[current_id] = [None, websocket]
                                break
                        else:
                            sent_data['name'] = None
                            sent_data['data'] = f'未注册action:{name}'
                            sent_data['code'] = False
                    #通过id找到websocket连接发送  浏览器转发给py消息
                    elif desc_id:
                        if desc_id in self.conn:
                            sent_data['data'] = data
                            sent_data['name'] = self.conn[current_id][0]
                            recv_websocket = self.conn[desc_id][1]
                        else:
                            sent_data['data'] = '目标连接已断开'
                    else:
                        sent_data['data'] = f'没有action参数'
                        sent_data['code'] = False
                    print(f'[ 发送 --> {id(recv_websocket)} ] {sent_data}')
                    await recv_websocket.send(json.dumps(sent_data))  # 发送处理结果给客户端
                except Exception as e:
                    打印错误(e)
                print('-' * 100)
        start_server = websockets.serve(handle_client, "127.0.0.1", self.port,loop=self.loop)
        print(f'run at http://127.0.0.1:{self.port}')
        print('*' * 100)
        self.loop.run_until_complete(start_server)
        self.loop.run_forever()
    def run_server_thread(self):
        th=threading.Thread(target=self.run_server)
        th.start()
        time.sleep(1)
    def run_client(self,name='cookie', data='你好'):
        import websockets
        try:
            async def rpc(name, data):
                try:
                    async with websockets.connect(f'ws://127.0.0.1:{self.port}') as websocket:
                        json_data = {'name': name, 'data': data}
                        await websocket.send(json.dumps(json_data))  # 发送消息
                        recv_msg = await websocket.recv()
                        return json.loads(recv_msg)  # 接收消息
                except Exception as e:
                    raise Exception(f'jsrpc错误:{e}')
            return asyncio.get_event_loop().run_until_complete(rpc(name, data))
        except Exception as e:
            return {'code': False, 'data': None,'msg':f'{e}'}
class 监听程序运行():
    def __init__(self, 程序EXE路径:str):
        self.程序EXE路径=程序EXE路径
        if isinstance(self.程序EXE路径, str):
            self.run_pid=subprocess.Popen(r'{}'.format(self.程序EXE路径), stdout=None, stderr=None, shell=False)
        else:
            raise Exception('cmd格式要求字符串')
        self.__close__ = False
    def __监听__(self):
        if self.run_pid:
            print('正在监听程序:{}'.format(self.程序EXE路径))
            stdoutdata, stderrdata = self.run_pid.communicate(input=None, timeout=None)
            print('stdoutdata:{}'.format(stdoutdata))
            print('stderrdata:{}'.format(stderrdata))
            # 没有强制退出
            if self.run_pid:
                code = self.run_pid.returncode  # returncode 0 正常退出
                print('returncode:{}'.format(code))
                if code != 0:
                    print('程序已异常退出')
                else:
                    print('程序已正常退出')
            # 强制退出了
            else:
                print('程序已强制退出')
            self.run_pid = False
    def 运行监听(self):
        th=threading.Thread(target=self.__监听__)
        th.setDaemon(True)
        th.start()
    def 关闭程序(self):
        if self.run_pid:
            try:
                self.run_pid.kill()
            except:
                pass
        self.run_pid=False
class 电脑信息():
    def __init__(self):
        '''
        wmic diskdrive 可以看出来牌子和大小.
        Wmic logicaldisk 每一个盘的文件系统和剩余空间
        wmic cpu
        wmic memorychip
        wmic bios

        '''
        pass
    def wmic_format(self, cmd):
        '''
        cmd='wmic cpu get name /format:list'
        wmic csproduct get name,uuid,vendor /format:list
        '''
        with os.popen(cmd) as f:
            res = f.read()  # 获取管道信息
        rt={}
        for x in res.split('\n'):
            if x.strip():
                a=x.split('=',1)
                rt[a[0].lower()]=a[1]
        return rt
    def wmic(self,cmd):
        with os.popen(cmd) as f:
            res = f.readlines()  # 获取管道信息
        keys=[x for x in res[0].split(' ') if x.strip()]
        cmd=cmd+' get {} /format:list'.format(','.join(keys))
        rt = self.wmic_format(cmd)
        return rt
    def 主机名(self):
        return socket.gethostname()
    def 网卡(self):
        return psutil.net_if_addrs()
    def 内网IP(self):
        return socket.gethostbyname_ex(self.主机名())[-1]
    def 硬盘分区(self):
        return psutil.disk_partitions()
    def 内存(self):
        return psutil.virtual_memory()
    def 系统开机时间(self):
        return datetime.datetime.fromtimestamp(psutil.boot_time ()).strftime("%Y-%m-%d %H: %M: %S")
    def 磁盘(self):
        硬盘空间=[]
        for disk in psutil.disk_partitions():
            try:
                硬盘空间.append(psutil.disk_usage(disk.device))
            except Exception as e:
                pass
        return 硬盘空间
    def 接收流量(self):
        return '{0:.2f} Mb'.format(self.mb(psutil.net_io_counters().bytes_recv))
    def 发送流量(self):
        return '{0:.2f} Mb'.format(self.mb(psutil.net_io_counters().bytes_sent))
    def 用户(self):
        return psutil.users()
    def mb(self,kb):
        return kb/1024/1024
    def 主板信息(self):
        return self.wmic('wmic csproduct')
    def cpu(self):
        return self.wmic('wmic cpu')
class 结构体:
    def __init__(self):
        ...
    def dict(self):
        return self.__dict__
    def keys(self):
        return list(self.__dict__.keys())
    def values(self):
        return list(self.__dict__.values())
class 过滤字符串():
    def __init__(self):
        self.keys=[]
        self.flag=False
    def 设置_关键字(self,keys=''):
        self.keys=[x.strip() for x in keys.strip().split(' ') if x.strip()]
    def 查找_关键字(self,name='',打印跳过=1):
        if not self.flag and all([x in name for x in self.keys]):
            self.flag=True
        if 打印跳过:
            if not self.flag:
                print('跳过',name)
        return self.flag
#静态类
class 弹窗:
    @classmethod
    def 信息框(cls,标题:str='信息框', 文本:str="信息框"):
        import win32api, win32con
        win32api.MessageBox(None, 文本, 标题, win32con.MB_OK | win32con.MB_ICONQUESTION)
    @classmethod
    def 选择框(cls,标题="选择框", 文本="选择框内容"):
        import win32api, win32con
        x = win32api.MessageBox(None, 文本, 标题, win32con.MB_YESNO | win32con.MB_DEFBUTTON1 | win32con.MB_ICONINFORMATION)
        return True if x == 6 else False  # 7
    @classmethod
    def 选择文件夹(cls,标题='选择文件夹'):
        from tkinter import Tk
        from tkinter import filedialog
        root = Tk()
        root.withdraw()  # 将Tkinter.Tk()实例隐藏
        path = filedialog.askdirectory(title=标题)
        root.destroy()
        return path
    @classmethod
    def 选择文件(cls, 路径: str = '.', 标题: str = '选择文件', 文件类型:tuple=(('excel文件', '*.xlsx'),)): #默认打开xlsx
        '''文件类型 (('png files', '*.png'), ('jpeg files', '*.jpeg'),)'''
        from tkinter import Tk
        from tkinter import filedialog
        root = Tk()
        root.withdraw()  # 将Tkinter.Tk()实例隐藏
        fname = filedialog.askopenfilename(title=标题, initialdir=路径, filetypes=文件类型)
        root.destroy()
        return fname
    @classmethod
    def 输入数字(cls,标题:str='整数录入',文本说明:str='请输入整数'):
        from tkinter import Tk
        from tkinter import simpledialog
        root = Tk()
        root.withdraw()  # 将Tkinter.Tk()实例隐藏
        d = simpledialog.askinteger(title=标题,prompt=文本说明,initialvalue=0)
        root.destroy()
        return d
    @classmethod
    def 输入浮点(cls,标题:str='浮点录入',文本说明:str='请输入浮点数'):
        from tkinter import Tk
        from tkinter import simpledialog
        root = Tk()
        root.withdraw()  # 将Tkinter.Tk()实例隐藏
        d = simpledialog.askfloat(title=标题,prompt=文本说明,initialvalue=0.0)
        root.destroy()
        return d
    @classmethod
    def 输入字符串(cls,标题:str='字符串录入',文本说明:str='请输入字符串'):
        from tkinter import Tk
        from tkinter import simpledialog
        root = Tk()
        root.withdraw()  # 将Tkinter.Tk()实例隐藏
        d = simpledialog.askstring(title=标题,prompt=文本说明,initialvalue='')
        root.destroy()
        return d
class 随机:
    '''
    随机0-1小数   random.random()
    随机整数      random.randint(1,10)
    随机小数      random.uniform(1,10)
    随机列表      random.choice([1,2,3])
    随机列表设置权重
    my_weights = [0.1,0.5,0.4]
    random.choices(['a','b','c'],weights=my_weights,k=1)
    '''
    @classmethod
    def 随机字符串(cls,字符串,长度):
        rt = []
        for i in range(长度):
            rt.append(random.choice(字符串))
        return ''.join(rt)
    @classmethod
    def 数字(cls,长度:int=20)->str:
        return cls.随机字符串(string.digits,长度)
    @classmethod
    def 大写字母(cls,长度:int=20)->str:
        return cls.随机字符串(string.ascii_uppercase,长度)
    @classmethod
    def 小写字母(cls, 长度: int=20) -> str:
        return cls.随机字符串(string.ascii_lowercase, 长度)
    @classmethod
    def 字母(cls, 长度: int=20) -> str:
        return cls.随机字符串(string.ascii_letters, 长度)
    @classmethod
    def 字母数字(cls, 长度: int=20) -> str:
        return cls.随机字符串(string.ascii_letters + string.digits,长度)
    @classmethod
    def 字母数字特殊符号(cls, 长度: int=20) -> str:
        return cls.随机字符串(string.ascii_letters + string.digits + string.punctuation,长度)
    @classmethod
    def 列表随机一个(cls,列表:list):
        assert 列表, '随机列表空'
        return random.choice(列表)
    @classmethod
    def 列表随机多个(cls,列表:list,个数)->list:
        assert 列表, '随机列表空'
        assert len(列表)>个数,'随机大于总长度'
        return random.sample(列表,个数)
#未安装依赖模块
def get_aliyun_token(AccessKeyId,AccessKeySecret):
    from aliyunsdkcore.client import AcsClient
    from aliyunsdkcore.request import CommonRequest
    # 创建AcsClient实例
    if AccessKeyId and AccessKeySecret:
        # client = AcsClient("<您的AccessKey Id>", "<您的AccessKey Secret>", "cn-shanghai")
        client = AcsClient(AccessKeyId,AccessKeySecret, "cn-shanghai")
    else:
        raise Exception('阿里云密钥错误')
    request = CommonRequest()
    request.set_method('POST')
    request.set_domain('nls-meta.cn-shanghai.aliyuncs.com')
    request.set_version('2019-02-28')
    request.set_action_name('CreateToken')
    response = client.do_action_with_exception(request)
    res=json.loads(response.decode())
    token=res['Token']['Id']
    return token
def 语音合成(text, APPKEY, fpath, token,voice="ailun"):
    '''
    [{'language': '中文普通话', 'sex': 'female', 'speaker': '知媛', 'speakerId': 'zhiyuan'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '知悦', 'speakerId': 'zhiyue'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '知莎', 'speakerId': 'zhistella'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '知柜', 'speakerId': 'zhigui'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '知硕', 'speakerId': 'zhishuo'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '知达', 'speakerId': 'zhida'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '艾琪', 'speakerId': 'aiqi'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '艾诚', 'speakerId': 'aicheng'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '艾佳', 'speakerId': 'aijia'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '思琪', 'speakerId': 'siqi'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '思佳', 'speakerId': 'sijia'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '马树', 'speakerId': 'mashu'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '悦儿', 'speakerId': 'yuer'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '若兮', 'speakerId': 'ruoxi'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '艾达', 'speakerId': 'aida'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '思诚', 'speakerId': 'sicheng'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '宁儿', 'speakerId': 'ninger'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '小云', 'speakerId': 'xiaoyun'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '小刚', 'speakerId': 'xiaogang'}, {'language': '中英', 'sex': 'female', 'speaker': '知妙_多情感', 'speakerId': 'zhimiao_emo'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '知米_多情感', 'speakerId': 'zhimi_emo'}, {'language': '中文普通话', 'sex': 'child', 'speaker': '知贝_多情感', 'speakerId': 'zhibei_emo'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '知燕_多情感', 'speakerId': 'zhiyan_emo'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '知甜_多情感', 'speakerId': 'zhitian_emo'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '知雅', 'speakerId': 'zhiya'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '艾夏', 'speakerId': 'aixia'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '艾悦', 'speakerId': 'aiyue'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '艾雅', 'speakerId': 'aiya'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '艾婧', 'speakerId': 'aijing'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '艾美', 'speakerId': 'aimei'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '思悦', 'speakerId': 'siyue'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '艾娜', 'speakerId': 'Aina'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '艾硕', 'speakerId': 'aishuo'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '艾雨', 'speakerId': 'aiyu'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '小美', 'speakerId': 'xiaomei'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '伊娜', 'speakerId': 'yina'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '思婧', 'speakerId': 'sijing'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '知甜', 'speakerId': 'zhitian'}, {'language': '方言场景', 'sex': 'female', 'speaker': '知青', 'speakerId': 'zhiqing'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '知猫', 'speakerId': 'zhimao'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '老妹', 'speakerId': 'laomei'}, {'language': '东北男声', 'sex': 'male', 'speaker': '老铁', 'speakerId': 'laotie'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '小仙', 'speakerId': 'xiaoxian'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '柜姐', 'speakerId': 'guijie'}, {'language': '中文普通话', 'sex': 'female', 'speaker': 'Stella', 'speakerId': 'stella'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '猫小美', 'speakerId': 'maoxiaomei'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '巧薇', 'speakerId': 'qiaowei'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '艾伦', 'speakerId': 'ailun'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '艾飞', 'speakerId': 'aifei'}, {'language': '中文普通话', 'sex': 'male', 'speaker': '亚群', 'speakerId': 'yaqun'}, {'language': '中文普通话', 'sex': 'male', 'speaker': 'Stanley', 'speakerId': 'stanley'}, {'language': '中文普通话', 'sex': 'male', 'speaker': 'Kenny', 'speakerId': 'kenny'}, {'language': '中文普通话', 'sex': 'female', 'speaker': 'Rosa', 'speakerId': 'rosa'}, {'language': '中文普通话', 'sex': 'female', 'speaker': '瑞琳', 'speakerId': 'ruilin'}, {'language': '中文普通话', 'sex': 'child', 'speaker': '艾彤', 'speakerId': 'aitong'}, {'language': '中文普通话', 'sex': 'child', 'speaker': '艾薇', 'speakerId': 'aiwei'}, {'language': '中文普通话', 'sex': 'child', 'speaker': '杰力豆', 'speakerId': 'jielidou'}, {'language': '中文普通话', 'sex': 'child', 'speaker': '小北', 'speakerId': 'xiaobei'}, {'language': '中文普通话', 'sex': 'child', 'speaker': '思彤', 'speakerId': 'sitong'}, {'language': '中文普通话', 'sex': 'child', 'speaker': '艾宝', 'speakerId': 'aibao'}, {'language': '意大利语', 'sex': 'female', 'speaker': 'Perla', 'speakerId': 'perla'}, {'language': '西班牙语', 'sex': 'female', 'speaker': 'Camila', 'speakerId': 'camila'}, {'language': '俄语', 'sex': 'female', 'speaker': 'masha', 'speakerId': 'masha'}, {'language': '韩语', 'sex': 'female', 'speaker': 'Kyong', 'speakerId': 'kyong'}, {'language': '越南语', 'sex': 'female', 'speaker': 'Tien', 'speakerId': 'tien'}, {'language': '泰语', 'sex': 'female', 'speaker': 'Waan', 'speakerId': 'waan'}, {'language': '德语', 'sex': 'female', 'speaker': 'Hanna', 'speakerId': 'hanna'}, {'language': '法语', 'sex': 'female', 'speaker': 'Clara', 'speakerId': 'clara'}, {'language': '美式英语', 'sex': 'female', 'speaker': 'ava', 'speakerId': 'ava'}, {'language': '英式英语', 'sex': 'male', 'speaker': 'Luca', 'speakerId': 'Luca'}, {'language': '英式英语', 'sex': 'female', 'speaker': 'Luna', 'speakerId': 'Luna'}, {'language': '英式英语', 'sex': 'female', 'speaker': 'Emily', 'speakerId': 'Emily'}, {'language': '英式英语', 'sex': 'male', 'speaker': 'Eric', 'speakerId': 'Eric'}, {'language': '美式英语', 'sex': 'female', 'speaker': 'Annie', 'speakerId': 'annie'}, {'language': '美式英语', 'sex': 'male', 'speaker': 'Andy', 'speakerId': 'Andy'}, {'language': '英式英语', 'sex': 'male', 'speaker': 'William', 'speakerId': 'William'}, {'language': '美式英语', 'sex': 'female', 'speaker': 'Abby', 'speakerId': 'Abby'}, {'language': '英式英语', 'sex': 'female', 'speaker': 'Lydia', 'speakerId': 'Lydia'}, {'language': '英式英语', 'sex': 'female', 'speaker': 'Olivia', 'speakerId': 'Olivia'}, {'language': '美式英文', 'sex': 'male', 'speaker': 'Brian', 'speakerId': 'brian'}, {'language': '美式英文', 'sex': 'female', 'speaker': 'Eva', 'speakerId': 'eva'}, {'language': '美式英文', 'sex': 'female', 'speaker': 'Donna', 'speakerId': 'donna'}, {'language': '美式英文', 'sex': 'female', 'speaker': 'Cally', 'speakerId': 'cally'}, {'language': '美式英文', 'sex': 'female', 'speaker': 'Cindy', 'speakerId': 'cindy'}, {'language': '美式英文', 'sex': 'female', 'speaker': 'Beth', 'speakerId': 'beth'}, {'language': '美式英文', 'sex': 'female', 'speaker': 'Betty', 'speakerId': 'betty'}, {'language': '英式英语', 'sex': 'female', 'speaker': 'Wendy', 'speakerId': 'Wendy'}, {'language': '香港粤语', 'sex': 'female', 'speaker': 'Kelly', 'speakerId': 'kelly'}, {'language': '粤语方言', 'sex': 'female', 'speaker': '佳佳', 'speakerId': 'jiajia'}, {'language': '东北男声', 'sex': 'male', 'speaker': '大虎', 'speakerId': 'dahu'}, {'language': '天津男声', 'sex': 'male', 'speaker': '艾侃', 'speakerId': 'aikan'}, {'language': '粤语方言', 'sex': 'female', 'speaker': '桃子', 'speakerId': 'taozi'}, {'language': '英式英语', 'sex': 'male', 'speaker': 'Harry', 'speakerId': 'Harry'}, {'language': '台湾方言', 'sex': 'female', 'speaker': '青青', 'speakerId': 'qingqing'}, {'language': '东北女声', 'sex': 'female', 'speaker': '翠姐', 'speakerId': 'cuijie'}, {'language': '湖南男声', 'sex': 'male', 'speaker': '小泽', 'speakerId': 'xiaoze'}, {'language': '粤语方言', 'sex': 'female', 'speaker': '姗姗', 'speakerId': 'shanshan'}, {'language': '日语', 'sex': 'female', 'speaker': '智香', 'speakerId': 'tomoka'}, {'language': '日语', 'sex': 'male', 'speaker': '智也', 'speakerId': 'tomoya'}, {'language': '印尼语', 'sex': 'female', 'speaker': 'Indah', 'speakerId': 'indah'}, {'language': '马来语', 'sex': 'female', 'speaker': 'Farah', 'speakerId': 'farah'}, {'language': '菲律宾语女声', 'sex': 'female', 'speaker': 'Tala', 'speakerId': 'tala'}, {'language': '四川方言', 'sex': 'female', 'speaker': '小玥', 'speakerId': 'xiaoyue'}]
    '''
    # token 24小时失效
    # pip install aliyun-nls
    import nls
    URL = "wss://nls-gateway.cn-shanghai.aliyuncs.com/ws/v1"
    # 参考https://help.aliyun.com/document_detail/450255.html获取token
    # 获取Appkey请前往控制台：https://nls-portal.console.aliyun.com/applist
    # 以下代码会根据上述TEXT文本反复进行语音合成
    class TestTts:
        def __init__(self, tid, test_file):
            self.__th = threading.Thread(target=self.__test_run)
            self.__id = tid
            self.__test_file = test_file
        def start(self, text):
            self.__text = text
            self.__f = open(self.__test_file, "wb")
            self.__th.start()
            self.__th.join()
        def test_on_metainfo(self, message, *args):
            pass
            #print("on_metainfo message=>{}".format(message))
        def test_on_error(self, message, *args):
            pass
            #print("on_error args=>{}".format(args))
        def test_on_close(self, *args):
            #print("on_close: args=>{}".format(args))
            try:
                self.__f.close()
            except Exception as e:
                print("关闭文件失败:", e)
        def test_on_data(self, data, *args):
            try:
                self.__f.write(data)
            except Exception as e:
                print("写入文件错误:", e,e.__traceback__.tb_lineno)
        def test_on_completed(self, message, *args):
            message=json.loads(message)
            header=message['header']
            if header['status']==20000000:
                print('语音生成成功-->{}'.format(self.__test_file))
            else:
                print('语音生成失败:{}'.format(header['status_text']))
        def __test_run(self):
            tts = nls.NlsSpeechSynthesizer(url=URL,
                                           token=token,
                                           appkey=APPKEY,
                                           on_metainfo=self.test_on_metainfo,
                                           on_data=self.test_on_data,
                                           on_completed=self.test_on_completed,
                                           on_error=self.test_on_error,
                                           on_close=self.test_on_close,
                                           callback_args=[self.__id])
            tts.start(self.__text, voice=voice, aformat="wav")
    def multiruntest(num=500):
        for i in range(0, num):
            name = "thread" + str(i)
            # t = TestTts(name, "tests/test_tts.pcm")
            # t = TestTts(name, "tests/test_tts.wav")
            t = TestTts(name, fpath)
            t.start(text)
    nls.enableTrace(False)
    multiruntest(1)

if __name__ == '__main__':
    pass