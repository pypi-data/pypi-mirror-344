#!/usr/bin/env python
# coding: utf-8

# ### load pyPC utilities

import os
import sys

print(os.getcwd())
cwd0 = './config/'
sys.path.append(cwd0)

import visualID_Eng as vID
from visualID_Eng import fg, hl, bg
vID.init(cwd0)
import tools4pyPC as t4p


# ### log MOs and basic functions

import re
import numpy as np
import pandas as pd
from scipy.optimize import linear_sum_assignment
from scipy.spatial.distance import pdist, squareform
from scipy.cluster.hierarchy import linkage, dendrogram
from ipywidgets import interact, Dropdown, Button, HBox
from IPython.display import display
import matplotlib.pyplot as plt
import seaborn as sns
import time
import os
import cclib
from cclib.parser.utils import PeriodicTable

from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

def extract_gaussian_info(logfile_path):
    """
    Extracts molecular orbital and structural information from a Gaussian log file using cclib.

    Parameters
    ----------
    logfile_path : str
        Path to the Gaussian .log file.

    Returns
    -------
    dict
        A dictionary containing UDFT/DFT type, basis size, molecular orbitals, geometry,
        occupation, HOMO index, spin values, and the AO overlap matrix.
    """
    if not os.path.isfile(logfile_path):
        raise FileNotFoundError(f"File not found: {logfile_path}")

    data = cclib.io.ccread(logfile_path)
    
    nbasis = getattr(data, "nbasis", None)
    if nbasis is None:
        raise ValueError("🛑 `nbasis` is missing. Cannot proceed without the number of basis functions.")

    final_geom = data.atomcoords[-1] if hasattr(data, "atomcoords") else []
    atomic_numbers = getattr(data, "atomnos", [])
    pt = PeriodicTable()
    optimized_geometry = [
        {
            "Z": int(Z),
            "symbol": pt.element[Z],
            "x": float(x),
            "y": float(y),
            "z": float(z)
        }
        for Z, (x, y, z) in zip(atomic_numbers, final_geom)
    ]

    MO_coeffs = data.mocoeffs if hasattr(data, "mocoeffs") else None

    if hasattr(data, "moenergies"):
        MO_energies = [e / 27.21139 for e in data.moenergies]
    else:
        MO_energies = None

    MO_occ = None
    homo_index = None
    DFT_type = None

    if hasattr(data, "homos") and MO_energies:
        homos = data.homos
        n = len(data.moenergies[0])
        if np.ndim(homos) == 1 and len(homos) == 2:
            DFT_type = "UDFT"
            MO_occ = []
            for homo in homos:
                occ = ["O" if i <= homo else "V" for i in range(n)]
                MO_occ.append(occ)
            homo_index = list(homos)
        else:
            DFT_type = "DFT"
            homo = int(homos.item())
            MO_occ = ["O" if i <= homo else "V" for i in range(n)]
            homo_index = homo

    S2_val = None
    S_val = None
    multiplicity = None

    try:
        with open(logfile_path, "r") as f:
            lines = f.readlines()
        for line in reversed(lines):
            if "<S**2>" in line and "S=" in line:
                match = re.search(r"<S\*\*2>=\s*([\d.]+)\s+S=\s*([\d.]+)", line)
                if match:
                    S2_val = float(match.group(1))
                    S_val = float(match.group(2))
                    multiplicity = round(2 * S_val + 1, 1)
                break
    except Exception:
        pass

    overlap_matrix = getattr(data, "aooverlaps", None)
    if overlap_matrix is None:
        print("⚠️ WARNING: AO overlap matrix not found. Using identity matrix instead.")
        print()
        overlap_matrix = np.identity(nbasis)

    return {
        "UDFT_or_DFT": DFT_type,
        "nbasis": nbasis,
        "n_MO": len(MO_energies[0]) if MO_energies else None,
        "MO_coeffs": MO_coeffs,
        "MO_energies": MO_energies,
        "MO_occ": MO_occ,
        "HOMO_index": homo_index,
        "optimized_geometry": optimized_geometry,
        "overlap_matrix": overlap_matrix,
        "spin": {
            "S2": S2_val,
            "S": S_val,
            "multiplicity": multiplicity
        }
    }

def load_mos_from_cclib(logfolder, filename):
    """
    Loads molecular orbital data from Gaussian output using cclib and organizes them into DataFrames.

    Parameters
    ----------
    logfolder : str or Path
        Directory containing the log file.
    filename : str
        Name of the Gaussian .log file.

    Returns
    -------
    tuple
        Alpha and beta DataFrames, coefficient matrices, basis count, overlap matrix, and full info dictionary.
    """
    from pathlib import Path
    logfile_path = Path(logfolder) / filename
    info = extract_gaussian_info(logfile_path)
    t4p.centerTitle("🚨 ENTERING load_mos_from_cclib 🚨")

    coeffs = info["MO_coeffs"]
    energies = info["MO_energies"]
    occupations = info["MO_occ"]
    nbasis = info["nbasis"]
    overlap_matrix = info["overlap_matrix"]

    if info["UDFT_or_DFT"] == "UDFT":
        print("UDFT")
        alpha_df = pd.DataFrame({
            "Index": np.arange(1, len(occupations[0]) + 1),
            "Occupation": occupations[0],
            "Energy (Ha)": energies[0]
        })
        beta_df = pd.DataFrame({
            "Index": np.arange(1, len(occupations[1]) + 1),
            "Occupation": occupations[1],
            "Energy (Ha)": energies[1]
        })
        print("✅ Finished load_mos_from_cclib")
        return alpha_df, beta_df, coeffs[0], coeffs[1], nbasis, overlap_matrix, info
    else:
        print("DFT")
        alpha_df = pd.DataFrame({
            "Index": np.arange(1, len(occupations) + 1),
            "Occupation": occupations,
            "Energy (Ha)": energies[0]
        })
        print("✅ Finished load_mos_from_cclib")
        return alpha_df, alpha_df, coeffs[0], coeffs[0], nbasis, overlap_matrix, info

def scalar_product_with_overlap(ci, cj, S):
    """
    Computes the scalar product between two coefficient vectors using an overlap matrix.

    Parameters
    ----------
    ci : np.ndarray
        Coefficient vector i.
    cj : np.ndarray
        Coefficient vector j.
    S : np.ndarray
        Overlap matrix.

    Returns
    -------
    float
        Scalar product ci^T S cj.
    """
    return np.dot(ci.T, S @ cj)



# ### Cosine similarity
# 
# #### Main calculations

# In[3]:


def cosine_similarity_with_overlap(ci, cj, S):
    """
    Computes the cosine similarity between two coefficient vectors using an overlap matrix.

    Parameters
    ----------
    ci : np.ndarray
        Coefficient vector i.
    cj : np.ndarray
        Coefficient vector j.
    S : np.ndarray
        Overlap matrix.

    Returns
    -------
    float
        Cosine similarity between ci and cj.
    """
    num = scalar_product_with_overlap(ci, cj, S)
    norm_i = np.sqrt(scalar_product_with_overlap(ci, ci, S))
    norm_j = np.sqrt(scalar_product_with_overlap(cj, cj, S))
    return num / (norm_i * norm_j)

def interactive_similarity(alpha_df, beta_df, alpha_mat, beta_mat, overlap_matrix):
    """
    Interactive widget to compute and display scalar product and cosine similarity between
    selected alpha and beta MOs using the overlap matrix.

    Parameters
    ----------
    alpha_df : pd.DataFrame
        DataFrame for alpha orbitals.
    beta_df : pd.DataFrame
        DataFrame for beta orbitals.
    alpha_mat : np.ndarray
        Coefficient matrix for alpha orbitals.
    beta_mat : np.ndarray
        Coefficient matrix for beta orbitals.
    overlap_matrix : np.ndarray
        Overlap matrix.
    """
    alpha_opts = {f"alpha #{row['Index']}": row['Index'] - 1 for _, row in alpha_df.iterrows()}
    beta_opts = {f"beta #{row['Index']}": row['Index'] - 1 for _, row in beta_df.iterrows()}

    def compute_similarity(alpha_idx, beta_idx):
        a = alpha_mat[alpha_idx,:]
        b = beta_mat[beta_idx,:]

        dot_product = np.dot(a, b)
        sp_with_overlap = a.T @ (overlap_matrix @ b)
        norm_a = np.sqrt(np.dot(a.T, overlap_matrix @ a))
        norm_b = np.sqrt(np.dot(b.T, overlap_matrix @ b))
        cos_sim_overlap = sp_with_overlap / (norm_a * norm_b)

        print(f"Norm of a                             = {norm_a:.2f}")
        print(f"Norm of b                             = {norm_b:.2f}")
        print(f"Simple dot product (a · b)           = {dot_product:.2f}")
        print(f"Dot product with S (aᵀ·S·b)          = {sp_with_overlap:.2f}")
        print(f"Cosine similarity with S (normalized) = {cos_sim_overlap:.2f}")

    interact(compute_similarity,
             alpha_idx=Dropdown(options=alpha_opts, description="Alpha OM"),
             beta_idx=Dropdown(options=beta_opts, description="Beta OM"))

def find_somo_candidates(alpha_df, beta_df, alpha_mat, beta_mat, nbasis, overlap_matrix, spin, threshold=0.99):
    """
    Identifies singly occupied molecular orbital (SOMO) candidates by comparing
    similarities between occupied alpha and all beta orbitals.

    Parameters
    ----------
    alpha_df : pd.DataFrame
        DataFrame for alpha orbitals.
    beta_df : pd.DataFrame
        DataFrame for beta orbitals.
    alpha_mat : np.ndarray
        Alpha orbital coefficients.
    beta_mat : np.ndarray
        Beta orbital coefficients.
    nbasis : int
        Number of basis functions.
    overlap_matrix : np.ndarray
        Overlap matrix.
    spin: dict
        spin["S2"]: eigenvalue of the S2 operator (float)
        spin["S"]: S-value (float)
        spin["multiplicity"] (float, calculated after 2S+1)
    threshold : float
        Maximum allowed similarity for SOMO detection.

    Returns
    -------
    pd.DataFrame
        Table listing SOMO-like orbital pairs and their properties.
    """
    occupied_alpha_idx = [i for i, occ in enumerate(alpha_df['Occupation']) if occ == 'O']
    similarity_matrix = np.zeros((len(occupied_alpha_idx), beta_mat.shape[0]))

    for i, alpha_idx in enumerate(occupied_alpha_idx):
        a = alpha_mat[alpha_idx,:]
        for j in range(beta_mat.shape[0]):
            b = beta_mat[j,:]
            similarity_matrix[i, j] = np.abs(cosine_similarity_with_overlap(a, b, overlap_matrix))

    row_ind, col_ind = linear_sum_assignment(-similarity_matrix)

    somos = []
    for row, col in zip(row_ind, col_ind):
        sim = similarity_matrix[row, col]
        alpha_idx = occupied_alpha_idx[row]
        beta_occ_status = beta_df.iloc[col]['Occupation'] if col < len(beta_df) else '?'
        if beta_occ_status == 'V' and sim < threshold:
            somos.append((alpha_idx + 1, col + 1, col + 1 + nbasis, sim,
                          alpha_df.iloc[alpha_idx]['Energy (Ha)'],'O',
                          beta_df.iloc[col]['Energy (Ha)'], 'V'))
    nMagMOs = len(somos)
    print(f"Eigenvalue of S2 operator = {spin['S2']}") 
    print(f"S-value = {spin['S']}") 
    print(f"Spin multiplicity = {spin['multiplicity']}") 
    expected_nMag = 2 * spin["S"]
    if abs(nMagMOs - expected_nMag) > 0.1:
        print(f"❌ Inconsistency detected:")
        print(f"   - Detected {nMagMOs} magnetic orbitals (SOMOs)")
        print(f"   - But 2×S = {expected_nMag:.2f} → expected ~{round(expected_nMag)}")
        print("   ⚠️ This might indicate incorrect SOMO detection or spin contamination.")
    else:
        print(f"✅ Number of magnetic orbitals ({nMagMOs}) is consistent with spin value = {expected_nMag} (within 0.1 tolerance).")

    return pd.DataFrame(somos, columns=["Alpha MO", "Beta MO", "Beta MO for Jmol", "Similarity",
                                        "Alpha Energy", "Alpha Occ", "Beta Energy", "Beta Occ"])

def cross_match_all(alpha_df, beta_df, alpha_mat, beta_mat, nbasis, overlap_matrix, n_virtual_alpha=0):
    """
    Matches alpha and beta MOs by maximizing similarity and computes their pairwise similarity and energy difference.

    Parameters
    ----------
    alpha_df : pd.DataFrame
        DataFrame for alpha orbitals.
    beta_df : pd.DataFrame
        DataFrame for beta orbitals.
    alpha_mat : np.ndarray
        Alpha orbital coefficients.
    beta_mat : np.ndarray
        Beta orbital coefficients.
    nbasis : int
        Number of basis functions.
    overlap_matrix : np.ndarray
        Overlap matrix.
    n_virtual_alpha : int
        Number of virtual alpha orbitals to include.

    Returns
    -------
    pd.DataFrame
        Table with matching alpha-beta pairs, similarity scores, and energy differences.
    """
    occ_alpha_idx = [i for i, occ in enumerate(alpha_df['Occupation']) if occ == 'O']
    virt_alpha_idx = [i for i, occ in enumerate(alpha_df['Occupation']) if occ == 'V']
    selected_alpha_idx = occ_alpha_idx + virt_alpha_idx[:n_virtual_alpha]

    similarity_matrix = np.zeros((len(selected_alpha_idx), beta_mat.shape[0]))

    for i, alpha_idx in enumerate(selected_alpha_idx):
        a = alpha_mat[alpha_idx,:]
        for j in range(beta_mat.shape[0]):
            b = beta_mat[j,:]
            similarity_matrix[i, j] = np.abs(cosine_similarity_with_overlap(a, b, overlap_matrix))

    row_ind, col_ind = linear_sum_assignment(-similarity_matrix)

    matches = []
    for row, col in zip(row_ind, col_ind):
        alpha_idx = selected_alpha_idx[row]
        occ_alpha = alpha_df.iloc[alpha_idx]['Occupation']
        occ_beta = beta_df.iloc[col]['Occupation']
        e_alpha = alpha_df.iloc[alpha_idx]['Energy (Ha)']
        e_beta = beta_df.iloc[col]['Energy (Ha)']
        delta_e_ev = (e_beta - e_alpha) * 27.2114  # Ha → eV
        matches.append({
            "Alpha MO": alpha_idx + 1,
            "Alpha Energy": e_alpha,
            "Alpha Occ": occ_alpha,
            "Beta MO": col + 1,
            "Beta MO for Jmol": col + 1 + nbasis,
            "Beta Energy": e_beta,
            "Beta Occ": occ_beta,
            "Similarity": similarity_matrix[row, col],
            "SOMO-like": (occ_alpha != occ_beta),
            "ΔE (eV)": f"{delta_e_ev:.2f}",
        })

    return pd.DataFrame(matches)

def cluster_orbitals(MOs, spin="alpha"):
    """
    Performs hierarchical clustering of molecular orbitals based on cosine similarity.

    Parameters
    ----------
    MOs : tuple of np.ndarray
        Tuple containing coefficient matrices for alpha and beta orbitals.
    spin : str
        Spin type to cluster ('alpha' or 'beta').
    """
    if spin == "alpha":
        matrix = MOs[0]
        title = "Clustering of Alpha MOs"
    elif spin == "beta":
        matrix = MOs[1]
        title = "Clustering of Beta MOs"
    else:
        raise SystemExit(f"{spin} spin is unknown. Use keywords 'alpha' or 'beta'.")

    sim_matrix = np.abs(np.dot(matrix.T, matrix))
    sim_matrix /= np.outer(np.linalg.norm(matrix, axis=0), np.linalg.norm(matrix, axis=0))
    dist_matrix = 1 - sim_matrix
    np.fill_diagonal(dist_matrix, 0.0)
    linkage_matrix = linkage(squareform(dist_matrix), method='average')

    sns.clustermap(sim_matrix, row_linkage=linkage_matrix, col_linkage=linkage_matrix, cmap="viridis")
    plt.title(title)
    plt.show()


def analyzeSimilarity(logfolder, logfile):
    """
    Full analysis pipeline to extract, match, and compare alpha and beta molecular orbitals.
    Displays interactive similarity widgets and saves annotated similarity results to Excel.

    Parameters
    ----------
    logfolder : str or Path
        Path to the folder containing the Gaussian log file.
    logfile : str
        Filename of the Gaussian log file.

    Returns
    -------
    tuple
        Alpha/beta DataFrames, coefficient matrices, nbasis, SOMO DataFrame, and overlap matrix.
    """
    alpha_df, beta_df, alpha_mat, beta_mat, nbasis, overlap_matrix, info = load_mos_from_cclib(logfolder, logfile)
    n_alpha_occ = (alpha_df["Occupation"] == "O").sum()
    n_beta_occ = (beta_df["Occupation"] == "O").sum()
    print(f"n_basis = {nbasis}")
    print(f"Occupied alpha MOs: {n_alpha_occ} (1 -> {n_alpha_occ})")
    print(f"Occupied beta MOs : {n_beta_occ} ({nbasis+1} -> {nbasis+n_beta_occ+1})")

    om_df = pd.concat([alpha_df, beta_df], ignore_index=True)
    display(om_df)

    t4p.centerTitle("Interactive similarity")
    interactive_similarity(alpha_df, beta_df, alpha_mat, beta_mat, overlap_matrix)

    t4p.centerTitle("Magnetic MOs")
    df_SOMOs = find_somo_candidates(alpha_df, beta_df, alpha_mat, beta_mat, nbasis, overlap_matrix, info["spin"], threshold=0.9991)
    display(df_SOMOs)

    t4p.centerTitle("Similarity table")
    similarity_df = cross_match_all(alpha_df, beta_df, alpha_mat, beta_mat, nbasis, overlap_matrix, n_virtual_alpha=5)
    pd.set_option('display.max_rows', None)
    display(similarity_df)
    pd.set_option('display.max_rows', 5)

    output_path = Path(logfolder) / f"{Path(logfile).stem}_similarity.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Similarity Table"

    somo_like_mask = (
        ((similarity_df["Alpha Occ"] == "O") & (similarity_df["Beta Occ"] == "V")) |
        ((similarity_df["Alpha Occ"] == "V") & (similarity_df["Beta Occ"] == "O"))
    )

    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    for r_idx, row in enumerate(dataframe_to_rows(similarity_df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            continue
        if somo_like_mask.iloc[r_idx - 2]:
            for cell in ws[r_idx]:
                cell.fill = yellow_fill

    wb.save(output_path)
    print(f"✅ Similarity table saved with SOMO highlights: {output_path}")

    return (alpha_df, beta_df), (alpha_mat, beta_mat), nbasis, df_SOMOs, overlap_matrix

def save_similarity_per_somo_from_df(df_SOMOs, lMOs, cMOs, nbasis, overlap_matrix, logfolder, logfile):
    """
    Saves one Excel sheet per SOMO candidate listing similarities with all beta MOs,
    sorted by decreasing similarity. Best match is highlighted in yellow.

    Parameters
    ----------
    df_SOMOs : pd.DataFrame
        DataFrame with identified SOMO candidates.
    lMOs : tuple of pd.DataFrame
        Alpha and beta orbital DataFrames.
    cMOs : tuple of np.ndarray
        Alpha and beta coefficient matrices.
    nbasis : int
        Number of basis functions.
    overlap_matrix : np.ndarray
        Overlap matrix.
    logfolder : str or Path
        Folder containing the log file.
    logfile : str
        Name of the log file.
    """
    alpha_df = lMOs[0]
    beta_df = lMOs[1]
    alpha_mat = cMOs[0]
    beta_mat = cMOs[1]

    wb = Workbook()
    wb.remove(wb.active)
    output_path = Path(logfolder) / f"{Path(logfile).stem}_similarityOfSOMOs.xlsx"
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    somo_alpha_indices = df_SOMOs["Alpha MO"].unique() - 1

    for alpha_idx in somo_alpha_indices:
        a = alpha_mat[alpha_idx,:]
        similarities = []

        for j in range(beta_mat.shape[0]):
            b = beta_mat[j,:]
            sim = np.abs(cosine_similarity_with_overlap(a, b, overlap_matrix))
            e_alpha = alpha_df.iloc[alpha_idx]['Energy (Ha)']
            e_beta = beta_df.iloc[j]['Energy (Ha)']
            delta_e_ev = (e_beta - e_alpha) * 27.2114

            similarities.append({
                "Alpha MO": alpha_idx + 1,
                "Alpha Energy": e_alpha,
                "Alpha Occ": alpha_df.iloc[alpha_idx]['Occupation'],
                "Beta MO": j + 1,
                "Beta MO for Jmol": j + 1 + nbasis,
                "Beta Energy": e_beta,
                "Beta Occ": beta_df.iloc[j]['Occupation'],
                "Similarity": sim,
                "ΔE (eV)": f"{delta_e_ev:.2f}"
            })

        df = pd.DataFrame(similarities).sort_values(by="Similarity", ascending=False).reset_index(drop=True)
        best_idx = 0

        sheet_name = f"alpha_{alpha_idx + 1}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                continue
            if r_idx - 2 == best_idx:
                for cell in ws[r_idx]:
                    cell.fill = yellow_fill

    wb.save(output_path)
    print(f"✅ Saved SOMO similarity sheets: {output_path}")



# #### heatmaps

# In[4]:


def build_full_similarity_table(lMOs, cMOs, nbasis, overlap_matrix, lumo_plusAlpha=5, lumo_plusBeta=5):
    """
    Builds a similarity matrix between selected alpha and beta MOs and returns optimal matches.

    Parameters
    ----------
    lMOs : tuple of pd.DataFrame
        Alpha and beta orbital DataFrames.
    cMOs : tuple of np.ndarray
        Coefficient matrices for alpha and beta MOs.
    nbasis : int
        Number of basis functions.
    overlap_matrix : np.ndarray
        Overlap matrix.
    lumo_plusAlpha : int
        Number of virtual alpha orbitals to include beyond LUMO.
    lumo_plusBeta : int
        Number of virtual beta orbitals to include beyond LUMO.

    Returns
    -------
    tuple
        DataFrame with matches, similarity matrix, and selected alpha indices.
    """
    alpha_occ_idx = [i for i, occ in enumerate(lMOs[0]['Occupation']) if occ == 'O']
    alpha_virt_idx = [i for i, occ in enumerate(lMOs[0]['Occupation']) if occ == 'V']
    alpha_selected = alpha_occ_idx + alpha_virt_idx[:lumo_plusAlpha + 1]
    beta_selected = alpha_occ_idx + alpha_virt_idx[:lumo_plusBeta + 1]

    similarity_matrix = np.zeros((len(alpha_selected), len(lMOs[1])))

    for i, alpha_idx in enumerate(alpha_selected):
        a = cMOs[0][alpha_idx, :]
        for j in range(cMOs[1].shape[0]):
            b = cMOs[1][j, :]
            sim = np.abs(cosine_similarity_with_overlap(a, b, overlap_matrix))
            similarity_matrix[i, j] = sim

    row_ind, col_ind = linear_sum_assignment(-similarity_matrix)

    data = []
    for row, col in zip(row_ind, col_ind):
        alpha_idx = alpha_selected[row]
        sim = similarity_matrix[row, col]
        data.append({
            "Alpha MO": alpha_idx + 1,
            "Alpha Occ": lMOs[0].iloc[alpha_idx]["Occupation"],
            "Alpha Energy": lMOs[0].iloc[alpha_idx]["Energy (Ha)"],
            "Beta MO": col + 1,
            "Beta MO for Jmol": col + 1 + nbasis,
            "Beta Occ": lMOs[1].iloc[col]["Occupation"],
            "Beta Energy": lMOs[1].iloc[col]["Energy (Ha)"],
            "Similarity": sim
        })

    df = pd.DataFrame(data)

    matched_beta = set(col_ind)
    unmatched_beta = [j for j in range(len(lMOs[1])) if j not in matched_beta and lMOs[1].iloc[j]['Occupation'] == 'O']
    for j in unmatched_beta:
        b = cMOs[1][:, j]
        sims = [np.abs(np.dot(cMOs[0][i, :], b)) / (np.linalg.norm(cMOs[0][i, :]) * np.linalg.norm(b))
                for i in alpha_selected]
        best_i = np.argmax(sims)
        data.append({
            "Alpha MO": alpha_selected[best_i] + 1,
            "Alpha Occ": lMOs[0].iloc[alpha_selected[best_i]]["Occupation"],
            "Alpha Energy": lMOs[0].iloc[alpha_selected[best_i]]["Energy (Ha)"],
            "Beta MO": j + 1,
            "Beta Occ": lMOs[1].iloc[j]["Occupation"],
            "Beta Energy": lMOs[1].iloc[j]["Energy (Ha)"],
            "Similarity": sims[best_i]
        })

    df = pd.DataFrame(data)
    return df, similarity_matrix, alpha_selected

def heatmap_MOs(lMOs, cMOs, nbasis, overlap_matrix, logfolder="./logs", logfilename="logfile.log"):
    """
    Interactive cosine similarity heatmap between alpha and beta MOs around the HOMO-LUMO frontier.

    Parameters
    ----------
    lMOs : tuple of pd.DataFrame
        Alpha and beta orbital DataFrames.
    cMOs : tuple of np.ndarray
        Coefficient matrices for alpha and beta orbitals.
    nbasis : int
        Number of basis functions.
    overlap_matrix : np.ndarray
        Overlap matrix.
    logfolder : str
        Directory to save the heatmap PNG.
    logfilename : str
        Filename used as prefix for saving.
    """
    from ipywidgets import interact, IntSlider, Button, HBox, Checkbox, Output
    from pathlib import Path
    import matplotlib.pyplot as plt
    import seaborn as sns
    from IPython.display import display, Markdown

    t4p.centerTitle("Cosine similarity of alpha/beta MOs around HOMO-LUMO frontier")

    fig_container = {"fig": None}
    output_msg = Output()
    output_msg.clear_output()

    def update_heatmap(n_occ=5, n_virt=5,
                       n_beta_occ=0, n_beta_virt=5,
                       show_values=False):
        alpha_occ_idx = [i for i, occ in enumerate(lMOs[0]['Occupation']) if occ == 'O']
        beta_occ_idx = [i for i, occ in enumerate(lMOs[1]['Occupation']) if occ == 'O']
        homo_alpha = max(alpha_occ_idx)
        homo_beta = max(beta_occ_idx)
        lumo_beta = min([i for i, occ in enumerate(lMOs[1]['Occupation']) if occ == 'V'])

        selected_alpha_occ = list(range(max(0, homo_alpha - n_occ), homo_alpha + 1))
        selected_alpha_virt = list(range(homo_alpha + 1, homo_alpha + 1 + n_virt))
        selected_alpha = selected_alpha_occ + selected_alpha_virt

        beta_start = min(selected_alpha)

        selected_beta_occ = list(range(max(0, homo_beta - n_beta_occ), homo_beta + 1))
        selected_beta_virt = list(range(lumo_beta, lumo_beta + n_beta_virt))
        selected_beta = sorted(set(selected_beta_occ + selected_beta_virt + list(range(beta_start, homo_beta + 1))))

        filtered_matrix = np.zeros((len(selected_alpha), len(selected_beta)))
        for i, ai in enumerate(selected_alpha):
            for j, bj in enumerate(selected_beta):
                a = cMOs[0][ai, :]
                b = cMOs[1][bj, :]
                filtered_matrix[i, j] = np.abs(cosine_similarity_with_overlap(a, b, overlap_matrix))

        y_labels = [f"α {i+1}" for i in selected_alpha]
        x_labels = [f"β {i+1}" for i in selected_beta]

        n_occ_alpha_in_plot = sum(1 for i in selected_alpha if lMOs[0].iloc[i]['Occupation'] == 'O')
        n_occ_beta_in_plot = sum(1 for i in selected_beta if lMOs[1].iloc[i]['Occupation'] == 'O')

        fig, ax = plt.subplots(figsize=(10, 10))
        sns.heatmap(filtered_matrix,
                    xticklabels=x_labels,
                    yticklabels=y_labels,
                    cmap="viridis",
                    annot=show_values,
                    fmt=".2f" if show_values else "",
                    ax=ax)

        ax.invert_yaxis()
        ax.axhline(n_occ_alpha_in_plot, color="red", linestyle="--", lw=1.5)
        ax.axvline(n_occ_beta_in_plot, color="red", linestyle="--", lw=1.5)

        # ax.set_title("Cosine similarity of alpha/beta MOs around HOMO-LUMO frontier")
        ax.set_xlabel("Beta MOs")
        ax.set_ylabel("Alpha MOs")
        fig.tight_layout()
        fig_container["fig"] = fig
        plt.show()

    def save_heatmap(_):
        fig = fig_container.get("fig")
        if fig is not None:
            filename_prefix = Path(logfilename).stem
            save_path = Path(logfolder) / f"{filename_prefix}_heatmap.png"
            fig.savefig(save_path, dpi=300, transparent=True)
            with output_msg:
                output_msg.clear_output()
                display(Markdown(f"✅ **Image saved as `{save_path}`**"))
        else:
            with output_msg:
                output_msg.clear_output()
                display(Markdown("❌ **No figure to save.**"))

    save_button = Button(description="💾 Save heatmap", tooltip=f"Save heatmap to PNG in {logfolder}")
    save_button.on_click(save_heatmap)
    display(HBox([save_button]))
    display(output_msg)

    slider_alpha_occ = IntSlider(value=5, min=1, max=30, step=1,
                                 description="HOMO–n > HOMO", continuous_update=False)
    slider_alpha_virt = IntSlider(value=5, min=1, max=30, step=1,
                                  description="LUMO > LUMO+n", continuous_update=False)
    slider_beta_occ = IntSlider(value=0, min=0, max=30, step=1,
                                description="β HOMO–n > HOMO", continuous_update=False)
    slider_beta_virt = IntSlider(value=5, min=1, max=30, step=1,
                                 description="β LUMO > LUMO+n", continuous_update=False)
    show_values_checkbox = Checkbox(value=False, description="Show values", indent=False)

    interact(
        update_heatmap,
        n_occ=slider_alpha_occ,
        n_virt=slider_alpha_virt,
        n_beta_occ=slider_beta_occ,
        n_beta_virt=slider_beta_virt,
        show_values=show_values_checkbox
    )


# #### tSNE

# In[5]:


def tsne(lMOs, cMOs, overlap_matrix, logfolder="./logs", logfilename="logfile.log"):
    """
    Performs a t-SNE projection of molecular orbitals (alpha and beta) using a cosine similarity
    metric invariant to phase, and displays an interactive Plotly visualization.

    Parameters
    ----------
    lMOs : tuple of pd.DataFrame
        DataFrames for alpha and beta molecular orbitals.
    cMOs : tuple of np.ndarray
        Coefficient matrices for alpha and beta orbitals.
    overlap_matrix : np.ndarray
        Overlap matrix used for computing cosine similarity.
    logfolder : str
        Path to the folder where the plot image will be saved.
    logfilename : str
        Name of the Gaussian log file used to prefix saved plots.
    """

    import time
    from sklearn.manifold import TSNE
    from sklearn.metrics.pairwise import cosine_similarity
    from ipywidgets import interactive_output, HBox, VBox, Output, SelectMultiple, FloatSlider, Checkbox, Button
    from pathlib import Path
    from IPython.display import display, Markdown
    import plotly.express as px
    import plotly.io as pio
    import numpy as np
    import pandas as pd

    alpha_mat = cMOs[0]
    beta_mat = cMOs[1]
    alpha_df = lMOs[0]
    beta_df = lMOs[1]

    fig_container = {"fig": None}
    output_msg = Output()
    output_msg.clear_output()
    pio.renderers.default = "notebook_connected"

    def run_tsne_phase_invariant(alpha_mat, beta_mat, alpha_df, beta_df, overlap_matrix, perplexity=30, max_iter=2000):
        vectors = np.vstack([alpha_mat, beta_mat])
        metadata = pd.concat([alpha_df, beta_df], ignore_index=True).copy()
        metadata["Type"] = ["Alpha"] * len(alpha_df) + ["Beta"] * len(beta_df)

        n = vectors.shape[0]
        cos_sim = np.zeros((n, n))
        for i in range(n):
            for j in range(n):
                sim = cosine_similarity_with_overlap(vectors[i], vectors[j], overlap_matrix)
                cos_sim[i, j] = np.abs(sim)

        distance_matrix = 1 - cos_sim
        distance_matrix = np.clip(distance_matrix, 0, None)

        print("Running t-SNE... please wait.")
        start = time.time()
        tsne = TSNE(
            n_components=2,
            metric='precomputed',
            perplexity=perplexity,
            max_iter=max_iter,
            init='random',
            learning_rate='auto'
        )
        projection = tsne.fit_transform(distance_matrix)
        end = time.time()
        print(f"t-SNE completed in {end - start:.2f} seconds.")
        metadata["x"], metadata["y"] = projection[:, 0], projection[:, 1]
        return metadata

    def get_homo_energy(df):
        energy_col = "Energy (Ha)" if "Energy (Ha)" in df.columns else "Énergie"
        occ_energies = df[df["Occupation"] == "O"][energy_col]
        return occ_energies.max()

    def plotTSNE(tsne_input_df, logfolder, logfilename):
        type_selector = SelectMultiple(options=["Alpha", "Beta"], value=["Alpha", "Beta"], description="Type")
        occ_selector = SelectMultiple(options=["O", "V"], value=["O", "V"], description="Occupation")
        alpha_slider = FloatSlider(value=1.0, min=0.0, max=5.0, step=0.1, description="±HOMO α (Ha)")
        beta_slider = FloatSlider(value=1.0, min=0.0, max=5.0, step=0.1, description="±HOMO β (Ha)")
        energy_filter_checkbox = Checkbox(value=True, description="Filter around HOMOs", indent=False)

        def plot_filtered_tsne_extended(selected_types, selected_occs, alpha_window, beta_window, filter_energy):
            if tsne_input_df.empty or "Type" not in tsne_input_df.columns:
                with output_msg:
                    output_msg.clear_output()
                    display(Markdown("⚠️ Nothing to display."))
                return

            homo_a = get_homo_energy(alpha_df)
            homo_b = get_homo_energy(beta_df)

            if filter_energy:
                mask_alpha = (
                    (tsne_input_df["Type"] == "Alpha") &
                    (tsne_input_df["Energy (Ha)"] >= homo_a - alpha_window) &
                    (tsne_input_df["Energy (Ha)"] <= homo_a + alpha_window)
                )
                mask_beta = (
                    (tsne_input_df["Type"] == "Beta") &
                    (tsne_input_df["Energy (Ha)"] >= homo_b - beta_window) &
                    (tsne_input_df["Energy (Ha)"] <= homo_b + beta_window)
                )
                filtered_df = tsne_input_df[mask_alpha | mask_beta]
            else:
                filtered_df = tsne_input_df

            final_df = filtered_df[
                filtered_df["Type"].isin(selected_types) &
                filtered_df["Occupation"].isin(selected_occs)
            ]

            if final_df.empty:
                with output_msg:
                    output_msg.clear_output()
                    display(Markdown("⚠️ No MOs match the current filters."))
                return

            fig = px.scatter(
                final_df,
                x="x", y="y",
                color="Type",
                symbol="Occupation",
                hover_data={
                    "Index": True,
                    "Energy (Ha)": True,
                    "Occupation": True,
                    "Type": True,
                    "x": False,
                    "y": False
                },
                labels={"Index": "MO number"}
            )

            for trace in fig.data:
                om_type = trace.name.split(", ")[0]
                occ = trace.name.split(", ")[1] if ", " in trace.name else ""
                if om_type == "Alpha" and occ == "O":
                    trace.update(marker=dict(symbol="circle", size=14, color="rgba(0,0,0,0)", line=dict(color="royalblue", width=2)))
                elif om_type == "Beta" and occ == "O":
                    trace.update(marker=dict(symbol="square", size=14, color="rgba(0,0,0,0)", line=dict(color="crimson", width=2)))
                if om_type == "Alpha" and occ == "V":
                    trace.update(marker=dict(symbol="circle", size=5, color="royalblue", line=dict(color="royalblue", width=1)))
                elif om_type == "Beta" and occ == "V":
                    trace.update(marker=dict(symbol="square", size=6, color="rgba(0,0,0,0)", line=dict(color="crimson", width=1)))

            fig.update_layout(
                title="Filtered t-SNE Projection Around HOMOs",
                legend_title="MO Type",
                height=900, width=900,
                plot_bgcolor="white",
                paper_bgcolor="white",
                xaxis=dict(showgrid=True, gridcolor="lightgray"),
                yaxis=dict(showgrid=True, gridcolor="lightgray"),
                shapes=[
                    dict(type="line", x0=min(tsne_input_df["x"]), x1=max(tsne_input_df["x"]),
                         y0=0, y1=0, line=dict(color="blue", width=2, dash="dot")),
                    dict(type="line", x0=0, x1=0,
                         y0=min(tsne_input_df["y"]), y1=max(tsne_input_df["y"]),
                         line=dict(color="blue", width=2, dash="dot"))
                ]
            )

            fig_container["fig"] = fig
            fig.show()

        def save_tsne(_):
            fig = fig_container.get("fig")
            output_msg.clear_output()
            if fig:
                filename_prefix = Path(logfilename).stem
                save_path = Path(logfolder) / f"{filename_prefix}_tSNE.png"
                fig.write_image(str(save_path), scale=3)
                with output_msg:
                    display(Markdown(f"✅ Image saved to `{save_path}`"))
            else:
                with output_msg:
                    display(Markdown("❌ No figure to save."))

        save_button = Button(description="💾 Save tSNE plot", tooltip=f"Save tSNE plot to PNG in {logfolder}")
        save_button.on_click(save_tsne)

        controls = VBox([
            HBox([type_selector, occ_selector]),
            HBox([alpha_slider, beta_slider, energy_filter_checkbox]),
            HBox([save_button]),
            output_msg
        ])
        out = interactive_output(
            plot_filtered_tsne_extended,
            {
                "selected_types": type_selector,
                "selected_occs": occ_selector,
                "alpha_window": alpha_slider,
                "beta_window": beta_slider,
                "filter_energy": energy_filter_checkbox,
            }
        )
        display(controls, out)

    tsne_input_df = run_tsne_phase_invariant(alpha_mat, beta_mat, alpha_df, beta_df, overlap_matrix)
    plotTSNE(tsne_input_df, logfolder, logfilename)


# ### Projection
# 
# #### Main projection scheme

# In[6]:


def project_occupied_alpha_onto_beta(logfolder, logfile, threshold_beta=20):
    """
    Projects each occupied alpha orbital onto the full set of beta orbitals (occupied + virtual)
    using the AO overlap matrix. Returns a summary DataFrame including projection norms,
    dominant beta contributions, and diagnostic flags.

    Parameters
    ----------
    logfolder : str
        Path to the folder containing the Gaussian log file.
    logfile : str
        Name of the Gaussian log file.
    threshold_beta : float, optional
        Percentage threshold (default: 20) above which a beta orbital is considered significant in the projection.

    Returns
    -------
    pd.DataFrame
        DataFrame with one row per occupied alpha orbital and the following columns:
        - 'Alpha OM': Index (1-based) of the alpha orbital
        - 'Occ α': Occupation of the alpha orbital (usually 'O')
        - 'Energy (Ha)': Energy of the alpha orbital
        - 'Projection² on β_virtual': Squared norm of the projection onto the virtual beta space
        - 'Projection² on β_occupied': Squared norm of the projection onto the occupied beta space
        - 'Dominant β MO': Index (1-based) of the beta orbital with the largest projection
        - 'Index4Jmol': Jmol-compatible index for the dominant beta orbital
        - 'Occ β': Occupation of the dominant beta orbital ('V' or 'O')
        - 'E (β, Ha)': Energy of the dominant beta orbital
        - 'Top 1 contrib (%)': Percentage of the total projection norm carried by the most contributing beta orbital
        - 'Top 2 contrib (%)': Cumulative contribution of the top 2 beta orbitals
        - 'Top 3 contrib (%)': Cumulative contribution of the top 3 beta orbitals
        - 'Dominance ratio': Largest single contribution / total projection
        - 'Spread?': Flag indicating whether the projection is distributed ("Yes" if <60% dominance)
        - 'β orbitals >{threshold_beta}%': List of tuples [OM index (1-based), contribution (%)] for beta orbitals contributing >{threshold_beta value}%
        - 'SOMO?': Yes if projection is dominant onto virtual space and small on occupied

    Notes
    -----
    The squared projection of an occupied alpha orbital \( \phi^\alpha_i \) onto the full beta space is computed as:

    \[
    \mathbf{v}_i = \phi^\alpha_i \cdot S \cdot (\phi^\beta)^T
    \]

    where \( S \) is the AO overlap matrix, and \( \phi^\beta \) is the matrix of beta MOs. The squared norm \( \|\mathbf{v}_i\|^2 \) represents the total overlap.

    Top-N contributions are computed by squaring the individual projections \( v_{ij} \), sorting them, and evaluating the cumulative contributions from the top 1, 2, or 3 beta orbitals. These are returned as "Top 1 contrib (%)", "Top 2 contrib (%)", and "Top 3 contrib (%)".

    The column "β orbitals >{threshold_beta}%" lists all beta orbitals contributing more than the specified percentage to the squared projection norm, with both their index (1-based) and contribution in percent.

    The flag "SOMO?" is set to "Yes" if the squared projection on the virtual beta subspace is greater than 0.5, and the projection on the occupied beta subspace is below 0.5.

    The total number of beta orbitals \( N \) used in the projection is equal to the total number of molecular orbitals in the beta spin channel. The projection is performed over the complete beta space, regardless of occupation.
    """

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    from pathlib import Path
    
    def save_projection_results_to_excel(df_sorted, logfolder, logfile):
        """
        Saves the sorted DataFrame of alpha → beta projections to an Excel file,
        highlighting the SOMO lines in light yellow.
    
        Parameters
        ----------
        df_sorted : pd.DataFrame
            DataFrame already sorted with SOMO rows first.
        logfolder : str
            Directory where the Excel file will be saved.
        logfile : str
            Name of the Gaussian log file used to generate the projection data.
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        from pathlib import Path
    
        # Convert 'β orbitals >X%' column (last dynamic key) into a string
        beta_colname = [col for col in df_sorted.columns if col.startswith("β orbitals >")][0]
        df_sorted[beta_colname] = df_sorted[beta_colname].apply(
            lambda lst: ", ".join([f"{idx}: {contrib:.1f}%" for idx, contrib in lst])
        )
    
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Alpha→Beta Projections"
    
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
        for r_idx, row in enumerate(dataframe_to_rows(df_sorted, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                continue  # skip header
            if r_idx > 1 and row[-1] == "Yes":  # la colonne "SOMO?" est en dernière position
                for cell in ws[r_idx]:
                    cell.fill = yellow_fill
    
        output_path = Path(logfolder) / f"{Path(logfile).stem}_projection_sorted.xlsx"
        wb.save(output_path)
        print(f"✅ Saved to: {output_path}")


    alpha_df, beta_df, alpha_mat, beta_mat, nBasis, overlap_matrix, info = load_mos_from_cclib(logfolder, logfile)

    t4p.centerTitle("Computes the squared projection of each occupied alpha orbital onto the subspaces spanned by both virtual and occupied beta orbitals")

    # alpha_occ_idx = [i for i, occ in enumerate(alpha_df["Occupation"]) if occ == 'O']
    alpha_occ_idx = list(range(len(alpha_df)))

    alpha_occ_mat = alpha_mat[alpha_occ_idx, :]  # (n_occ_alpha, n_basis)
    beta_mat_all = beta_mat  # (n_beta, n_basis)

    projection_data = []
    for i, a in enumerate(alpha_occ_mat):
        proj_vec_all = a @ overlap_matrix @ beta_mat_all.T
        proj2_all = proj_vec_all**2
        norm2_total = float(np.dot(proj_vec_all, proj_vec_all))

        dominant_idx = int(np.argmax(proj2_all))
        norm2_occ = float(np.dot(a @ overlap_matrix @ beta_mat[[j for j, o in enumerate(beta_df["Occupation"]) if o == 'O']].T,
                                  a @ overlap_matrix @ beta_mat[[j for j, o in enumerate(beta_df["Occupation"]) if o == 'O']].T))
        norm2_virt = float(np.dot(a @ overlap_matrix @ beta_mat[[j for j, o in enumerate(beta_df["Occupation"]) if o == 'V']].T,
                                   a @ overlap_matrix @ beta_mat[[j for j, o in enumerate(beta_df["Occupation"]) if o == 'V']].T))

        sorted_proj2 = np.sort(proj2_all)[::-1]
        top1 = sorted_proj2[:1].sum()
        top2 = sorted_proj2[:2].sum()
        top3 = sorted_proj2[:3].sum()
        dominance_ratio = float(sorted_proj2[0] / norm2_total) if norm2_total > 0 else 0.0
        spread_flag = dominance_ratio < 0.6

        rel_contrib = proj2_all / norm2_total if norm2_total > 0 else np.zeros_like(proj2_all)
        significant_idx = [(j + 1, round(float(val * 100), 1)) for j, val in enumerate(rel_contrib) if val > threshold_beta/100]

        # is_somo = "Yes" if norm2_virt > 0.5 and norm2_occ < 0.5 else "No"
        occ_alpha = alpha_df.iloc[alpha_occ_idx[i]]["Occupation"]
        is_somo = "Yes" if (occ_alpha == "O" and norm2_virt > 0.5 and norm2_occ < 0.5) else "No"

        projection_data.append({
            "Alpha MO": alpha_occ_idx[i] + 1,
            "Occ α": alpha_df.iloc[alpha_occ_idx[i]]["Occupation"],
            "Energy (Ha)": alpha_df.iloc[alpha_occ_idx[i]]["Energy (Ha)"],
            "Projection² on β_virtual": float(f"{norm2_virt:.3f}"),
            "Projection² on β_occupied": float(f"{norm2_occ:.3f}"),
            "Dominant β MO": dominant_idx + 1,
            "Index4Jmol": dominant_idx + 1 + nBasis,
            "Occ β": beta_df.iloc[dominant_idx]["Occupation"],
            "E (β, Ha)": beta_df.iloc[dominant_idx]["Energy (Ha)"],
            "Top 1 contrib (%)": float(f"{100 * top1 / norm2_total:.1f}" if norm2_total > 0 else 0),
            "Top 2 contrib (%)": float(f"{100 * top2 / norm2_total:.1f}" if norm2_total > 0 else 0),
            "Top 3 contrib (%)": float(f"{100 * top3 / norm2_total:.1f}" if norm2_total > 0 else 0),
            "Dominance ratio": round(dominance_ratio, 3),
            "Spread?": "Yes" if spread_flag else "No",
            f"β orbitals >{threshold_beta}%": significant_idx,
            "SOMO?": is_somo
        })
    df = pd.DataFrame(projection_data)

    def custom_sort_alpha_df(df):
        """
        Trie le DataFrame des projections alpha → beta selon :
        - d'abord les alpha virtuelles (Occ α == "V")
        - ensuite les SOMOs (Occ α == "O" et SOMO? == "Yes")
        - puis les autres alpha occupées
        Chaque bloc est trié en Alpha MO décroissant.
        """
        df_virtuals = df[df["Occ α"] == "V"].copy()
        df_somos = df[(df["Occ α"] == "O") & (df["SOMO?"] == "Yes")].copy()
        df_others = df[(df["Occ α"] == "O") & (df["SOMO?"] != "Yes")].copy()
        
        df_virtuals = df_virtuals.sort_values(by="Alpha MO", ascending=False)
        df_somos = df_somos.sort_values(by="Alpha MO", ascending=False)
        df_others = df_others.sort_values(by="Alpha MO", ascending=False)
        return pd.concat([df_virtuals, df_somos, df_others], ignore_index=True)
   
    df_sorted = custom_sort_alpha_df(df)
    save_projection_results_to_excel(df_sorted, logfolder, logfile)
    return df_sorted, info

def show_alpha_to_homo(df_proj, logfolder, logfile, highlight_somo=True):
    """
    Affiche les lignes du DataFrame df_proj correspondant aux orbitales alpha
    allant de l’α 1 jusqu’à la HOMO, avec surlignage facultatif des SOMOs.

    Paramètres
    ----------
    df_proj : pd.DataFrame
        DataFrame contenant les résultats de projection alpha → beta.
    logfolder : str
        Dossier contenant le fichier log.
    logfile : str
        Nom du fichier log.
    highlight_somo : bool
        Si True, surligne les lignes avec SOMO? == "Yes".

    Retourne
    --------
    pd.DataFrame ou Styler
        Un sous-ensemble stylisé ou brut du DataFrame.
    """
    alpha_df, *_ = load_mos_from_cclib(logfolder, logfile)
    homo_alpha = max(i for i, occ in enumerate(alpha_df["Occupation"]) if occ == "O")
    homo_index = homo_alpha + 1  # indices dans df_proj sont 1-based
    filtered = df_proj[df_proj["Alpha MO"] <= homo_index].copy()

    if not highlight_somo:
        return filtered

    def somo_highlight(row):
        return ['background-color: #ffff99' if row["SOMO?"] == "Yes" else '' for _ in row]

    return filtered.style.apply(somo_highlight, axis=1)


# In[7]:


#=========================================================
def compute_projection_matrix_and_eigenvalues(lMOs, cMOs, nbasis, overlap_matrix):
    """
    Computes the projection matrix P = A A^T where A = alpha · S · beta^T,
    and returns its eigenvalues and eigenvectors.

    Parameters
    ----------
    lMOs : np.ndarray 
    cMOs : np.ndarray
    nbasis : int
        Number of basis functions.
    overlap_matrix : np.ndarray
        AO overlap matrix (shape: n_basis, n_basis).

    Returns
    -------
    eigenvalues : np.ndarray
        Eigenvalues of the projection matrix P.
    eigenvectors : np.ndarray
        Eigenvectors of the projection matrix P.
    P : np.ndarray
        The projection matrix.
    """
    alpha_df = lMOs[0]
    beta_df = lMOs[1]
    alpha_mat = cMOs[0]
    beta_mat = cMOs[1]

    alpha_occ_idx = [i for i, occ in enumerate(alpha_df["Occupation"]) if occ == "O"]
    alpha_occ_mat = alpha_mat[alpha_occ_idx, :]

    # A = <alpha_occ | beta> : (n_alpha_occ × n_beta)
    A = alpha_occ_mat @ overlap_matrix @ beta_mat.T
    
    # P = A A† : (n_alpha_occ × n_alpha_occ)
    P = A @ A.T
    

    # Diagonalize P
    eigenvalues, eigenvectors = np.linalg.eigh(P)  # Use eigh since P is symmetric

    return eigenvalues[::-1], eigenvectors[:, ::-1], P  # Return in descending order

# Simulate a call with dummy values (the actual call should pass real data)
# compute_projection_matrix_and_eigenvalues(alpha_df, beta_df, alpha_mat, beta_mat, nbasis, overlap_matrix)

def compute_projection_matrix_and_eigenvalues(lMOs, cMOs, nbasis, overlap_matrix):
    """
    Computes the projection matrix P = A Aᵀ where A = alpha_occ · S · beta.T,
    and returns its eigenvalues and eigenvectors.

    Parameters
    ----------
    lMOs : tuple
        Tuple containing two DataFrames: (alpha_df, beta_df), each with MO occupations and energies.
    cMOs : tuple
        Tuple of two np.ndarrays: (alpha_mat, beta_mat), each of shape (n_OMs, n_basis).
        MOs are stored in rows (i.e., row i = MO_i).
    nbasis : int
        Number of basis functions.
    overlap_matrix : np.ndarray
        AO overlap matrix (shape: n_basis, n_basis).

    Returns
    -------
    eigenvalues : np.ndarray
        Eigenvalues of the projection matrix P.
    eigenvectors : np.ndarray
        Eigenvectors of the projection matrix P.
    P : np.ndarray
        The projection matrix P = A Aᵀ.
    """
    alpha_df, beta_df = lMOs
    alpha_mat, beta_mat = cMOs

    # Filter only occupied alpha orbitals
    occ_alpha_idx = [i for i, occ in enumerate(alpha_df["Occupation"]) if occ == "O"]
    alpha_occ = alpha_mat[occ_alpha_idx, :]  # (n_occ_alpha, n_basis)

    # Compute A = alpha_occ · S · beta.T
    A = alpha_occ @ overlap_matrix @ beta_mat.T  # (n_occ_alpha, n_beta)

    # Compute P = A Aᵀ
    P = A @ A.T  # (n_occ_alpha, n_occ_alpha)

    # Diagonalize
    eigenvalues, eigenvectors = np.linalg.eigh(P)

    return eigenvalues, eigenvectors, P


import pandas as pd
def compute_orbital_projections(lMOs, cMOs, overlap_matrix):
    """
    Computes how much each alpha orbital is represented in the beta orbital space
    using the AO overlap matrix S.

    Parameters
    ----------
    lMOs : tuple of DataFrames
        Tuple (alpha_df, beta_df), each containing orbital metadata.
    cMOs : tuple of np.ndarray
        Tuple (alpha_mat, beta_mat), each of shape (n_orbs, n_basis), with rows as orbitals.
    overlap_matrix : np.ndarray
        AO overlap matrix, shape (n_basis, n_basis).

    Returns
    -------
    pd.DataFrame
        DataFrame with alpha orbital number, energy, occupation, and squared projection norm.
    """
    alpha_df, beta_df = lMOs
    alpha_mat, beta_mat = cMOs

    projections = []
    for i, a in enumerate(alpha_mat):
        # Project orbital a onto the beta space
        A_i = a @ overlap_matrix @ beta_mat.T  # shape (n_beta,)
        proj_norm2 = np.dot(A_i, A_i)          # scalar: ||Proj_beta(a)||²

        projections.append({
            "Alpha OM": i + 1,
            "Energy (Ha)": alpha_df.iloc[i]["Energy (Ha)"],
            "Occupation": alpha_df.iloc[i]["Occupation"],
            "Projection²": proj_norm2
        })

    return pd.DataFrame(projections)

def print_eigen_analysis(eigenvalues, threshold=0.8):
    """
    Prints and analyzes the eigenvalues of the projection matrix.

    Parameters
    ----------
    eigenvalues : np.ndarray
        Eigenvalues of the projection matrix (real and ≥ 0).
    threshold : float
        Eigenvalues below this threshold are considered "low" (possible SOMO signature).
    """
    print("=== Projection Matrix Eigenvalue Analysis ===")
    print(f"Total eigenvalues: {len(eigenvalues)}")
    print()

    n_high = np.sum(eigenvalues > 0.95)
    n_low = np.sum(eigenvalues < threshold)

    for i, val in enumerate(sorted(eigenvalues, reverse=True)):
        status = ""
        if val > 0.95:
            status = "✅ well projected"
        elif val < threshold:
            status = "⚠️ low projection (possible SOMO)"
        else:
            status = "↔️ intermediate"

        print(f"Eigenvalue {i+1:2d}: {val:.3f} {status}")

    print()
    print(f"🔹 {n_high} strongly projected α-OMs")
    print(f"🔸 {n_low} possibly unpaired α-OMs (SOMO candidates)")

def identify_somos_from_projection(logfolder, logfile):
    """
    Identifies potential SOMOs by projecting occupied alpha orbitals onto the beta orbital space.

    Parameters
    ----------
    logfolder : str
        Path to the folder containing the Gaussian log file.
    logfile : str
        Name of the Gaussian .log file.

    This function:
    - Loads orbital data from the log file.
    - Computes the projection matrix P = A Aᵀ where A = α_occ · S · βᵀ.
    - Diagonalizes P and plots its eigenvalues.
    - Flags alpha orbitals with eigenvalues > 0.5 that project mainly onto virtual beta orbitals.
    """
    import matplotlib.pyplot as plt

    alpha_df, beta_df, alpha_mat, beta_mat, nBasis, overlap_matrix, info = load_mos_from_cclib(logfolder, logfile)
    n_alpha_occ = (alpha_df["Occupation"] == "O").sum()
    n_beta_occ = (beta_df["Occupation"] == "O").sum()
    alpha_occ_idx = [i for i, occ in enumerate(alpha_df["Occupation"]) if occ == 'O']
    alpha_occ_mat = alpha_mat[alpha_occ_idx, :]
    
    print(f"n_basis = {nBasis}")
    print(f"Occupied alpha MOs: {n_alpha_occ} (1 -> {n_alpha_occ})")
    print(f"Occupied beta MOs : {n_beta_occ} ({nBasis+1} -> {nBasis+n_beta_occ+1})")
    
    listMOs = (alpha_df, beta_df)
    coeffMOs = (alpha_mat, beta_mat)
    
    eigenvalues, eigenvectors, P = compute_projection_matrix_and_eigenvalues(listMOs, coeffMOs, nBasis, overlap_matrix)
    eigenvalues = np.clip(eigenvalues, 0, 1)

    
    plt.figure(figsize=(8, 5))
    plt.plot(eigenvalues, marker='o')
    plt.xlabel("Orbital index")
    plt.ylabel("Projection eigenvalue")
    plt.title("Eigenvalues of α → β projection matrix")
    plt.grid(True)
    plt.show()
    
    print(eigenvalues)
    
    A = alpha_occ_mat @ overlap_matrix @ beta_mat.T
    dominant_beta_index = np.argmax(A**2, axis=1)  # ou abs si non-normalisé
    for i, beta_idx in enumerate(dominant_beta_index):
        if beta_df.iloc[beta_idx]["Occupation"] == 'V' and eigenvalues[i] > 0.5:
            e_alpha = alpha_df.iloc[alpha_occ_idx[i]]["Energy (Ha)"]
            e_beta = beta_df.iloc[beta_idx]["Energy (Ha)"]
            print(f"🧲 OM alpha #{alpha_occ_idx[i]+1} (E={e_alpha:.3f} Ha) may be a SOMO — projects onto virtual beta #{beta_idx+1} (E={e_beta:.3f} Ha)")


# #### Heatmap

# In[8]:


from pathlib import Path

def parse_beta_contrib_string(s):
    if not isinstance(s, str) or not s.strip():
        return []
    parts = s.split(",")
    result = []
    for p in parts:
        try:
            idx, contrib = p.strip().split(":")
            idx = int(idx.strip())
            contrib = float(contrib.strip().replace("%", ""))
            result.append((idx, contrib))
        except Exception:
            continue
    return result

def projection_heatmap_from_df(df, nbasis, logfolder="./logs", logfilename="logfile.log"):
    from ipywidgets import interact, interactive_output, IntSlider, Button, HBox, Checkbox, Output, VBox
    from pathlib import Path
    import matplotlib.pyplot as plt
    import seaborn as sns
    from IPython.display import display, Markdown
    from functools import partial

    alpha_df, beta_df, alpha_mat, beta_mat, nBasis, overlap_matrix, info = load_mos_from_cclib(logfolder, logfile)
    lMOs = (alpha_df, beta_df)

    del alpha_mat, beta_mat, overlap_matrix, info, nBasis
    n_alpha_occ = (alpha_df["Occupation"] == "O").sum()
    n_beta_occ = (beta_df["Occupation"] == "O").sum()
    
    print(f"n_basis = {nbasis}")
    print(f"Occupied alpha MOs: {n_alpha_occ} (1 -> {n_alpha_occ})")
    print(f"Occupied beta MOs : {n_beta_occ} ({nbasis+1} -> {nbasis+n_beta_occ+1})")
    
    t4p.centerTitle("Main projection contribution of Alpha MOs on Beta MOs")

    fig_container = {"fig": None}
    output_msg = Output()
    output_msg.clear_output()

    def update_heatmap(n_occ, n_virt,n_beta_occ, n_beta_virt,
                       show_values,
                      ):
        alpha_indices = []
        beta_indices = set()
        contrib_dict = {}
        
        alpha_occ_idx = [i for i, occ in enumerate(lMOs[0]['Occupation']) if occ == 'O']
        beta_occ_idx = [i for i, occ in enumerate(lMOs[1]['Occupation']) if occ == 'O']
        homo_alpha = max(alpha_occ_idx)
        homo_beta = max(beta_occ_idx)
        lumo_beta = min([i for i, occ in enumerate(lMOs[1]['Occupation']) if occ == 'V'])

        selected_alpha_occ = list(range(max(0, homo_alpha - n_occ), homo_alpha + 1))
        selected_alpha_virt = list(range(homo_alpha + 1, homo_alpha + 1 + n_virt))
        selected_alpha = selected_alpha_occ + selected_alpha_virt

        beta_start = min(selected_alpha)

        selected_beta_occ = list(range(max(0, homo_beta - n_beta_occ), homo_beta + 1))
        selected_beta_virt = list(range(lumo_beta, lumo_beta + n_beta_virt))
        selected_beta = sorted(set(selected_beta_occ + selected_beta_virt + list(range(beta_start, homo_beta + 1))))

        for _, row in df.iterrows():
            alpha_idx = int(row["Alpha MO"])
            alpha_indices.append(alpha_idx)
            contribs = parse_beta_contrib_string(row[next(c for c in row.index if c.startswith("β orbitals >"))])
            contrib_dict[alpha_idx] = {b: w for b, w in contribs}
            beta_indices.update([b for b, _ in contribs])

        alpha_indices = sorted(alpha_indices)
        beta_indices = sorted(beta_indices)

        matrix = np.zeros((len(selected_alpha), len(selected_beta)))

        for i, a in enumerate(selected_alpha):
            for j, b in enumerate(selected_beta):
                matrix[i, j] = contrib_dict.get(a+1, {}).get(b+1, 0.0)/100

        y_labels = [f"α {i+1}" for i in selected_alpha]
        x_labels = [f"β {i+1}" for i in selected_beta]

        n_occ_alpha_in_plot = sum(1 for i in selected_alpha if lMOs[0].iloc[i]['Occupation'] == 'O')
        n_occ_beta_in_plot = sum(1 for i in selected_beta if lMOs[1].iloc[i]['Occupation'] == 'O')
        
        fig, ax = plt.subplots(figsize=(10, 10))
        sns.heatmap(matrix,
                    xticklabels=x_labels,
                    yticklabels=y_labels,
                    cmap="viridis",
                    annot=show_values,
                    fmt=".2f" if show_values else "",
                    ax=ax)
        ax.invert_yaxis()
        ax.axhline(n_occ_alpha_in_plot, color="red", linestyle="--", lw=1.5)
        ax.axvline(n_occ_beta_in_plot, color="red", linestyle="--", lw=1.5)

        ax.set_xlabel("Beta MOs")
        ax.set_ylabel("Alpha MOs")

        fig.tight_layout()
        fig_container["fig"] = fig
        plt.show()

    def save_heatmap(_):
        fig = fig_container.get("fig")
        if fig is not None:
            filename_prefix = Path(logfilename).stem
            save_path = Path(logfolder) / f"{filename_prefix}_projection_heatmap.png"
            fig.savefig(save_path, dpi=300, transparent=True)
            with output_msg:
                output_msg.clear_output()
                display(Markdown(f"✅ **Image saved as `{save_path}`**"))
        else:
            with output_msg:
                output_msg.clear_output()
                display(Markdown("❌ **No figure to save.**"))

    save_button = Button(description="💾 Save map", tooltip=f"Save map to PNG in {logfolder}")
    save_button.on_click(save_heatmap)

    display(HBox([save_button]))
    display(output_msg)

    slider_alpha_occ = IntSlider(value=5, min=1, max=30, step=1,
                                 description="HOMO–n > HOMO", continuous_update=False)
    slider_alpha_virt = IntSlider(value=5, min=1, max=30, step=1,
                                  description="LUMO > LUMO+n", continuous_update=False)
    slider_beta_occ = IntSlider(value=0, min=0, max=30, step=1,
                                description="β HOMO–n > HOMO", continuous_update=False)
    slider_beta_virt = IntSlider(value=5, min=1, max=30, step=1,
                                 description="β LUMO > LUMO+n", continuous_update=False)
    show_values_checkbox = Checkbox(value=False, description="Show values", indent=False)

    interact(
        update_heatmap,
        n_occ=slider_alpha_occ,
        n_virt=slider_alpha_virt,
        n_beta_occ=slider_beta_occ,
        n_beta_virt=slider_beta_virt,
        show_values=show_values_checkbox
    )    


