import pandas
import os
from openpyxl import load_workbook
import psutil
import win32com.client
import datetime


def insert_data_to_excel(file_path, data, sheet_arg=None):
    """
    Inserts numerical data into specific cells in the Excel file.

    Parameters:
    - file_path: Path to the Excel file.
    - data: Dictionary where keys are cell addresses (e.g., "A1") and values are the numerical data to insert.
    """
    wb = load_workbook(file_path,data_only=False, keep_vba= True, keep_links= True, rich_text=True)
    sheet = wb[sheet_arg]

    for cell, value in data.items():
        sheet[cell] = check_isnumber(value)



    # Save changes
    wb.save(file_path)
    wb.close()
    print(f"Data inserted successfully to: {file_path} into {sheet_arg}")


def reopen_excel_file(file_path):
    """
    Reopens the Excel file using the default application.
    """
    os.startfile(file_path)
    print(f"Reopened the file: {os.path.basename(file_path)}")


def close_excel_file_if_open(file_path):
    """
    Checks if the Excel file is open, and if so, closes it without closing the entire Excel application.
    """
    file_name = os.path.basename(file_path)
    file_closed = False

    for proc in psutil.process_iter(attrs=["pid", "name"]):
        if proc.info["name"].lower() == "excel.exe":
            # Attach to the running Excel application
            excel_app = win32com.client.Dispatch("Excel.Application")
            try:
                # Iterate over open workbooks
                for wb in excel_app.Workbooks:
                    if os.path.basename(wb.FullName).lower() == os.path.basename(
                        file_name.lower()
                    ):
                        wb.Close(SaveChanges=True)  # Close the workbook
                        print(f"Closed the file: {file_name}")
                        file_closed = True
                        break
            except Exception as e:
                print(f"An error occurred: {e}")
            finally:
                # Release COM object
                del excel_app
            break

    if not file_closed:
        print(f"The file {file_name} was not open.")

    return file_closed


def prepare_source(
    date="",
    postfix=None,
    base_path=None,
    sep=None,
    columns_to_drop=None,
    white_list=False,
    columns_to_keep=None,
    infer=True,
    header_start=0,
    range_of_interest_start=0,
    range_of_interest_end=-1,
):
    prefix = f"{date}"
    postfix = postfix
    base_path = base_path
    if infer:
        file = pandas.read_csv(base_path, sep=sep, on_bad_lines='warn')
    else:
        file = pandas.read_csv(base_path, sep=sep, header=header_start, on_bad_lines='warn')

    raw_column_list = list(file.columns)
    target_list = []
    if white_list:
        for i in range(len(columns_to_keep)):
            target_list.append(raw_column_list.pop(int(columns_to_keep[i])))
            print(f"target list: {target_list}")
    if not white_list:
        clean_file = file.drop(columns=columns_to_drop)
    elif white_list:
        clean_file = file.filter(items=target_list)
        return clean_file.iloc[range_of_interest_start: range_of_interest_end]
    
    return clean_file


def prepare_columns(column_name="", source=None, numeric = False, column_number= None, is_series = False, already_list= False):
    values = []
    
    if already_list is True and numeric is False and is_series is False:
        return source

    if numeric is False and is_series is True and already_list is False:
        values = source.tolist()

    if numeric is False and is_series is False and already_list is False:
        values = source[column_name].values.tolist()
    
    if numeric is True and is_series is False and already_list is False:
        values = source.iloc[:, column_number].values.tolist()
    
    return values


def build_dict(keys=None, values=None, header=None, include_header= False):
    column_dict = {}
    if include_header:
        values.insert(0, header)

    key_len = len(keys) - 1
    offset_range = range(key_len)[-1]
    
    if include_header:
        offset_with_header = offset_range+1

    else:
        offset_with_header = offset_range

    for i in range(0, offset_with_header):
        column_dict[keys[i]] = values[i]
    return column_dict


def prepare_cells(prefix="A", leng=-1, start=1, skip_after_four= False):
    cells = []
    offset = 0
    cell_counter= 0
    start = start + 1
    leng = leng + 1
    offset = leng + 1 + start
   

    for n in range(start, offset):

        if skip_after_four is True:
            
            cell_counter= cell_counter + 1
            if cell_counter < 5:
                
                cell = prefix + str(n)

            if cell_counter == 5:
                cell_counter= 0
                n= n + 1
                cell= prefix + str(n) 

        else:
            cell = prefix + str(n)
        
        cells.append(cell)
    
    return cells


def check_isnumber(value=None):

    if isinstance(value, (int, float)):
        return float(value)
    
    if isinstance(value, str):
        try:
            return float(value.replace(",", "."))
        except ValueError:
            return value
    
    return value


def set_source(
    file_path=None,
    prefix_letter=None,
    skip_after_four= False,
    column_target=None,
    date_target=None,
    start_int=1,
    header=None,
    postfix=None,
    base_path=None,
    sep=None,
    columns_to_drop=None,
    white_list=False,
    columns_to_keep=None,
    infer=True,
    header_start=0,
    range_of_interest_start=0,
    range_of_interest_end=-1,
    include_header= False
):
    raw_path = r"{}".format(file_path)
    file_path = raw_path
    clean_file = prepare_source(
        date=date_target,
        postfix=postfix,
        base_path=base_path,
        sep=sep,
        columns_to_drop=columns_to_drop,
        white_list= white_list,
        columns_to_keep=columns_to_keep,
        infer=infer,
        header_start=header_start,
        range_of_interest_start=range_of_interest_start,
        range_of_interest_end=range_of_interest_end,
    )
    column_to_insert = build_dict(
        keys=prepare_cells(prefix=prefix_letter, leng=len(clean_file), start=start_int, skip_after_four= skip_after_four),
        values=prepare_columns(column_name=column_target, source=clean_file),
        header=header, include_header= include_header
    )
    return {"file_path": file_path, "column_insert": column_to_insert}


def get_date(file_path=None, sheet=None):

    workbook = load_workbook(file_path, data_only=True, keep_vba= True, keep_links= True, rich_text=True)
    sheet = workbook[sheet]
    for row in sheet.iter_rows(
        min_row=1, max_row=1, min_col=1, max_col=1, values_only=True
    ):
        date_object = row[0]
        date_formatted = date_object.strftime("%Y-%m-%d")
        return date_formatted


def get_weekday(date=None):
    try:
        weeekday_num = datetime.datetime.strptime(date, "%Y-%m-%d").weekday()
        if weeekday_num == 0:
            return "Mon"
        elif weeekday_num == 1:
            return "Tue"
        elif weeekday_num == 2:
            return "Wed"
        elif weeekday_num == 3:
            return "Thu"
        elif weeekday_num == 4:
            return "Fri"
        elif weeekday_num == 5:
            return "Sat"
        elif weeekday_num == 6:
            return "Sun"
    except Exception as e:
        print(f"error: {e}")


def parse_range(
    source_path=None,
    sheet=None,
    start_row=None,
    end_row=None,
    start_column=None,
    end_column=None,
):
    workbook = load_workbook(source_path, data_only=True, keep_vba= True, keep_links= True, rich_text=True)
    sheet = workbook[sheet]
    for row in sheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_column,
        max_col=end_column,
        values_only=True,
    ):
        cell_values= []
        for cell in row:
            if cell is not None:
                cell_values.append(check_isnumber(cell))
        return cell_values 


def split_positives(numbers= None):
    positives= []
    for x in numbers:
        if x >= 0:
            positives.append(x)
        else:
            positives.append(0)
    return positives

def split_negatives(numbers= None):
    negatives= []
    for x in numbers:
        if x <= 0:
            negatives.append(x)
        else:
            negatives.append(0)
    return negatives


def split_on_header(header= None, target_header=None, source_list= None):
    if header== target_header:

        positive_list= split_positives(source_list)
        negative_list= split_negatives(source_list)

        return {"positive_list": positive_list, "negative_list": negative_list}
    
    return source_list

def parse_split_lists(positive_sign= None, input= None):
    if positive_sign:
        return input["positive_list"]
    else:
        return input["negative_list"]


def parse_from_excel(base_path, source_sheet, start_row, end_row, start_column, end_column, file_path, header, include_header, prefix_letter, start_int):
    
    values = []
    
    values = parse_range(
    source_path=base_path,
    sheet=source_sheet,
    start_row=start_row,
    end_row=end_row,
    start_column=start_column,
    end_column=end_column
    )
    
    values_len= len(values)
    
    column_to_insert = build_dict(
        keys=prepare_cells(prefix=prefix_letter, leng=values_len, start=start_int, skip_after_four= False),
        values=values,
        header=header, include_header= include_header
    )
    return {"file_path": file_path, "column_insert": column_to_insert}