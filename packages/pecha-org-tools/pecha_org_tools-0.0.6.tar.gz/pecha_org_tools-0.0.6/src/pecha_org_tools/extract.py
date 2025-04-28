import re
from pathlib import Path
from typing import List, Optional, Union

from openpyxl import load_workbook

from pecha_org_tools.config import download_spreedsheet


class CategoryExtractor:
    def __init__(self, input_file: Optional[Path] = None):
        input_file = input_file or Path(download_spreedsheet())
        self.input_file = input_file
        self.bo_categories, self.en_categories = self.process_file()

    @staticmethod
    def read_xlsx_file(file_path: Path):
        """
        Read the xlsx file and return its contents as a list of rows.
        """
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        rows_data = []
        for row in worksheet.iter_rows(values_only=True):
            rows_data.append(row)
        return rows_data

    def extract_categories(self):
        """
        Process the xlsx file and extract hierarchical categories from its contents.
        """
        self.rows_data = self.read_xlsx_file(self.input_file)
        bo_categories = []
        cur_bo_cat: List[Union[str, None]] = []

        en_categories = []
        cur_en_cat: List[Union[str, None]] = []

        for row in self.rows_data:
            # Find the first non-None value and its index
            flag = False
            for col_index, cell_value in enumerate(row):
                if cell_value is not None:
                    flag = True
                    break
            if not flag:
                continue

            cur_cat_len = len(cur_bo_cat)
            cell_value = cell_value.strip()

            en_cell_value = row[col_index + 1].strip()

            # Update or extend the current category hierarchy
            if cur_cat_len == col_index:
                cur_bo_cat.append(cell_value)
                cur_en_cat.append(en_cell_value)
            else:
                cur_bo_cat[col_index] = cell_value
                cur_en_cat[col_index] = en_cell_value

            # Reset trailing elements to None
            cur_bo_cat[col_index + 1 :] = [None] * (cur_cat_len - col_index - 1)  # noqa
            cur_en_cat[col_index + 1 :] = [None] * (cur_cat_len - col_index - 1)  # noqa

            # Add non-empty elements to the result
            active_category = [
                category for category in cur_bo_cat if category is not None
            ]
            bo_categories.append(active_category)

            active_en_category = [
                category for category in cur_en_cat if category is not None
            ]
            en_categories.append(active_en_category)

        self.bo_extracted_categories = bo_categories
        self.en_extracted_categories = en_categories

    def process_file(self):
        """
        Extract and format categories from the provided xlsx file.
        """
        self.extract_categories()
        bo_categories, en_categories = [], []
        for bo_category_hierarchy, en_category_hierarchy in zip(
            self.bo_extracted_categories, self.en_extracted_categories
        ):
            bo_formatted_category = format_categories(bo_category_hierarchy, "bo")
            en_formatted_category = format_categories(en_category_hierarchy, "en")

            bo_categories.append(bo_formatted_category)
            en_categories.append(en_formatted_category)

        return bo_categories, en_categories

    def get_category(self, category_name: str):
        """
        Get the category hierarchy for a given category name.
        """

        bo_category = self.get_category_by_lang(category_name, lang="bo")
        en_category = self.get_en_category_by_bo(category_name)

        return {"bo": bo_category, "en": en_category}

    def get_en_category_by_bo(
        self,
        category_name: str,
    ):
        """
        Get english category hierarchy by matching with category name in Tibetan.
        """

        matched_idx = None
        for idx, bo_category in enumerate(self.bo_categories):
            if bo_category[-1]["name"] == category_name:
                matched_idx = idx
                break

        matched_category = self.en_categories[matched_idx]
        return matched_category

    def get_category_by_lang(
        self,
        category_name: str,
        lang: str,
    ):
        """
        Get the category hierarchy for a given category name.
        """

        if lang == "bo":
            formatted_categories = self.bo_categories
        elif lang == "en":
            formatted_categories = self.en_categories
        else:
            raise ValueError(f"Unsupported language: {lang}")

        matched_category = None
        for formatted_category in formatted_categories:
            if formatted_category[-1]["name"] == category_name:
                matched_category = formatted_category.copy()
                break

        if not matched_category:
            raise ValueError(f"Category not found for {category_name}")

        return matched_category


def parse_category_text(text: str):
    """
    Extract the main text and any descriptions (in parentheses) from a given string.
    """
    pattern = r"^(.*?)\s*(?:\((.*?)\))?(?:\((.*?)\))?$"
    match = re.search(pattern, text)

    if match:
        name = match.group(1).strip() if match.group(1) else ""
        description = match.group(2).strip() if match.group(2) else ""
        short_description = match.group(3).strip() if match.group(3) else ""
        return name, description, short_description
    else:
        return None, None, None


def format_categories(category_hierarchy: List[str], lang: str):
    """
    Format each category hierarchy into a structured format with main text and descriptions.
    """
    formatted_category = []
    for category in category_hierarchy:
        name, description, short_description = parse_category_text(category)
        if lang == "bo":
            category_data = {
                "name": name,
                "heDesc": description,
                "heShortDesc": short_description,
            }
        else:
            category_data = {
                "name": name,
                "enDesc": description,
                "enShortDesc": short_description,
            }

        formatted_category.append(category_data)

    return formatted_category


if __name__ == "__main__":
    categorizer = CategoryExtractor()
    print(categorizer)
