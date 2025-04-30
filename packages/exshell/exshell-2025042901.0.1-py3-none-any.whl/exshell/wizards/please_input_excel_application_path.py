import openpyxl as xl
import subprocess
import time

from pathlib import Path
from tomlkit import dumps as toml_dumps


class PleaseInputExcelApplicationPath():
    """TOML形式の設定ファイルに、以下の行（例）を追加させるためのウィザード。

    [excel]
    path = "C:\\Program Files\\Microsoft Office\\root\\Office16\\EXCEL.EXE"
    """

    @staticmethod
    def play(exshell_builder):

        print() # 改行

        while True:
            text = f"""\
🙋　Tutorial
-------------
このアプリケーションでは、 Excel アプリケーションを自動的に開いたり閉じたりしたいです。

これに同意できる方は、後述の説明を参考に Excel アプリケーションへのファイルパスを入力してください。
そうでない方は、[Ctrl] + [C] キーで強制終了していただくことができます。

Excel アプリケーションへのファイルパスの調べ方を説明します。

◆ Windows 11 を使っていて、Excel をすでにインストールしている方：
    タスクバーの検索ボックスに `Excel` と入力し、
    出てきた Excel のアイコンを右クリックして［ファイルの場所を開く］をクリックしてください。
    ショートカット・アイコンが出てくるのでさらに右クリックして［ファイルの場所を開く］をクリックしてください。
    📄［EXCEL.EXE］ファイルが出てくるので右クリックして［パスのコピー］をクリックしてください。
    これでクリップボードにファイルパスがコピーされました。
    これをターミナルに貼り付けてください。
    両端にダブルクォーテーションが付いているので、ダブルクォーテーションは削除してください。

◆ それ以外の方
    がんばってください。


    Example of input
    ----------------
    C:\\Program Files\\Microsoft Office\\root\\Office16\\EXCEL.EXE

Input
-----"""

            lines = text.split('\n')

            for line in lines:
                print(line)
                time.sleep(0.05)

            temporary_excel_application_path = input()
            print() # 空行

            # ワークブックを新規生成
            wb = xl.Workbook()

            # ワークシート
            ws = wb['Sheet']

            cell = ws[f'A1']
            cell.value = "ありがとうございます。 Excel ファイルを開けました。"

            cell = ws[f'A2']
            cell.value = "この画面は、プログラムの方から閉じますので、このままにしておいてください。"

            cell = ws[f'A3']
            cell.value = "引き続き、プログラムの指示に従ってください。よろしくお願いします。"


            try:
                print(f'🔧　Save 📄［ {exshell_builder.abs_path_to_workbook} ］workbook file...')
                wb.save(exshell_builder.abs_path_to_workbook)
            
            except Exception as ex:
                print(f"""\
{ex}
""")
                time.sleep(1)

                message = f"""\
🙋　Tutorial
-------------
何らかの理由で 📄［ {exshell_builder.abs_path_to_workbook} ］ファイルの上書きに失敗しました。
もしかして、指定したディレクトリーは無かった、といったことはありませんか？

問題をがんばって取り除いたあとで、
もう一度、最初からやり直してください...

"""
                print(message)
                time.sleep(1)
                continue


            # Excel のファイルパス入力完了


            # エクセルを開く
            print(f'🔧　Open virtual display...')
            opened_excel_process = subprocess.Popen([temporary_excel_application_path, exshell_builder.abs_path_to_workbook])   # Excel が開くことを期待
            time.sleep(1)


            message = f"""\
🙋　Tutorial
-------------
Excel アプリケーションが自動的に開かれた方は `y` を、
そうでない場合は　それ以外を入力してください。

    Example of input
    ----------------
    y

Input
-----
"""
            line = input(message)
            print() # 空行


            if line == 'y':
                # 設定ファイルへ保存
                exshell_builder.config_doc_rw['excel']['path'] = temporary_excel_application_path

                print(f"""\
{exshell_builder.config_doc_rw=}
{exshell_builder.config_doc_rw['excel']['path']=}
""")

                print(f'🔧　Save 📄［ {exshell_builder.abs_path_to_config} ］config file...')
                with open(exshell_builder.abs_path_to_config, mode='w', encoding='utf-8') as f:
                    f.write(toml_dumps(exshell_builder.config_doc_rw))

                # エクセルを閉じる
                print(f'🔧　Close virtual display...')
                opened_excel_process.terminate()
                time.sleep(1)
                break
                
            else:
                message = f"""\
🙋　Tutorial
-------------
もう一度、最初からやり直してください...

"""
                print(message)
                time.sleep(1)

        message = f"""\
🙋　Tutorial
-------------
このアプリケーションと、 Excel アプリケーションの紐付けが完了しました。
引き続き、よろしくお願いします...

"""
        print(message)
        time.sleep(1)
